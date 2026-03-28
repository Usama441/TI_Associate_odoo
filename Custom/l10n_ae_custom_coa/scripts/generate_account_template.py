#!/usr/bin/env python3
"""Generate account.account-ae_custom.csv from Account (account.account).xlsx."""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_HEADERS = (
    'Code',
    'Account Name',
    'Type',
    'Allow Reconciliation',
    'Account Currency',
    'Companies',
)

TYPE_MAPPING = {
    'Receivable': 'asset_receivable',
    'Payable': 'liability_payable',
    'Bank and Cash': 'asset_cash',
    'Current Assets': 'asset_current',
    'Non-current Assets': 'asset_non_current',
    'Prepayments': 'asset_prepayments',
    'Current Liabilities': 'liability_current',
    'Non-current Liabilities': 'liability_non_current',
    'Equity': 'equity',
    'Current Year Earnings': 'equity_unaffected',
    'Income': 'income',
    'Other Income': 'income_other',
    'Expenses': 'expense',
    'Cost of Revenue': 'expense_direct_cost',
}

DEFAULT_TAX_XMLIDS_BY_CODE = {
    '41010102': 'uae_export_tax',  # Revenue - International Clients
}


def normalize_name(name: str) -> str:
    # Normalize mojibake from spreadsheet exports and keep output ASCII-safe.
    normalized = name.replace('â€“', '-').replace('–', '-')
    normalized = re.sub(r'\s+', ' ', normalized).strip()
    return normalized


def as_code(raw_value) -> str:
    code = str(raw_value).strip()
    if isinstance(raw_value, (int, float)):
        code = str(int(raw_value))
    if not code:
        raise ValueError('Missing code')
    if not code.isdigit():
        raise ValueError(f"Code must be numeric: {code}")
    if len(code) != 8:
        raise ValueError(f"Code must be exactly 8 digits: {code}")
    return code


def as_bool(raw_value) -> bool:
    if isinstance(raw_value, bool):
        return raw_value
    return str(raw_value or '').strip().lower() in {'1', 'true', 'yes'}


def parse_rows(input_path: Path) -> list[dict]:
    workbook = load_workbook(input_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    headers = tuple(worksheet.cell(1, col_idx).value for col_idx in range(1, 7))
    if headers != EXPECTED_HEADERS:
        raise ValueError(f'Unexpected headers. Expected {EXPECTED_HEADERS}, got {headers}')

    rows = []
    errors = []
    seen_codes = set()
    type_counter = Counter()
    for row_idx in range(2, worksheet.max_row + 1):
        values = [worksheet.cell(row_idx, col_idx).value for col_idx in range(1, 7)]
        if all(value in (None, '') for value in values):
            continue
        try:
            code = as_code(values[0])
            name = normalize_name(str(values[1] or '').strip())
            raw_type = str(values[2] or '').strip()
            reconcile = as_bool(values[3])

            if not name:
                raise ValueError('Account name is required')
            if raw_type not in TYPE_MAPPING:
                raise ValueError(f'Unknown account type: {raw_type}')
            if code in seen_codes:
                raise ValueError(f'Duplicate account code: {code}')

            seen_codes.add(code)
            type_counter[raw_type] += 1
            rows.append({
                'id': f'ae_custom_account_{code}',
                'name': name,
                'code': code,
                'account_type': TYPE_MAPPING[raw_type],
                'reconcile': 'True' if reconcile else 'False',
                'tax_ids': DEFAULT_TAX_XMLIDS_BY_CODE.get(code, ''),
            })
        except ValueError as err:
            errors.append(f'Row {row_idx}: {err}')

    required_singletons = ('Receivable', 'Payable', 'Current Year Earnings')
    for account_type in required_singletons:
        if type_counter[account_type] != 1:
            errors.append(f"Expected exactly one '{account_type}', found {type_counter[account_type]}")

    if errors:
        raise ValueError('\n'.join(errors))
    return rows


def write_rows(rows: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open('w', newline='', encoding='utf-8') as csv_file:
        writer = csv.DictWriter(
            csv_file,
            fieldnames=('id', 'name', 'code', 'account_type', 'reconcile', 'tax_ids'),
            quoting=csv.QUOTE_ALL,
        )
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--input', required=True, type=Path, help='Source XLSX path')
    parser.add_argument('--output', required=True, type=Path, help='Destination CSV path')
    args = parser.parse_args()

    rows = parse_rows(args.input)
    write_rows(rows, args.output)
    print(f'Wrote {len(rows)} accounts to {args.output}')


if __name__ == '__main__':
    main()
