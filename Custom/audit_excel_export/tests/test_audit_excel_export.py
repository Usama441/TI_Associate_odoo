import base64
import datetime
import io
import unittest
from unittest.mock import patch

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = load_workbook = None

from odoo import Command, fields
from odoo.addons.audit_excel_export.utils import clean_bank_narration
from odoo.addons.account.tests.common import AccountTestInvoicingCommon
from odoo.exceptions import ValidationError
from odoo.tests.common import tagged


@tagged('post_install', '-at_install')
@unittest.skipUnless(load_workbook, 'openpyxl is required for XLSX assertions')
class TestAuditExcelExport(AccountTestInvoicingCommon):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        cls.export_date_from = fields.Date.from_string('2025-01-01')
        cls.export_date_to = fields.Date.from_string('2025-12-31')

        cls.customer_partner = cls.partner_a
        cls.vendor_partner = cls.partner_b

        cls.customer_invoice_posted = cls.init_invoice(
            'out_invoice',
            partner=cls.customer_partner,
            invoice_date=fields.Date.from_string('2025-01-10'),
            post=True,
            amounts=[100.0],
        )
        cls.customer_invoice_posted_second = cls.init_invoice(
            'out_invoice',
            partner=cls.customer_partner,
            invoice_date=fields.Date.from_string('2025-01-18'),
            post=True,
            amounts=[50.0],
        )
        cls.customer_invoice_grouped = cls.init_invoice(
            'out_invoice',
            partner=cls.customer_partner,
            invoice_date=fields.Date.from_string('2025-01-22'),
            post=False,
            amounts=[60.0, 40.0],
        )
        cls.customer_invoice_grouped.invoice_line_ids[0].name = 'Grouped line 1'
        cls.customer_invoice_grouped.invoice_line_ids[1].name = 'Grouped line 2'
        cls.customer_invoice_grouped.action_post()
        cls.vendor_bill_posted = cls.init_invoice(
            'in_invoice',
            partner=cls.vendor_partner,
            invoice_date=fields.Date.from_string('2025-01-14'),
            post=True,
            amounts=[80.0],
        )

    def _create_wizard(self, **overrides):
        vals = {
            'company_ids': [Command.set(self.env.company.ids)],
            'date_from': self.export_date_from,
            'date_to': self.export_date_to,
            'export_sheet_key': 'all',
            'include_draft_entries': True,
            'unfold_all': True,
            'hide_zero_lines': False,
            'aging_based_on': 'base_on_maturity_date',
            'aging_interval': 30,
            'show_currency': True,
            'show_account': True,
            'include_dynamic_columns': True,
            'invoice_bill_scope': 'all_states',
            'include_refunds': True,
        }
        vals.update(overrides)
        return self.env['audit.excel.export.wizard'].create(vals)

    def _export_workbook(self, wizard):
        action = wizard.action_export_xlsx()
        self.assertEqual(action.get('type'), 'ir.actions.act_url')
        self.assertTrue(wizard.file_data)
        binary = base64.b64decode(wizard.file_data)
        self.assertGreater(len(binary), 0)
        wb = load_workbook(io.BytesIO(binary), data_only=False)
        return action, wb

    def _sheet_values(self, ws):
        values = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    values.append(str(cell.value))
        return values

    def _find_header_position(self, ws, header_name, *, max_scan_rows=80):
        for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
            for col_idx in range(1, ws.max_column + 1):
                if ws.cell(row=row_idx, column=col_idx).value == header_name:
                    return row_idx, col_idx
        return None, None

    def _normalize_label(self, value):
        return ''.join(ch for ch in str(value or '').strip().lower() if ch.isalnum())

    def _find_row_by_label(self, ws, label, *, start_row=1, occurrence=1):
        expected = self._normalize_label(label)
        seen = 0
        for row_idx in range(start_row, ws.max_row + 1):
            if self._normalize_label(ws.cell(row=row_idx, column=1).value) != expected:
                continue
            seen += 1
            if seen == occurrence:
                return row_idx
        return None

    def _find_row_by_keys(self, ws, keys, *, start_row=1):
        normalized = {self._normalize_label(key) for key in keys}
        for row_idx in range(start_row, ws.max_row + 1):
            if self._normalize_label(ws.cell(row=row_idx, column=1).value) in normalized:
                return row_idx
        return None

    def _find_row_containing_text(self, ws, text, *, start_row=1, occurrence=1):
        expected = self._normalize_label(text)
        seen = 0
        for row_idx in range(start_row, ws.max_row + 1):
            if expected not in self._normalize_label(ws.cell(row=row_idx, column=1).value):
                continue
            seen += 1
            if seen == occurrence:
                return row_idx
        return None

    def _find_cell_by_value(self, ws, expected):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == expected:
                    return cell
        return None

    def _find_column_by_label(self, ws, label, *, start_row=1, end_row=10):
        expected = self._normalize_label(label)
        for row_idx in range(start_row, min(ws.max_row, end_row) + 1):
            for col_idx in range(1, ws.max_column + 1):
                if self._normalize_label(ws.cell(row=row_idx, column=col_idx).value) == expected:
                    return col_idx
        return None

    def _sheet_formulas(self, ws, *, max_scan_rows=400, max_scan_cols=12):
        formulas = []
        for row in ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, max_scan_rows),
            min_col=1,
            max_col=min(ws.max_column, max_scan_cols),
        ):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas.append((cell.coordinate, cell.value))
        return formulas

    def _coerce_openpyxl_date(self, value):
        if isinstance(value, datetime.datetime):
            return value.date()
        return value

    def _static_numeric_like_values(self, ws):
        values = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 250), min_col=1, max_col=min(ws.max_column, 8)):
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                if isinstance(value, str) and value.startswith('='):
                    continue
                if isinstance(value, (int, float, bool, datetime.date, datetime.datetime)):
                    values.append((cell.coordinate, value))
        return values

    def test_01_export_download_and_sheet_order(self):
        wizard = self._create_wizard()
        action, wb = self._export_workbook(wizard)

        self.assertIn('download=true', action.get('url', ''))
        self.assertTrue(wizard.file_name.endswith('.xlsx'))
        self.assertEqual(
            wb.sheetnames,
            [
                'Client Details',
                'General Ledger',
                'Trial Balance',
                'SOCI',
                'SOCE',
                'SOFP',
                'SOCF',
                'PPE',
                'Prepayment',
                'Bank Summary',
                'Customer Invoices',
                'Aged Receivables',
                'Vendor Bills',
                'Aged Payables',
                'VAT Control',
                'Accruals',
                'Summary Sheet',
                'Share Capital',
            ],
        )

    def test_02_template_formulas_and_formatting_survive(self):
        wizard = self._create_wizard()
        _action, wb = self._export_workbook(wizard)

        self.assertTrue(str(wb['Summary Sheet']['A1'].value).startswith('='))
        self.assertTrue(str(wb['Summary Sheet']['B7'].value).startswith('='))
        self.assertTrue(str(wb['SOFP']['A1'].value).startswith('='))
        self.assertTrue(str(wb['VAT Control']['B18'].value).startswith('='))

        sofp = wb['SOFP']
        self.assertIn('A2:E2', {str(rng) for rng in sofp.merged_cells.ranges})
        self.assertTrue(sofp['A2'].font.bold)
        self.assertEqual(sofp.page_setup.orientation, 'landscape')
        self.assertEqual(sofp.column_dimensions['B'].width, 18.0)
        self.assertEqual(sofp.row_dimensions[2].height, 28.0)

    def test_03_first_five_sheets_have_data(self):
        wizard = self._create_wizard()
        _action, wb = self._export_workbook(wizard)

        gl_values = self._sheet_values(wb['General Ledger'])
        self.assertIn('Code', gl_values)
        self.assertIn('Account Name', gl_values)
        self.assertIn('Debit', gl_values)
        self.assertIn('Credit', gl_values)
        self.assertIn('Balance', gl_values)
        self.assertGreater(len(gl_values), 20)

        customer_ws = wb['Customer Invoices']
        customer_values = self._sheet_values(customer_ws)
        expected_customer_headers = [
            'Invoice/Bill Date',
            'Number',
            'Invoice Partner Display Name',
            'Invoice lines/Label',
            'Currency',
            'Currency/Inverse Rate',
            'Invoice lines/Quantity',
            'Invoice lines/Unit Price',
            'Invoice lines/Subtotal',
            'Amount paid',
            'Payments/Date',
            'Amount Due',
            'Amount Due Signed',
            'Total',
            'Total Signed',
        ]
        for header in expected_customer_headers:
            self.assertIn(header, customer_values)
        self.assertGreater(len(customer_values), 20)
        self.assertIn(self.customer_invoice_posted.name, customer_values)
        self.assertIn(self.customer_invoice_posted_second.name, customer_values)
        self.assertIn(self.customer_invoice_grouped.name, customer_values)
        self.assertIn('Grouped line 1', customer_values)
        self.assertIn('Grouped line 2', customer_values)
        self.assertIn('Invoice', customer_values)
        self.assertIn('1. Grouped line 1', customer_values)
        self.assertIn('2. Grouped line 2', customer_values)
        self.assertEqual(customer_values.count('Invoice/Bill Date'), 1)
        self.assertEqual(customer_values.count('Invoice lines/Label'), 1)
        self.assertEqual(customer_values.count('Currency/Inverse Rate'), 1)
        self.assertEqual(customer_values.count(self.customer_invoice_grouped.name), 1)

        number_col = self._find_column_by_label(customer_ws, 'Number')
        label_col = self._find_column_by_label(customer_ws, 'Invoice lines/Label')
        grouped_invoice_row = None
        for row_idx in range(1, customer_ws.max_row + 1):
            if customer_ws.cell(row=row_idx, column=number_col).value == self.customer_invoice_grouped.name:
                grouped_invoice_row = row_idx
                break
        self.assertTrue(grouped_invoice_row)
        self.assertEqual(customer_ws.cell(row=grouped_invoice_row, column=label_col).value, 'Invoice')
        self.assertFalse(customer_ws.cell(row=grouped_invoice_row + 1, column=number_col).value)
        self.assertEqual(customer_ws.cell(row=grouped_invoice_row + 1, column=label_col).value, '1. Grouped line 1')
        self.assertFalse(customer_ws.cell(row=grouped_invoice_row + 2, column=number_col).value)
        self.assertEqual(customer_ws.cell(row=grouped_invoice_row + 2, column=label_col).value, '2. Grouped line 2')

        vendor_ws = wb['Vendor Bills']
        vendor_values = self._sheet_values(vendor_ws)
        expected_line_headers = [
            'Label',
            'Account',
            'Quantity',
            'Price',
            'Taxes',
            'VAT Amount',
            'Amount',
            'Currency',
        ]
        for header in ('Invoice No', 'Invoice Date', 'Accounting Date', 'Currency', 'Vendor', 'Due Date', 'Conversion Rate'):
            self.assertIn(header, vendor_values)
        for header in expected_line_headers:
            self.assertIn(header, vendor_values)
        self.assertGreater(len(vendor_values), 20)
        self.assertIn(self.vendor_bill_posted.name, vendor_values)

        for sheet_name, ws in (('Customer Invoices', customer_ws), ('Vendor Bills', vendor_ws)):
            header_row, line_description_col = self._find_header_position(ws, 'Label')
            self.assertIsNotNone(header_row)
            has_line_rows = any(
                ws.cell(row=row_idx, column=line_description_col).value not in (None, '')
                for row_idx in range(header_row + 1, min(ws.max_row, header_row + 250) + 1)
            )
            self.assertTrue(has_line_rows, f'Expected invoice/bill line rows in {sheet_name}')

            for numeric_header in ('Quantity', 'Price', 'VAT Amount', 'Amount'):
                _hdr_row, numeric_col = self._find_header_position(ws, numeric_header)
                self.assertIsNotNone(numeric_col)
                has_numeric_value = any(
                    isinstance(ws.cell(row=row_idx, column=numeric_col).value, (int, float))
                    for row_idx in range(header_row + 1, min(ws.max_row, header_row + 250) + 1)
                )
                self.assertTrue(has_numeric_value, f'Expected numeric values under {numeric_header} in {sheet_name}')

            _hdr_row, conversion_col = self._find_header_position(ws, 'Conversion Rate')
            self.assertIsNotNone(conversion_col)
            has_conversion_value = any(
                isinstance(ws.cell(row=row_idx, column=conversion_col).value, (int, float))
                for row_idx in range(_hdr_row + 1, min(ws.max_row, _hdr_row + 250) + 1)
            )
            self.assertTrue(has_conversion_value, f'Expected conversion rate values in {sheet_name}')

        for sheet_name in ('Aged Receivables', 'Aged Payables'):
            ws = wb[sheet_name]
            values = self._sheet_values(ws)
            self.assertGreater(ws.max_row, 5)
            self.assertGreater(ws.max_column, 5)
            self.assertGreater(len(values), 20)

    def test_04_trial_balance_is_populated(self):
        wizard = self._create_wizard()
        _action, wb = self._export_workbook(wizard)

        ws = wb['Trial Balance']
        values = self._sheet_values(ws)
        self.assertIn('Code', values)
        self.assertIn('Account Name', values)
        self.assertIn('Debit', values)
        self.assertIn('Credit', values)
        self.assertIn('Balance', values)

        has_numeric_tb_rows = any(
            isinstance(cell.value, (int, float)) and abs(cell.value) > 0
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 300), min_col=1, max_col=min(ws.max_column, 12))
            for cell in row
        )
        self.assertTrue(has_numeric_tb_rows)

    def test_04b_trial_balance_layout_distinguishes_opening_and_ending_balance_columns(self):
        wizard = self._create_wizard(export_sheet_key='trial_balance')
        workbook = Workbook()
        ws = workbook.active
        ws.title = 'Trial Balance'

        ws['A3'] = 'Code'
        ws['B3'] = 'Account Name'
        ws['C3'] = 'Balance'
        ws['D3'] = 'Debit'
        ws['E3'] = 'Credit'
        ws['F3'] = 'Balance'

        ws['A4'] = '12030101'
        ws['B4'] = 'Prepayment'
        ws['C4'] = 125.0
        ws['D4'] = 40.0
        ws['E4'] = 15.0
        ws['F4'] = 150.0
        ws['A5'] = 'Total'

        layout, rows, total_row = wizard._extract_tb_data_rows_from_worksheet(ws)

        self.assertEqual(layout.get('opening_balance_col'), 3)
        self.assertEqual(layout.get('balance_col'), 6)
        self.assertEqual(total_row, 5)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].get('opening_balance'), 125.0)
        self.assertEqual(rows[0].get('balance'), 150.0)

    def test_04c_trial_balance_helper_columns_reference_balance_columns_as_source_of_truth(self):
        wizard = self._create_wizard(export_sheet_key='trial_balance')
        workbook = Workbook()
        ws = workbook.active
        ws.title = 'Trial Balance'

        ws['A3'] = 'Code'
        ws['B3'] = 'Account Name'
        ws['C3'] = 'Balance'
        ws['D3'] = 'Debit'
        ws['E3'] = 'Credit'
        ws['F3'] = 'Balance'
        ws['A4'] = '12030101'
        ws['B4'] = 'Prepayment'
        ws['C4'] = 4163.00
        ws['D4'] = 21050.00
        ws['E4'] = 19252.85
        ws['F4'] = 5960.15
        ws['A5'] = 'Total'

        layout, rows, total_row = wizard._extract_tb_data_rows_from_worksheet(ws)
        metric_refs = wizard._prepare_tb_formula_links(
            tb_sheet=ws,
            layout=layout,
            rows=rows,
            total_row=total_row,
            tb_maps={
                'periods': {},
                'prior_period': {
                    '12030101': {'debit': 0.0, 'credit': 0.0, 'balance': 4163.00},
                },
                'opening_current': {},
                'opening_prior': {
                    '12030101': {'balance': 0.0},
                },
            },
        )

        self.assertTrue(metric_refs)
        self.assertEqual(ws['L4'].value, '=F4')
        self.assertEqual(ws['M4'].value, '=I4')

    def test_05_client_details_are_populated(self):
        wizard = self._create_wizard()
        _action, wb = self._export_workbook(wizard)

        ws = wb['Client Details']
        self.assertEqual(ws['B4'].value, self.env.company.name)
        self.assertEqual(self._coerce_openpyxl_date(ws['B5'].value), self.export_date_from)
        self.assertEqual(self._coerce_openpyxl_date(ws['B6'].value), self.export_date_to)

    def test_06_share_capital_rows_are_populated_when_company_data_exists(self):
        self.env.company.write({
            'shareholder_1': 'Owner A',
            'number_of_shares_1': 100,
            'share_value_1': 10.0,
        })

        wizard = self._create_wizard()
        _action, wb = self._export_workbook(wizard)

        ws = wb['Share Capital']
        self.assertEqual(ws['A7'].value, 'Owner A')
        self.assertEqual(ws['B7'].value, 100)
        self.assertEqual(ws['C7'].value, 10.0)

    def test_07_non_tb_template_sheets_have_no_static_numeric_samples(self):
        wizard = self._create_wizard()
        _action, wb = self._export_workbook(wizard)

        non_tb_template_sheets = [
            'VAT Control',
            'Prepayment',
            'Summary Sheet',
            'SOFP',
            'SOCI',
            'SOCE',
            'SOCF',
            'Accruals',
        ]

        for sheet_name in non_tb_template_sheets:
            ws = wb[sheet_name]
            static_values = self._static_numeric_like_values(ws)
            self.assertFalse(static_values, f'Static sample numeric values found on {sheet_name}: {static_values[:5]}')

    def test_08_trial_balance_forces_flat_unfolded_options(self):
        wizard = self._create_wizard(unfold_all=False)

        with patch.object(type(wizard), '_build_native_report_sheet_payload', autospec=True, return_value={}) as mocked_builder:
            wizard._prepare_trial_balance_payload()

        options = mocked_builder.call_args.kwargs['options']
        self.assertFalse(options.get('hierarchy'))
        self.assertTrue(options.get('unfold_all'))
        self.assertEqual(options.get('unfolded_lines'), [])

    def test_09_customer_invoice_report_options_are_forwarded(self):
        wizard = self._create_wizard(invoice_bill_scope='posted_only', include_refunds=False)

        with patch.object(type(wizard), '_get_invoice_bill_moves', autospec=True, return_value=self.env['account.move']) as mocked_get_moves:
            wizard._prepare_invoice_bill_payload(is_customer=True)

        options = mocked_get_moves.call_args.kwargs['options']
        self.assertEqual(options.get('invoice_bill_report_kind'), 'customer')
        self.assertEqual(options.get('invoice_bill_scope'), 'posted_only')
        self.assertFalse(options.get('include_refunds'))

    def test_10_statement_sheets_use_live_link_formula_graph_without_loops(self):
        wizard = self._create_wizard()
        _action, wb = self._export_workbook(wizard)

        sofp = wb['SOFP']
        soci = wb['SOCI']
        soce = wb['SOCE']
        socf = wb['SOCF']
        summary = wb['Summary Sheet']

        sofp_cash_row = self._find_row_by_label(sofp, 'Cash and bank balances')
        sofp_retained_row = self._find_row_by_label(sofp, 'Retained earnings')
        soci_revenue_row = self._find_row_by_label(soci, 'Revenue')
        soce_retained_col = self._find_column_by_label(soce, 'Retained earnings')
        soce_prior_profit_row = self._find_row_by_label(soce, 'Net profit / (loss)', occurrence=1) or self._find_row_by_label(soce, 'Net profit', occurrence=1)
        soce_current_profit_row = self._find_row_by_label(soce, 'Net profit / (loss)', occurrence=2) or self._find_row_by_label(soce, 'Net profit', occurrence=2)
        soce_prior_profit_formula = str(soce.cell(row=soce_prior_profit_row, column=soce_retained_col).value or '')
        soce_current_profit_formula = str(soce.cell(row=soce_current_profit_row, column=soce_retained_col).value or '')
        socf_net_profit_row = self._find_row_by_label(socf, 'Net profit for the year') or self._find_row_by_label(socf, 'Net profit for the period')
        socf_change_assets_row = self._find_row_by_label(socf, '(Increase) / decrease in current assets') or self._find_row_by_label(socf, 'Increase / decrease in current assets')

        self.assertIsNotNone(sofp_cash_row)
        self.assertIsNotNone(sofp_retained_row)
        self.assertIsNotNone(soci_revenue_row)
        self.assertIsNotNone(soce_retained_col)
        self.assertIsNotNone(soce_prior_profit_row)
        self.assertIsNotNone(soce_current_profit_row)
        self.assertIsNotNone(socf_net_profit_row)
        self.assertIsNotNone(socf_change_assets_row)
        self.assertIn('Trial Balance', str(sofp[f'B{sofp_cash_row}'].value or ''))
        self.assertIn('SOCE!', str(sofp[f'B{sofp_retained_row}'].value or ''))
        self.assertNotIn('SOCI!', str(sofp[f'B{sofp_retained_row}'].value or ''))
        self.assertIn('Trial Balance', str(soci[f'B{soci_revenue_row}'].value or ''))
        self.assertIn('SOCI!', soce_prior_profit_formula)
        self.assertIn('SOCI!', soce_current_profit_formula)
        self.assertIn('SOCI!', str(socf[f'B{socf_net_profit_row}'].value or ''))
        self.assertIn('SOFP!', str(socf[f'B{socf_change_assets_row}'].value or ''))
        self.assertIn('SOCI!', str(summary['B7'].value or ''))
        self.assertIn('SOFP!', str(summary['B11'].value or ''))

        for _coord, formula in self._sheet_formulas(soci):
            self.assertNotIn('SOFP!', formula)
            self.assertNotIn('SOCE!', formula)
            self.assertNotIn('SOCF!', formula)

        for _coord, formula in self._sheet_formulas(soce):
            self.assertNotIn('SOFP!', formula)
            self.assertNotIn('SOCF!', formula)

        for _coord, formula in self._sheet_formulas(socf):
            self.assertNotIn('SOCE!', formula)
            self.assertNotIn('SOCF!', formula)

    def test_11_last_saved_settings_are_loaded(self):
        default_key_wizard = self._create_wizard()
        settings_key = default_key_wizard._previous_settings_key()
        self.env['ir.config_parameter'].sudo().set_param(settings_key, '')

        custom_wizard = self._create_wizard(
            date_from=fields.Date.from_string('2025-02-01'),
            date_to=fields.Date.from_string('2025-02-28'),
            export_sheet_key='general_ledger',
            year_span='2y',
            balance_sheet_date_mode='range',
            prior_year_mode='manual',
            prior_balance_sheet_date_mode='range',
            prior_date_start=fields.Date.from_string('2024-02-01'),
            prior_date_end=fields.Date.from_string('2024-02-29'),
            include_draft_entries=False,
            unfold_all=False,
            hide_zero_lines=True,
            aging_based_on='base_on_invoice_date',
            aging_interval=45,
            show_currency=False,
            show_account=False,
            include_dynamic_columns=False,
            invoice_bill_scope='posted_only',
            include_refunds=False,
            gl_options_json='{"unfold_all": false}',
            aged_receivable_options_json='{"show_currency": false}',
            aged_payable_options_json='{"show_currency": false}',
        )
        custom_wizard._store_previous_settings()

        field_names = [
            'use_previous_settings',
            'date_from',
            'date_to',
            'export_sheet_key',
            'year_span',
            'balance_sheet_date_mode',
            'prior_year_mode',
            'prior_balance_sheet_date_mode',
            'prior_date_start',
            'prior_date_end',
            'include_draft_entries',
            'unfold_all',
            'hide_zero_lines',
            'aging_based_on',
            'aging_interval',
            'show_currency',
            'show_account',
            'include_dynamic_columns',
            'invoice_bill_scope',
            'include_refunds',
            'gl_options_json',
            'aged_receivable_options_json',
            'aged_payable_options_json',
        ]
        defaults = self.env['audit.excel.export.wizard'].default_get(field_names)

        self.assertEqual(defaults.get('date_from'), fields.Date.from_string('2025-02-01'))
        self.assertEqual(defaults.get('date_to'), fields.Date.from_string('2025-02-28'))
        self.assertEqual(defaults.get('export_sheet_key'), 'general_ledger')
        self.assertEqual(defaults.get('year_span'), '2y')
        self.assertEqual(defaults.get('balance_sheet_date_mode'), 'range')
        self.assertEqual(defaults.get('prior_year_mode'), 'manual')
        self.assertEqual(defaults.get('prior_balance_sheet_date_mode'), 'range')
        self.assertEqual(defaults.get('prior_date_start'), fields.Date.from_string('2024-02-01'))
        self.assertEqual(defaults.get('prior_date_end'), fields.Date.from_string('2024-02-29'))
        self.assertFalse(defaults.get('include_draft_entries'))
        self.assertFalse(defaults.get('unfold_all'))
        self.assertTrue(defaults.get('hide_zero_lines'))
        self.assertEqual(defaults.get('aging_based_on'), 'base_on_invoice_date')
        self.assertEqual(defaults.get('aging_interval'), 45)
        self.assertFalse(defaults.get('show_currency'))
        self.assertFalse(defaults.get('show_account'))
        self.assertFalse(defaults.get('include_dynamic_columns'))
        self.assertEqual(defaults.get('invoice_bill_scope'), 'posted_only')
        self.assertFalse(defaults.get('include_refunds'))
        self.assertEqual(defaults.get('gl_options_json'), '{"unfold_all": false}')
        self.assertEqual(defaults.get('aged_receivable_options_json'), '{"show_currency": false}')
        self.assertEqual(defaults.get('aged_payable_options_json'), '{"show_currency": false}')

    def test_12_can_export_only_general_ledger_sheet(self):
        wizard = self._create_wizard(export_sheet_key='general_ledger')
        _action, wb = self._export_workbook(wizard)

        self.assertEqual(wb.sheetnames, ['General Ledger'])
        values = self._sheet_values(wb['General Ledger'])
        self.assertIn('Code', values)
        self.assertIn('Account Name', values)
        self.assertIn('Debit', values)
        self.assertIn('Credit', values)
        self.assertIn('Balance', values)
        self.assertIn('general_ledger', wizard.file_name)

    def test_13_can_export_only_template_sheet(self):
        wizard = self._create_wizard(export_sheet_key='summary_sheet')
        _action, wb = self._export_workbook(wizard)

        self.assertEqual(
            wb.sheetnames,
            ['Client Details', 'Trial Balance', 'SOCI', 'SOCE', 'SOFP', 'Summary Sheet'],
        )
        self.assertTrue(str(wb['Summary Sheet']['A1'].value).startswith('='))

    def test_14_narration_parser_samples(self):
        parsed_1 = clean_bank_narration(
            'BNK1/2025/00042 INWARD T/T Return of deposit MARINA TSYBINA'
        )
        self.assertEqual(parsed_1.get('direction'), 'IN')
        self.assertEqual(parsed_1.get('rail'), 'TT')
        self.assertEqual(parsed_1.get('bank_ref'), 'BNK1/2025/00042')
        self.assertIn('IN TT', parsed_1.get('clean', ''))
        self.assertIn('BNK1/2025/00042', parsed_1.get('clean', ''))

        parsed_2 = clean_bank_narration(
            'BNK1/2025/00043 CHARGE COLLECTION-INCL. VAT '
            'Charge Collection-INCL. VAT OUTWARD REMITTANCE CHARGE'
        )
        self.assertEqual(parsed_2.get('rail'), 'FEE')
        self.assertIn('VAT', parsed_2.get('clean', ''))
        self.assertEqual(parsed_2.get('bank_ref'), 'BNK1/2025/00043')

        parsed_3 = clean_bank_narration(
            'BNK1/2025/00047 CREDIT CARD PAYMENT 536272XXXXXX7871'
        )
        self.assertEqual(parsed_3.get('rail'), 'CARD')
        self.assertEqual(parsed_3.get('card_last4'), '7871')
        self.assertIn('****7871', parsed_3.get('clean', ''))

        parsed_4 = clean_bank_narration(
            'BNK1/2025/00053 INWARD T/T /REF//BENEFRES/AE//PRS INVOICE NUMB ER 224 GOVINDER SINGH SIDHU'
        )
        self.assertEqual(parsed_4.get('direction'), 'IN')
        self.assertEqual(parsed_4.get('odoo_ref'), 'PRS-224')
        self.assertEqual(parsed_4.get('bank_ref'), 'BNK1/2025/00053')
        self.assertNotIn('1.0000', parsed_4.get('clean', ''))

        long_text = (
            'BNK1/2025/00999 INWARD T/T '
            + ('VERY LONG DESCRIPTION ' * 20)
            + 'PRS INVOICE NUMBER 998'
        )
        parsed_long = clean_bank_narration(long_text, max_len=100)
        self.assertLessEqual(len(parsed_long.get('clean', '')), 100)

    def test_15_clean_general_ledger_rows_and_keep_non_detail_rows(self):
        wizard = self._create_wizard()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'General Ledger'

        sheet['A3'] = 'Code'
        sheet['B3'] = 'Account Name'
        sheet['G3'] = 'Debit'
        sheet['H3'] = 'Credit'
        sheet['I3'] = 'Balance'

        sheet['A4'] = '12020101'
        sheet['B4'] = 'Accounts Receivable'
        sheet['G4'] = 100.0
        sheet['H4'] = 0.0
        sheet['I4'] = 100.0

        sheet['B5'] = 'BNK1/2025/00047 CREDIT CARD PAYMENT 536272XXXXXX7871'
        sheet['G5'] = 0.0
        sheet['H5'] = 10.0
        sheet['I5'] = -10.0

        sheet['B6'] = 'Initial Balance'
        sheet['B7'] = 'Total 12020101 Account Receivable'
        sheet['B8'] = 'Load more...'
        sheet['B9'] = 'BNK1/2025/00053 INWARD T/T /REF//BENEFRES/AE//PRS INVOICE NUMB ER 224 GOVINDER SINGH SIDHU'

        before_amount_tuple = (sheet['G5'].value, sheet['H5'].value, sheet['I5'].value)
        wizard._clean_general_ledger_sheet_narrations(sheet)

        self.assertEqual(sheet['B4'].value, 'Accounts Receivable')
        self.assertEqual(sheet['B6'].value, 'Initial Balance')
        self.assertEqual(sheet['B7'].value, 'Total 12020101 Account Receivable')
        self.assertEqual(sheet['B8'].value, 'Load more...')
        self.assertTrue(str(sheet['B5'].value).startswith('CARD | Payment | ****7871'))
        self.assertIn('Raw narration:', sheet['B5'].comment.text)
        self.assertIn('CREDIT CARD PAYMENT 536272XXXXXX7871', sheet['B5'].comment.text)
        self.assertEqual(before_amount_tuple, (sheet['G5'].value, sheet['H5'].value, sheet['I5'].value))
        self.assertIn('PRS-224', str(sheet['B9'].value))

    def test_16_general_ledger_cleaner_feature_flag_in_stage_b(self):
        wizard = self._create_wizard(export_sheet_key='general_ledger')
        config = self.env['ir.config_parameter'].sudo()
        parameter_key = 'audit_excel_export.gl_narration_cleaner_enabled'

        source_workbook = Workbook()
        source_sheet = source_workbook.active
        source_sheet.title = 'General Ledger'
        source_sheet['A3'] = 'Code'
        source_sheet['B3'] = 'Account Name'
        source_sheet['B4'] = 'BNK1/2025/00047 CREDIT CARD PAYMENT 536272XXXXXX7871'

        stage_a_payload = {
            'data_sheets': [
                {
                    'sheet_name': 'General Ledger',
                    'native_source_sheet': source_sheet,
                },
            ],
            'trial_balance': {},
        }

        config.set_param(parameter_key, 'false')
        stage_b_disabled = wizard._stage_b_create_workbook_and_write_data_sheets(stage_a_payload)
        disabled_sheet = stage_b_disabled['workbook']['General Ledger']
        self.assertEqual(disabled_sheet['B4'].value, 'BNK1/2025/00047 CREDIT CARD PAYMENT 536272XXXXXX7871')
        self.assertFalse(disabled_sheet['B4'].comment)

        config.set_param(parameter_key, '')
        source_workbook_enabled = Workbook()
        source_sheet_enabled = source_workbook_enabled.active
        source_sheet_enabled.title = 'General Ledger'
        source_sheet_enabled['A3'] = 'Code'
        source_sheet_enabled['B3'] = 'Account Name'
        source_sheet_enabled['B4'] = 'BNK1/2025/00047 CREDIT CARD PAYMENT 536272XXXXXX7871'

        stage_b_enabled = wizard._stage_b_create_workbook_and_write_data_sheets(
            {
                'data_sheets': [
                    {
                        'sheet_name': 'General Ledger',
                        'native_source_sheet': source_sheet_enabled,
                    },
                ],
                'trial_balance': {},
            }
        )
        enabled_sheet = stage_b_enabled['workbook']['General Ledger']
        self.assertTrue(str(enabled_sheet['B4'].value).startswith('CARD | Payment | ****7871'))
        self.assertTrue(enabled_sheet['B4'].comment)

    def test_17_statement_export_uses_live_links_without_audit_report_snapshot_data(self):
        wizard = self._create_wizard(export_sheet_key='sofp')

        with patch.object(
            type(wizard),
            '_get_audit_report_statement_data',
            autospec=True,
            side_effect=AssertionError('audit report snapshot path should not be used for live-link exports'),
        ):
            _action, wb = self._export_workbook(wizard)

        self.assertEqual(
            wb.sheetnames,
            ['Client Details', 'Trial Balance', 'SOCI', 'SOCE', 'SOFP'],
        )
        retained_row = self._find_row_by_label(wb['SOFP'], 'Retained earnings')
        self.assertIsNotNone(retained_row)
        self.assertIn('SOCE!', str(wb['SOFP'][f'B{retained_row}'].value or ''))

    def test_18_one_year_period_category_skips_prior_period_windows(self):
        wizard = self._create_wizard(
            year_span='1y',
            prior_year_mode='manual',
            prior_date_start=False,
            prior_date_end=False,
        )

        periods = wizard._get_reporting_periods()
        self.assertFalse(periods.get('show_prior_year'))
        self.assertFalse(periods.get('prior_date_start'))
        self.assertFalse(periods.get('prior_date_end'))
        self.assertFalse(periods.get('prior_opening_date_end'))

    def test_19_one_year_sofp_hides_prior_year_column_data(self):
        wizard = self._create_wizard(
            export_sheet_key='sofp',
            year_span='1y',
            prior_year_mode='manual',
            prior_date_start=False,
            prior_date_end=False,
        )
        _action, wb = self._export_workbook(wizard)

        sofp = wb['SOFP']
        self.assertIsNone(sofp['C6'].value)
        for row_idx in range(7, sofp.max_row + 1):
            if sofp.cell(row=row_idx, column=1).value in (None, ''):
                continue
            self.assertIsNone(sofp[f'C{row_idx}'].value)

    def test_20_soci_export_auto_includes_trial_balance_dependency(self):
        wizard = self._create_wizard(export_sheet_key='soci')
        _action, wb = self._export_workbook(wizard)

        self.assertEqual(wb.sheetnames, ['Client Details', 'Trial Balance', 'SOCI'])
        soci = wb['SOCI']
        revenue_row = self._find_row_by_label(soci, 'Revenue')
        net_profit_row = self._find_row_by_keys(soci, ['Net profit / (loss)', 'Net profit'])
        self.assertIsNotNone(revenue_row)
        self.assertIsNotNone(net_profit_row)
        self.assertIn('Trial Balance', str(soci[f'B{revenue_row}'].value or ''))
        self.assertTrue(str(soci[f'B{net_profit_row}'].value or '').startswith('='))

    def test_21_soce_export_auto_includes_trial_balance_and_soci_dependencies(self):
        wizard = self._create_wizard(export_sheet_key='soce')
        _action, wb = self._export_workbook(wizard)

        self.assertEqual(wb.sheetnames, ['Client Details', 'Trial Balance', 'SOCI', 'SOCE'])
        soce = wb['SOCE']
        share_col = self._find_column_by_label(soce, 'Share capital')
        retained_col = self._find_column_by_label(soce, 'Retained earnings')
        total_col = self._find_column_by_label(soce, 'Total')
        opening_row = self._find_row_by_label(soce, 'Balance as at start of period')
        prior_profit_row = self._find_row_by_label(soce, 'Net profit / (loss)', occurrence=1) or self._find_row_by_label(soce, 'Net profit', occurrence=1)
        current_profit_row = self._find_row_by_label(soce, 'Net profit / (loss)', occurrence=2) or self._find_row_by_label(soce, 'Net profit', occurrence=2)
        closing_row = self._find_row_by_label(soce, 'Balance c/f')
        self.assertIn('Trial Balance', str(soce.cell(row=opening_row, column=share_col).value or ''))
        self.assertIn('SOCI!', str(soce.cell(row=prior_profit_row, column=retained_col).value or ''))
        self.assertIn('SOCI!', str(soce.cell(row=current_profit_row, column=retained_col).value or ''))
        self.assertTrue(str(soce.cell(row=closing_row, column=total_col).value or '').startswith('=SUM('))

    def test_22_sofp_export_auto_includes_upstream_dependencies_and_links_equity_to_soce(self):
        wizard = self._create_wizard(export_sheet_key='sofp')
        _action, wb = self._export_workbook(wizard)

        self.assertEqual(
            wb.sheetnames,
            ['Client Details', 'Trial Balance', 'SOCI', 'SOCE', 'SOFP'],
        )
        sofp = wb['SOFP']
        soce = wb['SOCE']
        soce_prior_close_row = self._find_row_containing_text(soce, 'Balance as at', occurrence=2)
        soce_current_close_row = self._find_row_containing_text(soce, 'Balance as at', occurrence=3)
        share_row = self._find_row_by_label(sofp, 'Share capital') or self._find_row_by_label(sofp, 'Share Capital')
        retained_row = self._find_row_by_label(sofp, 'Retained earnings')
        owner_row = self._find_row_by_label(sofp, 'Owner current account')
        cash_row = self._find_row_by_label(sofp, 'Cash and bank balances')
        self.assertIsNotNone(soce_prior_close_row)
        self.assertIsNotNone(soce_current_close_row)
        self.assertIsNotNone(share_row)
        self.assertIsNotNone(retained_row)
        self.assertIsNotNone(owner_row)
        self.assertIsNotNone(cash_row)
        self.assertEqual(str(sofp[f'B{share_row}'].value or ''), f'=SOCE!B{soce_current_close_row}')
        self.assertEqual(str(sofp[f'C{share_row}'].value or ''), f'=SOCE!B{soce_prior_close_row}')
        self.assertEqual(str(sofp[f'B{retained_row}'].value or ''), f'=SOCE!D{soce_current_close_row}')
        self.assertEqual(str(sofp[f'C{retained_row}'].value or ''), f'=SOCE!D{soce_prior_close_row}')
        self.assertEqual(str(sofp[f'B{owner_row}'].value or ''), f'=SOCE!C{soce_current_close_row}')
        self.assertEqual(str(sofp[f'C{owner_row}'].value or ''), f'=SOCE!C{soce_prior_close_row}')
        self.assertIn('Trial Balance', str(sofp[f'B{cash_row}'].value or ''))
        self.assertNotIn('SOCI!', str(sofp[f'B{retained_row}'].value or ''))

    def test_22b_sofp_equity_lines_follow_existing_sofp_labels(self):
        wizard = self._create_wizard(export_sheet_key='sofp')
        wb = Workbook()
        sofp = wb.active
        sofp.title = 'SOFP'
        sofp['A20'] = 'Equity'
        sofp['A21'] = 'Retained earnings'
        sofp['A22'] = 'Share Capital'
        sofp['A23'] = 'Total Equity'

        soce = wb.create_sheet('SOCE')
        soce['B6'] = 'Share capital'
        soce['C6'] = "Owner's current account"
        soce['D6'] = 'Retained earnings'
        soce['E6'] = 'Statutory reserve'
        soce['A7'] = 'Balance as at 00 Month 0000'
        soce['A12'] = 'Balance as at 31 December 2024'
        soce['A17'] = 'Balance as at 31 December 2025'

        lines = wizard._build_sofp_equity_formula_lines(wb, sofp_sheet=sofp)

        self.assertEqual(
            [line['label'] for line in lines],
            ['Retained earnings', 'Share Capital'],
        )
        self.assertEqual(lines[0]['current_formula'], '=SOCE!D17')
        self.assertEqual(lines[0]['prior_formula'], '=SOCE!D12')
        self.assertEqual(lines[1]['current_formula'], '=SOCE!B17')
        self.assertEqual(lines[1]['prior_formula'], '=SOCE!B12')

    def test_22_socf_export_auto_includes_upstream_dependencies_with_sofp_links(self):
        wizard = self._create_wizard(export_sheet_key='socf')
        _action, wb = self._export_workbook(wizard)

        self.assertEqual(wb.sheetnames, ['Client Details', 'Trial Balance', 'SOCI', 'SOCF'])
        socf = wb['SOCF']
        net_profit_row = self._find_row_by_label(socf, 'Net profit for the year') or self._find_row_by_label(socf, 'Net profit for the period')
        change_assets_row = self._find_row_by_label(socf, '(Increase) / decrease in current assets') or self._find_row_by_label(socf, 'Increase / decrease in current assets')
        self.assertIn('SOCI!', str(socf[f'B{net_profit_row}'].value or ''))
        self.assertIn('SOFP!', str(socf[f'B{change_assets_row}'].value or ''))
        for _coord, formula in self._sheet_formulas(socf):
            self.assertNotIn('SOCE!', formula)
            self.assertNotIn('SOCF!', formula)

    def test_23_summary_sheet_links_only_to_final_statement_outputs(self):
        wizard = self._create_wizard(export_sheet_key='summary_sheet')
        _action, wb = self._export_workbook(wizard)

        self.assertEqual(
            wb.sheetnames,
            ['Client Details', 'Trial Balance', 'SOCI', 'SOCE', 'SOFP', 'Summary Sheet'],
        )
        formulas = [formula for _coord, formula in self._sheet_formulas(wb['Summary Sheet'])]
        self.assertTrue(any('SOCI!' in formula for formula in formulas))
        self.assertTrue(any('SOFP!' in formula for formula in formulas))
        self.assertFalse(any('Trial Balance!' in formula for formula in formulas))
        self.assertFalse(any('SOCE!' in formula for formula in formulas))
        self.assertFalse(any('SOCF!' in formula for formula in formulas))

    def test_24_year_end_date_uses_date_to(self):
        year_end_date = fields.Date.from_string('2025-12-31')
        wizard = self._create_wizard(date_to=year_end_date)

        self.assertEqual(wizard.date_to, year_end_date)
        self.assertEqual(wizard._get_effective_year_end_date(), year_end_date)

    def test_25_aged_payload_uses_year_end_date(self):
        year_end_date = fields.Date.from_string('2025-12-31')
        wizard = self._create_wizard(date_to=year_end_date)

        with patch.object(type(wizard), '_build_native_report_sheet_payload', autospec=True, return_value={}) as mocked_builder:
            wizard._prepare_aged_payload(
                xmlid='account_reports.aged_receivable_report',
                sheet_name='Aged Receivables',
                title='Aged Receivables',
                overrides={},
            )

        options = mocked_builder.call_args.kwargs['options']
        self.assertEqual(options.get('date', {}).get('date_from'), '2025-12-31')
        self.assertEqual(options.get('date', {}).get('date_to'), '2025-12-31')

    def test_26_statement_data_uses_native_trial_balance_movement_rows(self):
        wizard = self._create_wizard(
            export_sheet_key='soci',
            year_span='2y',
            prior_year_mode='manual',
            prior_date_start=fields.Date.from_string('2024-01-01'),
            prior_date_end=fields.Date.from_string('2024-12-31'),
        )
        audit_report_model = type(self.env['audit.report'])
        fetched_ranges = []

        def _movement_rows(_wizard_self, company, date_start, date_end):
            fetched_ranges.append((company.id, date_start, date_end))
            if fields.Date.to_string(date_end) == '2025-12-31':
                return [
                    {
                        'id': 1,
                        'code': '51010000',
                        'initial_balance': 10.0,
                        'debit': 250.0,
                        'credit': 50.0,
                        'movement_balance': 200.0,
                        'end_balance': 999.0,
                        'balance': 999.0,
                    },
                    {
                        'id': 2,
                        'code': '41030000',
                        'initial_balance': 0.0,
                        'debit': 0.0,
                        'credit': 75.0,
                        'movement_balance': -75.0,
                        'end_balance': 123.0,
                        'balance': 123.0,
                    },
                ]
            return [
                {
                    'id': 3,
                    'code': '51010000',
                    'initial_balance': 5.0,
                    'debit': 90.0,
                    'credit': 10.0,
                    'movement_balance': 80.0,
                    'end_balance': 321.0,
                    'balance': 321.0,
                },
            ]

        with patch.object(
            audit_report_model,
            '_get_report_data',
            autospec=True,
            return_value={'show_prior_year': True},
        ), patch.object(
            type(wizard),
            '_tb_fetch_movement_rows',
            autospec=True,
            side_effect=_movement_rows,
        ), patch.object(
            audit_report_model,
            '_fetch_grouped_account_rows_from_odoo_trial_balance',
            autospec=True,
            side_effect=AssertionError('audit.report TB helper should not be used here.'),
        ), patch.object(
            audit_report_model,
            '_fetch_grouped_account_rows',
            autospec=True,
            side_effect=AssertionError('audit.report raw grouped rows should not be used here.'),
        ):
            stmt_data = wizard._get_audit_report_statement_data()

        self.assertEqual(
            fetched_ranges,
            [
                (
                    self.env.company.id,
                    fields.Date.from_string('2025-01-01'),
                    fields.Date.from_string('2025-12-31'),
                ),
                (
                    self.env.company.id,
                    fields.Date.from_string('2024-01-01'),
                    fields.Date.from_string('2024-12-31'),
                ),
            ],
        )
        self.assertEqual(stmt_data['period_totals'].get('5101'), 200.0)
        self.assertEqual(stmt_data['period_totals'].get('4103'), -75.0)
        self.assertEqual(stmt_data['prior_period_totals'].get('5101'), 80.0)

    def test_27_socf_current_liabilities_metric_includes_bank_overdraft_and_credit_card(self):
        wizard = self._create_wizard()
        _action, wb = self._export_workbook(wizard)

        tb = wb['Trial Balance']
        metric_label_cell = self._find_cell_by_value(tb, 'socf_change_current_liabilities')
        self.assertIsNotNone(metric_label_cell, 'Missing SOCF current liabilities helper metric')

        current_formula = str(
            tb.cell(row=metric_label_cell.row, column=metric_label_cell.column + 1).value or ''
        )
        prior_formula = str(
            tb.cell(row=metric_label_cell.row, column=metric_label_cell.column + 2).value or ''
        )

        for formula in (current_formula, prior_formula):
            self.assertIn('220101*', formula)
            self.assertIn('220103*', formula)
            self.assertIn('220201*', formula)
            self.assertIn('220301*', formula)
            self.assertIn('220302*', formula)
            self.assertIn('220303*', formula)

    def test_28_soci_other_income_metric_uses_abs_for_positive_display(self):
        wizard = self._create_wizard(export_sheet_key='soci')
        _action, wb = self._export_workbook(wizard)

        tb = wb['Trial Balance']
        metric_label_cell = self._find_cell_by_value(tb, 'soci_other_income')
        self.assertIsNotNone(metric_label_cell, 'Missing SOCI other income helper metric')

        current_formula = str(
            tb.cell(row=metric_label_cell.row, column=metric_label_cell.column + 1).value or ''
        )
        prior_formula = str(
            tb.cell(row=metric_label_cell.row, column=metric_label_cell.column + 2).value or ''
        )

        self.assertIn('ABS(', current_formula)
        self.assertIn('ABS(', prior_formula)

    def test_29_soci_lines_use_direct_trial_balance_cell_link_formulas(self):
        wizard = self._create_wizard(export_sheet_key='soci')
        _action, wb = self._export_workbook(wizard)

        soci = wb['SOCI']
        revenue_row = self._find_row_by_label(soci, 'Revenue')
        director_salary_row = self._find_row_by_label(soci, 'Director salary')
        other_income_row = self._find_row_by_label(soci, 'Other income')
        self.assertIsNotNone(revenue_row)
        self.assertIsNotNone(director_salary_row)
        self.assertIsNotNone(other_income_row)

        revenue_formula = str(soci[f'B{revenue_row}'].value or '')
        director_salary_formula = str(soci[f'B{director_salary_row}'].value or '')
        other_income_formula = str(soci[f'B{other_income_row}'].value or '')

        self.assertIn("'Trial Balance'!", revenue_formula)
        self.assertRegex(revenue_formula, r"'Trial Balance'!\$[A-Z]+\$\d+")
        self.assertNotIn('SUMIFS(', revenue_formula)

        self.assertIn("'Trial Balance'!", director_salary_formula)
        self.assertRegex(director_salary_formula, r"'Trial Balance'!\$[A-Z]+\$\d+")
        self.assertNotIn('SUMIFS(', director_salary_formula)

        self.assertIn("'Trial Balance'!", other_income_formula)
        self.assertRegex(other_income_formula, r"'Trial Balance'!\$[A-Z]+\$\d+")
        self.assertIn('ABS(', other_income_formula)
        self.assertNotIn('SUMIFS(', other_income_formula)

    def test_30_sofp_cash_line_uses_direct_trial_balance_cell_link_formula(self):
        wizard = self._create_wizard(export_sheet_key='sofp')
        _action, wb = self._export_workbook(wizard)

        sofp = wb['SOFP']
        cash_row = self._find_row_by_label(sofp, 'Cash and bank balances')
        self.assertIsNotNone(cash_row)

        current_formula = str(sofp[f'B{cash_row}'].value or '')
        prior_formula = str(sofp[f'C{cash_row}'].value or '')

        for formula in (current_formula, prior_formula):
            self.assertIn("'Trial Balance'!", formula)
            self.assertRegex(formula, r"'Trial Balance'!\$[A-Z]+\$\d+")
            self.assertNotIn('SUMIFS(', formula)

    def test_31_tb_exact_balance_formula_prefers_direct_row_reference(self):
        wizard = self._create_wizard(export_sheet_key='sofp')
        tb_context = {
            'sheet_name': 'Trial Balance',
            'code_range_ref': "'Trial Balance'!$K$4:$K$20",
            'current_balance_range_ref': "'Trial Balance'!$L$4:$L$20",
            'prior_balance_range_ref': "'Trial Balance'!$M$4:$M$20",
            'rows': [
                {
                    'code': '12030101',
                    'code_raw': '12030101',
                    'current_balance_ref': "'Trial Balance'!$L$7",
                    'prior_balance_ref': "'Trial Balance'!$M$7",
                },
            ],
        }

        current_formula = wizard._tb_exact_balance_formula(tb_context, '12030101')
        prior_formula = wizard._tb_exact_balance_formula(tb_context, '12030101', is_prior=True)

        self.assertIn("'Trial Balance'!$L$7", current_formula)
        self.assertIn("'Trial Balance'!$M$7", prior_formula)
        self.assertNotIn('SUMIFS(', current_formula)
        self.assertNotIn('SUMIFS(', prior_formula)

    def test_32_tb_prefix_formula_uses_matching_row_ranges(self):
        wizard = self._create_wizard(export_sheet_key='sofp')
        tb_context = {
            'sheet_name': 'Trial Balance',
            'code_range_ref': "'Trial Balance'!$K$4:$K$50",
            'current_balance_range_ref': "'Trial Balance'!$L$4:$L$50",
            'prior_balance_range_ref': "'Trial Balance'!$M$4:$M$50",
            'rows': [
                {
                    'code': '12040101',
                    'row_index': 7,
                    'current_balance_ref': "'Trial Balance'!$L$7",
                    'prior_balance_ref': "'Trial Balance'!$M$7",
                },
                {
                    'code': '12040102',
                    'row_index': 8,
                    'current_balance_ref': "'Trial Balance'!$L$8",
                    'prior_balance_ref': "'Trial Balance'!$M$8",
                },
                {
                    'code': '12040103',
                    'row_index': 9,
                    'current_balance_ref': "'Trial Balance'!$L$9",
                    'prior_balance_ref': "'Trial Balance'!$M$9",
                },
            ],
        }

        current_formula = wizard._tb_prefix_balance_formula(tb_context, '120401')
        prior_formula = wizard._tb_prefix_balance_formula(tb_context, '120401', is_prior=True)

        self.assertIn("'Trial Balance'!$L$7:'Trial Balance'!$L$9", current_formula)
        self.assertIn("'Trial Balance'!$M$7:'Trial Balance'!$M$9", prior_formula)
        self.assertNotIn('SUMIFS(', current_formula)
        self.assertNotIn('SUMIFS(', prior_formula)

    def test_33_cash_bank_formula_avoids_sumifs_when_some_prefixes_missing(self):
        wizard = self._create_wizard(export_sheet_key='sofp')
        tb_context = {
            'sheet_name': 'Trial Balance',
            'code_range_ref': "'Trial Balance'!$K$4:$K$80",
            'current_balance_range_ref': "'Trial Balance'!$L$4:$L$80",
            'prior_balance_range_ref': "'Trial Balance'!$M$4:$M$80",
            'rows': [
                {
                    'code': '12040201',
                    'row_index': 11,
                    'current_balance_ref': "'Trial Balance'!$L$11",
                    'prior_balance_ref': "'Trial Balance'!$M$11",
                },
                {
                    'code': '12060101',
                    'row_index': 12,
                    'current_balance_ref': "'Trial Balance'!$L$12",
                    'prior_balance_ref': "'Trial Balance'!$M$12",
                },
            ],
        }

        current_formula = wizard._tb_sofp_cash_bank_formula(tb_context)
        prior_formula = wizard._tb_sofp_cash_bank_formula(tb_context, is_prior=True)

        self.assertIn("'Trial Balance'!$L$11", current_formula)
        self.assertIn("'Trial Balance'!$L$12", current_formula)
        self.assertIn("'Trial Balance'!$M$11", prior_formula)
        self.assertIn("'Trial Balance'!$M$12", prior_formula)
        self.assertNotIn('SUMIFS(', current_formula)
        self.assertNotIn('SUMIFS(', prior_formula)

    def test_34_socf_lines_use_sofp_formulas_for_current_period(self):
        wizard = self._create_wizard(export_sheet_key='socf')
        _action, wb = self._export_workbook(wizard)

        socf = wb['SOCF']

        change_assets_row = self._find_row_by_label(socf, '(Increase) / decrease in current assets') or self._find_row_by_label(
            socf, 'Increase / decrease in current assets'
        )
        change_liabilities_row = self._find_row_by_label(
            socf, 'Increase / (decrease) in current liabilities'
        ) or self._find_row_by_label(socf, 'Increase / decrease in current liabilities')
        property_row = self._find_row_by_label(socf, 'Property, plant and equipment')

        self.assertIsNotNone(change_assets_row)
        self.assertIsNotNone(change_liabilities_row)
        self.assertIsNotNone(property_row)

        for row_idx in (change_assets_row, change_liabilities_row, property_row):
            current_formula = str(socf[f'B{row_idx}'].value or '')
            self.assertIn('SOFP!', current_formula)
