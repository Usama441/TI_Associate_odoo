import base64
import datetime
import io
import json
import logging
import os
import re
from copy import copy
from dateutil.relativedelta import relativedelta

from odoo import _, api, fields, models
from odoo.exceptions import ValidationError

from .native_trial_balance_mixin import NativeTrialBalanceMixin
from ..utils import clean_bank_narration

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = load_workbook = None
    Comment = None
    Alignment = Border = Font = PatternFill = Side = None
    get_column_letter = None

_logger = logging.getLogger(__name__)


class AuditExcelExportWizard(NativeTrialBalanceMixin, models.TransientModel):
    _name = 'audit.excel.export.wizard'
    _description = 'Audit Excel Export Wizard'
    _SOCF_CT_EXPENSE_ACCOUNT_CODES = ('51270101',)
    _SOCF_CURRENT_LIABILITY_PREFIXES = (
        '220201',
        '220301',
        '220302',
        '220303',
        '220101',
        '220103',
    )
    _SOCF_EOSB_LIABILITY_CODE = '21040101'
    _SOCF_SECURITY_DEPOSIT_ACCOUNT_CODE = '11050102'
    _SOCF_RELATED_PARTY_LOAN_PREFIX = '2102'
    _LINKED_SHEET_DEPENDENCIES = {
        'Trial Balance': ('Client Details',),
        'SOCI': ('Client Details', 'Trial Balance'),
        'SOCE': ('Client Details', 'Trial Balance', 'SOCI'),
        'SOFP': ('Client Details', 'Trial Balance', 'SOCI', 'SOCE'),
        'SOCF': ('Client Details', 'Trial Balance', 'SOCI'),
        'Summary Sheet': ('Client Details', 'Trial Balance', 'SOCI', 'SOCE', 'SOFP'),
    }
    _WORKBOOK_SHEET_ORDER = (
        'Client Details',
        'General Ledger',
        'Trial Balance',
        'SOCI',
        'SOCE',
        'SOFP',
        'SOCF',
        'PPE',
        'Prepayment',
        'Bank Summary',
        'Customer Invoices',
        'Aged Receivables',
        'Vendor Bills',
        'Aged Payables',
        'VAT Control',
        'Accruals',
        'Summary Sheet',
        'Share Capital',
    )

    _DATA_SHEET_SEQUENCE = (
        'General Ledger',
        'Customer Invoices',
        'Vendor Bills',
        'Aged Receivables',
        'Aged Payables',
    )
    _TEMPLATE_SHEET_SOURCE_PLAN = (
        ('vat_control', 'VAT Control', 'VAT Control'),
        ('prepayment', 'Prepayment', 'Prepayment'),
        ('client_details', 'Client Details', 'Client Details'),
        ('summary_sheet', 'SUMMARY', 'Summary Sheet'),
        ('sofp', 'SoFP', 'SOFP'),
        ('soci', 'SoCI', 'SOCI'),
        ('soce', 'SoCE', 'SOCE'),
        ('socf', 'SoCF', 'SOCF'),
        ('share_capital', 'Share Capital', 'Share Capital'),
        ('accruals', 'Accruals', 'Accruals'),
    )
    _TRIAL_BALANCE_SHEET_PLAN = ('trial_balance', 'Trial Balance')
    _TEMPLATE_FILE_RELATIVE_PATH = ('data', 'afs_excel_template.xlsx')
    # Keep the export-all tab order aligned with the audit workbook sequence.
    # Existing generated tabs map to the legacy names as follows:
    # Sale -> Customer Invoices, SL Control -> Aged Receivables,
    # Purchase -> Vendor Bills, PL Control -> Aged Payables.
    # PPE and Bank Summary are currently exported as empty placeholders.
    _EXPORTABLE_SHEET_KEYS = (
        ('client_details', 'Client Details'),
        ('summary_sheet', 'Summary Sheet'),
        ('general_ledger', 'General Ledger'),
        ('trial_balance', 'Trial Balance'),
        ('sofp', 'SOFP'),
        ('soci', 'SOCI'),
        ('soce', 'SOCE'),
        ('socf', 'SOCF'),
        ('ppe', 'PPE'),
        ('prepayment', 'Prepayment'),
        ('customer_invoices', 'Sale'),
        ('aged_receivables', 'SL Control'),
        ('vendor_bills', 'Purchase'),
        ('aged_payables', 'PL Control'),
        ('bank_summary', 'Bank Summary'),
        ('share_capital', 'Share Capital'),
        ('accruals', 'Accruals'),
        ('vat_control', 'VAT Control'),
    )
    _EXPORT_SHEET_NAME_BY_KEY = {
        'client_details': 'Client Details',
        'summary_sheet': 'Summary Sheet',
        'general_ledger': 'General Ledger',
        'trial_balance': 'Trial Balance',
        'sofp': 'SOFP',
        'soci': 'SOCI',
        'soce': 'SOCE',
        'socf': 'SOCF',
        'ppe': 'PPE',
        'prepayment': 'Prepayment',
        # Keep dropdown labels as legacy names, but map to generated sheet names.
        'customer_invoices': 'Customer Invoices',
        'aged_receivables': 'Aged Receivables',
        'vendor_bills': 'Vendor Bills',
        'aged_payables': 'Aged Payables',
        'bank_summary': 'Bank Summary',
        'share_capital': 'Share Capital',
        'accruals': 'Accruals',
        'vat_control': 'VAT Control',
    }

    company_ids = fields.Many2many(
        'res.company',
        string='Companies',
        required=True,
        default=lambda self: self.env.company,
    )
    date_from = fields.Date(
        string='Date From',
        required=True,
        default=lambda self: fields.Date.context_today(self).replace(day=1),
    )
    date_to = fields.Date(
        string='Date To',
        required=True,
        default=lambda self: fields.Date.context_today(self),
    )
    year_span = fields.Selection(
        [
            ('1y', '1 Year'),
            ('2y', '2 Years (with prior year)'),
        ],
        string='Export Period',
        required=True,
        default='2y',
    )
    balance_sheet_date_mode = fields.Selection(
        [
            ('end_only', 'End date only (snapshot)'),
            ('range', 'Date range (start to end)'),
        ],
        string='Balance Sheet Dates',
        required=True,
        default='end_only',
    )
    prior_year_mode = fields.Selection(
        [
            ('auto', 'Auto (1 year back)'),
            ('manual', 'Manual selection'),
        ],
        string='Prior Year Dates',
        required=True,
        default='auto',
    )
    prior_balance_sheet_date_mode = fields.Selection(
        [
            ('end_only', 'End date only (snapshot)'),
            ('range', 'Date range (start to end)'),
        ],
        string='Prior Balance Sheet Dates',
        required=True,
        default='end_only',
    )
    prior_date_start = fields.Date(string='Prior Date Start')
    prior_date_end = fields.Date(string='Prior Date End')
    use_previous_settings = fields.Boolean(
        string='Use Previous Settings',
        default=True,
    )
    export_sheet_key = fields.Selection(
        [('all', 'All sheets')] + [(key, label) for key, label in _EXPORTABLE_SHEET_KEYS],
        string='Sheets To Export',
        required=True,
        default='all',
    )

 

    include_draft_entries = fields.Boolean(string='Include Draft Entries', default=False)
    unfold_all = fields.Boolean(string='Unfold All', default=True)
    hide_zero_lines = fields.Boolean(string='Hide Zero Lines', default=False)

    aging_based_on = fields.Selection(
        [
            ('base_on_invoice_date', 'Based on Invoice Date'),
            ('base_on_maturity_date', 'Based on Maturity Date'),
        ],
        string='Aging Based On',
        default='base_on_maturity_date',
        required=True,
    )
    aging_interval = fields.Integer(string='Aging Interval (Days)', default=30)
    show_currency = fields.Boolean(string='Show Currency Column', default=True)
    show_account = fields.Boolean(string='Show Account Column', default=True)

    include_dynamic_columns = fields.Boolean(string='Include Dynamic Columns', default=True)
    invoice_bill_scope = fields.Selection(
        [
            ('all_states', 'All States'),
            ('posted_only', 'Posted Only'),
            ('posted_cancelled', 'Posted + Cancelled'),
        ],
        string='Invoices/Bills State Scope',
        default='all_states',
        required=True,
    )
    include_refunds = fields.Boolean(string='Include Credit Notes/Refunds', default=True)

    gl_options_json = fields.Text(string='General Ledger Options Override (JSON)')
    aged_receivable_options_json = fields.Text(string='Aged Receivable Options Override (JSON)')
    aged_payable_options_json = fields.Text(string='Aged Payable Options Override (JSON)')

    file_name = fields.Char(string='File Name', readonly=True)
    file_data = fields.Binary(string='File', readonly=True)

    @api.constrains('date_from', 'date_to')
    def _check_date_range(self):
        for wizard in self:
            year_end_date = wizard._get_effective_year_end_date()
            if wizard.date_from and year_end_date and wizard.date_from > year_end_date:
                raise ValidationError(_('Date From must be earlier than or equal to Year End Date.'))

    @api.constrains('prior_year_mode', 'prior_date_start', 'prior_date_end')
    def _check_prior_date_range(self):
        for wizard in self:
            if wizard._is_one_year_export():
                continue
            if wizard.prior_year_mode != 'manual':
                continue
            if not wizard.prior_date_start or not wizard.prior_date_end:
                raise ValidationError(_('Prior Date Start and Prior Date End are required in manual mode.'))
            if wizard.prior_date_start > wizard.prior_date_end:
                raise ValidationError(_('Prior Date Start must be earlier than or equal to Prior Date End.'))

    @api.constrains('aging_interval')
    def _check_aging_interval(self):
        for wizard in self:
            if wizard.aging_interval <= 0:
                raise ValidationError(_('Aging Interval must be greater than zero.'))

    @staticmethod
    def _normalize_year_end_vals(vals):
        normalized_vals = dict(vals or {})
        if 'aged_as_of_date' in normalized_vals and 'date_to' not in normalized_vals:
            normalized_vals['date_to'] = normalized_vals.get('aged_as_of_date')
        normalized_vals.pop('aged_as_of_date', None)
        normalized_vals.pop('journal_ids', None)
        normalized_vals.pop('partner_ids', None)
        normalized_vals.pop('analytic_account_ids', None)

        legacy_period_category = normalized_vals.pop('audit_period_category', None)
        if 'year_span' not in normalized_vals:
            mapped_year_span = AuditExcelExportWizard._legacy_period_category_to_year_span(legacy_period_category)
            if mapped_year_span:
                normalized_vals['year_span'] = mapped_year_span
        return normalized_vals

    @api.model_create_multi
    def create(self, vals_list):
        normalized_vals_list = [self._normalize_year_end_vals(vals) for vals in vals_list]
        return super().create(normalized_vals_list)

    def write(self, vals):
        normalized_vals = self._normalize_year_end_vals(vals)
        return super().write(normalized_vals)

    def _get_effective_year_end_date(self):
        self.ensure_one()
        return fields.Date.to_date(self.date_to) if self.date_to else False

    def _previous_settings_key(self):
        return (
            f'audit_excel_export_wizard.prev_settings.user_{self.env.user.id}.'
            f'company_{self.env.company.id}'
        )

    def _get_previous_settings(self):
        raw = self.env['ir.config_parameter'].sudo().get_param(self._previous_settings_key(), default='')
        if not raw:
            return {}
        try:
            data = json.loads(raw)
        except (TypeError, ValueError):
            return {}
        return data if isinstance(data, dict) else {}

    def _selection_values(self, field_name):
        selection = self._fields[field_name].selection or []
        return {value for value, _label in selection}

    @staticmethod
    def _legacy_period_category_to_year_span(value):
        category = (value or '').strip().lower()
        if category.endswith('_1y'):
            return '1y'
        if category.endswith('_2y'):
            return '2y'
        return False

    @staticmethod
    def _safe_date_from_settings(value):
        if not value:
            return False
        try:
            return fields.Date.to_date(value)
        except Exception:
            return False

    def _store_previous_settings(self):
        self.ensure_one()
        data = {
            'company_ids': self.company_ids.ids,
            'date_from': self.date_from.isoformat() if self.date_from else False,
            'date_to': self.date_to.isoformat() if self.date_to else False,
            'year_span': self.year_span,
            'balance_sheet_date_mode': self.balance_sheet_date_mode,
            'prior_year_mode': self.prior_year_mode,
            'prior_balance_sheet_date_mode': self.prior_balance_sheet_date_mode,
            'export_sheet_key': self.export_sheet_key,
            'prior_date_start': self.prior_date_start.isoformat() if self.prior_date_start else False,
            'prior_date_end': self.prior_date_end.isoformat() if self.prior_date_end else False,
            'include_draft_entries': self.include_draft_entries,
            'unfold_all': self.unfold_all,
            'hide_zero_lines': self.hide_zero_lines,
            'aging_based_on': self.aging_based_on,
            'aging_interval': self.aging_interval,
            'show_currency': self.show_currency,
            'show_account': self.show_account,
            'include_dynamic_columns': self.include_dynamic_columns,
            'invoice_bill_scope': self.invoice_bill_scope,
            'include_refunds': self.include_refunds,
            'gl_options_json': self.gl_options_json or '',
            'aged_receivable_options_json': self.aged_receivable_options_json or '',
            'aged_payable_options_json': self.aged_payable_options_json or '',
        }
        self.env['ir.config_parameter'].sudo().set_param(self._previous_settings_key(), json.dumps(data))

    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        if not res.get('use_previous_settings', True):
            return res

        data = self._get_previous_settings()
        if not data:
            return res

        if 'year_span' in fields_list:
            year_span = data.get('year_span')
            if year_span not in self._selection_values('year_span'):
                year_span = self._legacy_period_category_to_year_span(data.get('audit_period_category'))
            if year_span in self._selection_values('year_span'):
                res['year_span'] = year_span

        selection_fields = (
            'balance_sheet_date_mode',
            'prior_year_mode',
            'prior_balance_sheet_date_mode',
            'aging_based_on',
            'invoice_bill_scope',
            'export_sheet_key',
        )
        for field_name in selection_fields:
            if field_name in fields_list and data.get(field_name) in self._selection_values(field_name):
                res[field_name] = data[field_name]

        date_fields = (
            'date_from',
            'date_to',
            'prior_date_start',
            'prior_date_end',
        )
        for field_name in date_fields:
            if field_name in fields_list and data.get(field_name):
                date_value = self._safe_date_from_settings(data.get(field_name))
                if date_value:
                    res[field_name] = date_value

        m2m_fields = (
            'company_ids',
        )
        for field_name in m2m_fields:
            if field_name in fields_list and isinstance(data.get(field_name), list):
                ids = [int(record_id) for record_id in data[field_name] if isinstance(record_id, int)]
                res[field_name] = [(6, 0, ids)]

        bool_int_fields = (
            'include_draft_entries',
            'unfold_all',
            'hide_zero_lines',
            'aging_interval',
            'show_currency',
            'show_account',
            'include_dynamic_columns',
            'include_refunds',
        )
        for field_name in bool_int_fields:
            if field_name in fields_list and field_name in data:
                res[field_name] = data[field_name]

        text_fields = (
            'gl_options_json',
            'aged_receivable_options_json',
            'aged_payable_options_json',
        )
        for field_name in text_fields:
            if field_name in fields_list and field_name in data:
                res[field_name] = data.get(field_name) or False

        return res

    def _apply_previous_settings(self, data):
        self.ensure_one()

        year_span = data.get('year_span')
        if year_span not in self._selection_values('year_span'):
            year_span = self._legacy_period_category_to_year_span(data.get('audit_period_category'))
        if year_span in self._selection_values('year_span'):
            self.year_span = year_span

        selection_fields = (
            'balance_sheet_date_mode',
            'prior_year_mode',
            'prior_balance_sheet_date_mode',
            'aging_based_on',
            'invoice_bill_scope',
            'export_sheet_key',
        )
        for field_name in selection_fields:
            value = data.get(field_name)
            if value in self._selection_values(field_name):
                setattr(self, field_name, value)

        date_fields = (
            'date_from',
            'date_to',
            'prior_date_start',
            'prior_date_end',
        )
        for field_name in date_fields:
            date_value = self._safe_date_from_settings(data.get(field_name))
            if date_value:
                setattr(self, field_name, date_value)

        m2m_models = (
            ('company_ids', 'res.company'),
        )
        for field_name, model_name in m2m_models:
            values = data.get(field_name)
            if isinstance(values, list):
                ids = [int(record_id) for record_id in values if isinstance(record_id, int)]
                setattr(self, field_name, self.env[model_name].browse(ids))

        bool_int_fields = (
            'include_draft_entries',
            'unfold_all',
            'hide_zero_lines',
            'aging_interval',
            'show_currency',
            'show_account',
            'include_dynamic_columns',
            'include_refunds',
        )
        for field_name in bool_int_fields:
            if field_name in data:
                setattr(self, field_name, data[field_name])

        text_fields = (
            'gl_options_json',
            'aged_receivable_options_json',
            'aged_payable_options_json',
        )
        for field_name in text_fields:
            if field_name in data:
                setattr(self, field_name, data.get(field_name) or False)

    @api.onchange('use_previous_settings')
    def _onchange_use_previous_settings(self):
        if not self.use_previous_settings:
            return
        data = self._get_previous_settings()
        if data:
            self._apply_previous_settings(data)

    def action_export_xlsx(self):
        self.ensure_one()
        self._ensure_openpyxl_available()
        year_end_date = self._get_effective_year_end_date()
        if not year_end_date:
            raise ValidationError(_('Year End Date is required.'))
        if self.date_from > year_end_date:
            raise ValidationError(_('Date From must be earlier than or equal to Year End Date.'))

        gl_overrides = self._parse_json_options(self.gl_options_json, _('General Ledger Options Override'))
        ar_overrides = self._parse_json_options(self.aged_receivable_options_json, _('Aged Receivable Options Override'))
        ap_overrides = self._parse_json_options(self.aged_payable_options_json, _('Aged Payable Options Override'))
        requested_sheet_names = self._get_selected_sheet_names()
        selected_sheet_names = self._expand_selected_sheet_dependencies(requested_sheet_names)
        include_trial_balance = 'Trial Balance' in set(selected_sheet_names)

        # Stage A: build data payloads for generated data sheets + trial balance rows.
        stage_a_payload = self._stage_a_build_data_payload(
            gl_overrides=gl_overrides,
            ar_overrides=ar_overrides,
            ap_overrides=ap_overrides,
            selected_sheet_names=selected_sheet_names,
        )

        # Stage B: load the formatted workbook template and write first five generated data sheets.
        stage_b_context = self._stage_b_create_workbook_and_write_data_sheets(stage_a_payload)

        # Stage C: append Trial Balance sheet after template-backed tabs.
        stage_c_context = self._stage_c_create_template_sheets(
            stage_b_context,
            stage_a_payload['trial_balance'],
            include_trial_balance=include_trial_balance,
        )

        # Stage D: remove static/sample values from template sheets except formulas and TB tab.
        self._stage_d_clean_template_inputs(stage_c_context, selected_sheet_names=selected_sheet_names)

        # Stage E: inject real trial balance figures only on the mapped TB sheet.
        self._stage_e_inject_trial_balance(stage_c_context, stage_a_payload['trial_balance'])

        # Stage F: populate template tabs with wizard/company context values.
        self._stage_f_populate_template_context(stage_c_context)

        # Stage G: build the live-link workbook graph from Trial Balance helper formulas.
        self._stage_g_link_statement_sheets_to_trial_balance(stage_c_context, stage_a_payload['trial_balance'])
        self._stage_g_hide_prior_year_live_link_columns(stage_c_context, selected_sheet_names)

        # Stage H: keep only requested sheet(s) in the final workbook.
        self._stage_h_filter_workbook_sheets(stage_c_context, selected_sheet_names)

        if self.use_previous_settings:
            self._store_previous_settings()

        # Stage I: finalize binary and keep existing wizard download UX.
        return self._stage_g_finalize_binary_download(
            stage_c_context['workbook'],
            selected_sheet_names=requested_sheet_names,
        )

    @classmethod
    def _get_export_sheet_name_map(cls):
        sheet_name_map = {}
        for key, _label in cls._EXPORTABLE_SHEET_KEYS:
            sheet_name = cls._EXPORT_SHEET_NAME_BY_KEY.get(key)
            if sheet_name:
                sheet_name_map[key] = sheet_name
        return sheet_name_map

    def _get_selected_sheet_names(self):
        self.ensure_one()
        sheet_name_map = self._get_export_sheet_name_map()
        if self.export_sheet_key == 'all':
            sheet_names = [sheet_name_map[key] for key, _label in self._EXPORTABLE_SHEET_KEYS if key in sheet_name_map]
            if not sheet_names:
                raise ValidationError(_('No exportable sheets are configured.'))
            return self._order_sheet_names_for_workbook(sheet_names)
        sheet_name = sheet_name_map.get(self.export_sheet_key)
        if not sheet_name:
            raise ValidationError(_('Please select a valid sheet to export.'))
        return [sheet_name]

    def _expand_selected_sheet_dependencies(self, selected_sheet_names):
        expanded = set(selected_sheet_names or [])
        queue = list(selected_sheet_names or [])
        while queue:
            sheet_name = queue.pop(0)
            for dependency in self._LINKED_SHEET_DEPENDENCIES.get(sheet_name, ()):
                if dependency in expanded:
                    continue
                expanded.add(dependency)
                queue.append(dependency)
        return self._order_sheet_names_for_workbook(expanded)

    @classmethod
    def _order_sheet_names_for_workbook(cls, sheet_names):
        remaining = list(dict.fromkeys(sheet_names or []))
        ordered = [name for name in cls._WORKBOOK_SHEET_ORDER if name in remaining]
        ordered.extend(name for name in remaining if name not in ordered)
        return ordered

    def _parse_json_options(self, raw_value, field_label):
        if not raw_value:
            return {}

        try:
            parsed = json.loads(raw_value)
        except (TypeError, ValueError) as exc:
            raise ValidationError(_('%(field)s must contain valid JSON. Error: %(error)s', field=field_label, error=str(exc))) from exc

        if not isinstance(parsed, dict):
            raise ValidationError(_('%(field)s must be a JSON object.', field=field_label))

        return parsed

    def _is_one_year_export(self):
        return (self.year_span or '2y') == '1y'

    def _get_reporting_periods(self):
        self.ensure_one()
        date_start = fields.Date.to_date(self.date_from) if self.date_from else False
        date_end = self._get_effective_year_end_date()
        year_span = self.year_span or '2y'
        show_prior_year = year_span == '2y'
        prior_year_mode = self.prior_year_mode or 'auto'
        prior_balance_sheet_date_mode = self.prior_balance_sheet_date_mode or 'end_only'

        prior_date_start = False
        prior_date_end = False
        if show_prior_year:
            if prior_year_mode == 'manual' and self.prior_date_start and self.prior_date_end:
                prior_date_start = fields.Date.to_date(self.prior_date_start)
                prior_date_end = fields.Date.to_date(self.prior_date_end)
            else:
                prior_date_start = date_start - relativedelta(years=1) if date_start else False
                prior_date_end = date_end - relativedelta(years=1) if date_end else False

            if prior_balance_sheet_date_mode == 'end_only' and prior_date_end:
                prior_date_start = prior_date_end - relativedelta(years=1) + relativedelta(days=1)

        opening_date_end = date_start - relativedelta(days=1) if date_start else False
        prior_opening_date_end = prior_date_start - relativedelta(days=1) if prior_date_start else False

        balance_sheet_date_start = date_start if self.balance_sheet_date_mode == 'range' else False
        prior_balance_sheet_date_start = (
            prior_date_start if prior_balance_sheet_date_mode == 'range' else False
        )

        return {
            'date_start': date_start,
            'date_end': date_end,
            'year_span': year_span,
            'show_prior_year': show_prior_year,
            'prior_date_start': prior_date_start,
            'prior_date_end': prior_date_end,
            'opening_date_end': opening_date_end,
            'prior_opening_date_end': prior_opening_date_end,
            'balance_sheet_date_start': balance_sheet_date_start,
            'prior_balance_sheet_date_start': prior_balance_sheet_date_start,
            'balance_sheet_date_mode': self.balance_sheet_date_mode,
            'prior_balance_sheet_date_mode': prior_balance_sheet_date_mode,
        }

    def _stage_a_build_data_payload(self, *, gl_overrides, ar_overrides, ap_overrides, selected_sheet_names):
        selected_sheet_set = set(selected_sheet_names or [])
        data_sheets = []
        if 'General Ledger' in selected_sheet_set:
            data_sheets.append(self._prepare_general_ledger_payload(gl_overrides))
        if 'Customer Invoices' in selected_sheet_set:
            data_sheets.append(self._prepare_invoice_bill_payload(is_customer=True))
        if 'Vendor Bills' in selected_sheet_set:
            data_sheets.append(self._prepare_invoice_bill_payload(is_customer=False))
        if 'Aged Receivables' in selected_sheet_set:
            data_sheets.append(
                self._prepare_aged_payload(
                    xmlid='account_reports.aged_receivable_report',
                    sheet_name='Aged Receivables',
                    title='Aged Receivables',
                    overrides=ar_overrides,
                )
            )
        if 'Aged Payables' in selected_sheet_set:
            data_sheets.append(
                self._prepare_aged_payload(
                    xmlid='account_reports.aged_payable_report',
                    sheet_name='Aged Payables',
                    title='Aged Payables',
                    overrides=ap_overrides,
                )
            )
        if 'PPE' in selected_sheet_set:
            data_sheets.append(self._prepare_empty_sheet_payload('PPE'))
        if 'Bank Summary' in selected_sheet_set:
            data_sheets.append(self._prepare_empty_sheet_payload('Bank Summary'))

        trial_balance_payload = {}
        if 'Trial Balance' in selected_sheet_set:
            trial_balance_payload = self._prepare_trial_balance_payload()

        return {
            'data_sheets': data_sheets,
            'trial_balance': trial_balance_payload,
        }

    def _stage_b_create_workbook_and_write_data_sheets(self, stage_a_payload):
        self._ensure_openpyxl_available()

        workbook, template_sheet_map = self._load_formatted_template_workbook()

        used_sheet_names = set(workbook.sheetnames)
        for insert_index, payload in enumerate(stage_a_payload['data_sheets']):
            if payload.get('sheet_mode') == 'empty':
                self._write_empty_sheet(
                    workbook,
                    used_sheet_names,
                    insert_index=insert_index,
                    sheet_name=payload.get('sheet_name') or _('Sheet'),
                )
                continue

            if payload.get('sheet_mode') == 'invoice_bill_blocks':
                self._write_invoice_bill_blocks_sheet(
                    workbook,
                    used_sheet_names,
                    insert_index=insert_index,
                    sheet_name=payload.get('sheet_name') or _('Sheet'),
                    title=payload.get('title') or payload.get('sheet_name') or _('Sheet'),
                    summary_headers=payload.get('summary_headers') or [],
                    line_headers=payload.get('line_headers') or [],
                    blocks=payload.get('blocks') or [],
                )
                continue

            if payload.get('sheet_mode') == 'invoice_bill_flat':
                self._write_invoice_bill_flat_sheet(
                    workbook,
                    used_sheet_names,
                    insert_index=insert_index,
                    sheet_name=payload.get('sheet_name') or _('Sheet'),
                    title=payload.get('title') or payload.get('sheet_name') or _('Sheet'),
                    headers=payload.get('headers') or [],
                    rows=payload.get('rows') or [],
                    column_types=payload.get('column_types') or {},
                )
                continue

            native_source_sheet = payload.get('native_source_sheet')
            if native_source_sheet is not None:
                sheet_title = self._write_native_sheet_copy(
                    workbook,
                    used_sheet_names,
                    insert_index=insert_index,
                    sheet_name=payload.get('sheet_name') or native_source_sheet.title or 'Sheet',
                    source_sheet=native_source_sheet,
                )
                if (
                    sheet_title == 'General Ledger'
                    and self._is_gl_narration_cleaner_enabled()
                    and sheet_title in workbook.sheetnames
                ):
                    self._clean_general_ledger_sheet_narrations(workbook[sheet_title])
                continue

            self._write_plain_sheet(
                workbook,
                used_sheet_names,
                insert_index=insert_index,
                **payload,
            )

        return {
            'workbook': workbook,
            'used_sheet_names': used_sheet_names,
            'template_sheet_map': template_sheet_map,
            'data_sheet_count': len(stage_a_payload.get('data_sheets') or []),
        }

    def _stage_c_create_template_sheets(self, stage_b_context, trial_balance_payload, include_trial_balance=True):
        workbook = stage_b_context['workbook']
        used_sheet_names = stage_b_context['used_sheet_names']

        template_sheet_map = list(stage_b_context.get('template_sheet_map', []))
        stage_b_context['template_sheet_map'] = template_sheet_map

        if not include_trial_balance:
            stage_b_context['trial_balance_sheet'] = False
            stage_b_context['trial_balance_prepopulated'] = False
            return stage_b_context

        _logical_key, requested_sheet_name = self._TRIAL_BALANCE_SHEET_PLAN
        insert_index = stage_b_context.get('data_sheet_count', 0) + len(template_sheet_map)
        native_source_sheet = trial_balance_payload.get('native_source_sheet') if trial_balance_payload else None
        if native_source_sheet is not None:
            trial_balance_sheet_name = self._write_native_sheet_copy(
                workbook,
                used_sheet_names,
                insert_index=insert_index,
                sheet_name=trial_balance_payload.get('sheet_name') or requested_sheet_name,
                source_sheet=native_source_sheet,
            )
            trial_balance_prepopulated = True
        else:
            trial_balance_sheet_name = self._get_unique_sheet_name(requested_sheet_name, used_sheet_names)
            sheet = workbook.create_sheet(title=trial_balance_sheet_name, index=insert_index)
            self._build_template_trial_balance_sheet(sheet)
            trial_balance_prepopulated = False

        stage_b_context['trial_balance_sheet'] = trial_balance_sheet_name
        stage_b_context['trial_balance_prepopulated'] = trial_balance_prepopulated
        return stage_b_context

    def _stage_d_clean_template_inputs(self, stage_c_context, selected_sheet_names=None):
        workbook = stage_c_context['workbook']
        selected_sheet_set = set(selected_sheet_names or [])

        for item in stage_c_context['template_sheet_map']:
            sheet_name = item['sheet_name']
            if selected_sheet_set and sheet_name not in selected_sheet_set:
                continue
            self._cleanup_template_sheet_inputs(workbook[sheet_name])

    def _stage_e_inject_trial_balance(self, stage_c_context, trial_balance_payload):
        if not stage_c_context.get('trial_balance_sheet'):
            return
        if stage_c_context.get('trial_balance_prepopulated'):
            return
        workbook = stage_c_context['workbook']
        sheet = workbook[stage_c_context['trial_balance_sheet']]
        self._write_trial_balance_to_template_sheet(sheet, trial_balance_payload)

    def _stage_f_populate_template_context(self, stage_c_context):
        workbook = stage_c_context['workbook']
        primary_company = self._get_primary_company_for_template()

        client_details_sheet = self._get_sheet_if_exists(workbook, 'Client Details')
        if client_details_sheet is not None:
            self._populate_client_details_sheet(client_details_sheet, primary_company)

        share_capital_sheet = self._get_sheet_if_exists(workbook, 'Share Capital')
        if share_capital_sheet is not None and primary_company:
            self._populate_share_capital_sheet(share_capital_sheet, primary_company)

    def _is_gl_narration_cleaner_enabled(self):
        raw_value = (
            self.env['ir.config_parameter']
            .sudo()
            .get_param('audit_excel_export.gl_narration_cleaner_enabled', default='')
        )
        if raw_value is None:
            return True
        value = str(raw_value).strip().lower()
        if not value:
            return True
        return value not in {'0', 'false', 'no', 'off'}

    def _find_general_ledger_header_row(self, sheet):
        if sheet is None:
            return None
        max_scan_rows = min(sheet.max_row, 80)
        for row_idx in range(1, max_scan_rows + 1):
            code_header = (sheet.cell(row=row_idx, column=1).value or '')
            name_header = (sheet.cell(row=row_idx, column=2).value or '')
            if str(code_header).strip().lower() == 'code' and str(name_header).strip().lower() == 'account name':
                return row_idx
        return None

    def _is_cleanable_gl_detail_row(self, code_value, name_value):
        if code_value not in (None, '') and str(code_value).strip():
            return False
        label = str(name_value or '').strip()
        if not label:
            return False
        label_lower = label.lower()
        if label_lower == 'initial balance':
            return False
        if label_lower.startswith('total'):
            return False
        if label_lower == 'load more...':
            return False
        return True

    def _clean_general_ledger_sheet_narrations(self, sheet):
        if sheet is None:
            return
        header_row = self._find_general_ledger_header_row(sheet)
        if not header_row:
            return

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            code_cell = sheet.cell(row=row_idx, column=1)
            name_cell = sheet.cell(row=row_idx, column=2)
            if not self._is_cleanable_gl_detail_row(code_cell.value, name_cell.value):
                continue

            raw_narration = str(name_cell.value or '').strip()
            if not raw_narration:
                continue

            try:
                parsed = clean_bank_narration(raw_narration, max_len=100)
            except Exception:
                _logger.warning(
                    "General Ledger narration cleaning failed on row %s for value: %r",
                    row_idx,
                    raw_narration,
                    exc_info=True,
                )
                continue

            clean_value = (parsed.get('clean') or '').strip()
            if not clean_value:
                continue

            name_cell.value = clean_value
            if Comment:
                name_cell.comment = Comment(f"Raw narration: {raw_narration}", "Narration Cleaner")

    def _stage_h_filter_workbook_sheets(self, stage_c_context, selected_sheet_names):
        workbook = stage_c_context['workbook']
        selected_sheet_set = set(selected_sheet_names or [])
        if not selected_sheet_set:
            raise ValidationError(_('Please select at least one sheet to export.'))

        for sheet_name in list(workbook.sheetnames):
            if sheet_name not in selected_sheet_set:
                workbook.remove(workbook[sheet_name])

        remaining_selected = [sheet_name for sheet_name in selected_sheet_names if sheet_name in workbook.sheetnames]
        if not remaining_selected:
            raise ValidationError(_('None of the selected sheets were generated for export.'))
        self._reorder_workbook_sheets(workbook, self._order_sheet_names_for_workbook(remaining_selected))

    @staticmethod
    def _sheet_name_to_filename_token(sheet_name):
        token = re.sub(r'[^A-Za-z0-9]+', '_', (sheet_name or '').strip().lower()).strip('_')
        return token or 'sheet'

    def _stage_g_finalize_binary_download(self, workbook, selected_sheet_names=None):
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        file_bytes = output.read()
        output.close()
        year_end_date = self._get_effective_year_end_date()
        year_end_date_string = fields.Date.to_string(year_end_date) if year_end_date else ''

        if selected_sheet_names and len(selected_sheet_names) == 1:
            sheet_token = self._sheet_name_to_filename_token(selected_sheet_names[0])
            filename = (
                f"audit_export_{sheet_token}_{fields.Date.to_string(self.date_from)}_"
                f"{year_end_date_string}.xlsx"
            )
        else:
            filename = f"audit_export_{fields.Date.to_string(self.date_from)}_{year_end_date_string}.xlsx"
        self.write({
            'file_name': filename,
            'file_data': base64.b64encode(file_bytes),
        })

        return {
            'type': 'ir.actions.act_url',
            'url': (
                f"/web/content/?model={self._name}&id={self.id}"
                "&field=file_data&filename_field=file_name&download=true"
            ),
            'target': 'self',
        }

    def _get_sheet_if_exists(self, workbook, sheet_name):
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]
        return None

    def _get_primary_company_for_template(self):
        companies = self.company_ids.sorted('id')
        if companies:
            return companies[0]
        return self.env.company

    def _populate_client_details_sheet(self, sheet, company):
        company_names = ', '.join(self.company_ids.mapped('name')) if self.company_ids else (company.name or '')
        shareholders = self._collect_company_shareholders(company)
        year_end_date = self._get_effective_year_end_date()

        self._write_cell_value(sheet, 'B4', company_names)
        self._write_cell_value(sheet, 'B5', self.date_from)
        self._write_cell_value(sheet, 'B6', year_end_date)

        # Additional company profile fields are only meaningful in single-company export.
        if len(self.company_ids) == 1 and company:
            self._write_cell_value(sheet, 'B8', getattr(company, 'company_license_number', '') or '')
            self._write_cell_value(sheet, 'B9', getattr(company, 'free_zone', '') or '')
            self._write_cell_value(sheet, 'B10', getattr(company, 'incorporation_date', False) or None)
            self._write_cell_value(sheet, 'B12', getattr(company, 'trade_license_activities', '') or '')
            self._write_cell_value(
                sheet,
                'B15',
                getattr(company, 'corporate_tax_registration_number', '') or '',
            )
            self._write_cell_value(sheet, 'B16', getattr(company, 'vat_registration_number', '') or '')
        else:
            self._write_cell_value(sheet, 'B8', '')
            self._write_cell_value(sheet, 'B9', '')
            self._write_cell_value(sheet, 'B10', None)
            self._write_cell_value(sheet, 'B12', '')
            self._write_cell_value(sheet, 'B15', '')
            self._write_cell_value(sheet, 'B16', '')

        self._write_cell_value(sheet, 'B13', ', '.join(shareholders) if shareholders else '')

        for coord in ('B5', 'B6', 'B10'):
            cell = sheet[coord]
            if cell.value:
                cell.number_format = 'yyyy-mm-dd'

    def _populate_share_capital_sheet(self, sheet, company):
        shareholders = self._collect_company_shareholder_rows(company)
        for row_offset in range(4):
            row_no = 7 + row_offset
            row_data = shareholders[row_offset] if row_offset < len(shareholders) else None
            if not row_data:
                self._write_cell_value(sheet, f'A{row_no}', '')
                self._write_cell_value(sheet, f'B{row_no}', None)
                self._write_cell_value(sheet, f'C{row_no}', None)
                continue

            self._write_cell_value(sheet, f'A{row_no}', row_data['name'])
            self._write_cell_value(sheet, f'B{row_no}', row_data['shares'])
            self._write_cell_value(sheet, f'C{row_no}', row_data['share_value'])

            sheet[f'B{row_no}'].number_format = '#,##0'
            sheet[f'C{row_no}'].number_format = '#,##0.00'

    def _collect_company_shareholders(self, company):
        result = []
        if not company:
            return result
        for idx in range(1, 11):
            name = (getattr(company, f'shareholder_{idx}', '') or '').strip()
            if name:
                result.append(name)
        return result

    def _collect_company_shareholder_rows(self, company):
        rows = []
        if not company:
            return rows

        for idx in range(1, 11):
            name = (getattr(company, f'shareholder_{idx}', '') or '').strip()
            shares = getattr(company, f'number_of_shares_{idx}', 0) or 0
            share_value = getattr(company, f'share_value_{idx}', 0.0) or 0.0
            if not name and not shares and not share_value:
                continue
            rows.append({
                'name': name,
                'shares': int(shares) if shares else None,
                'share_value': float(share_value) if share_value else None,
            })
        return rows

    def _write_cell_value(self, sheet, coordinate, value):
        sheet[coordinate].value = value

    def _ensure_openpyxl_available(self):
        if not (load_workbook and Workbook):
            raise ValidationError(_('Python package "openpyxl" is required for this export.'))

    def _resolve_template_workbook_path(self):
        module_root = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
        candidate = os.path.abspath(os.path.join(module_root, *self._TEMPLATE_FILE_RELATIVE_PATH))
        if os.path.isfile(candidate):
            return candidate
        raise ValidationError(
            _('Template workbook file was not found at "%(path)s".', path=candidate)
        )

    def _build_formatted_template_workbook(self):
        """Build the template-backed workbook structure directly in code.

        The legacy XLSX file remains on disk as a visual reference, but export
        generation should not depend on loading workbook-level sample data,
        defined names, or stale links from that file at runtime.
        """
        builder_map = self._get_template_sheet_builder_map()
        workbook = Workbook()
        template_sheet_map = []

        for index, (logical_key, _source_sheet_name, target_sheet_name) in enumerate(self._TEMPLATE_SHEET_SOURCE_PLAN):
            if logical_key not in builder_map:
                raise ValidationError(
                    _('No template sheet builder is configured for "%(sheet)s".', sheet=target_sheet_name)
                )

            if index == 0:
                sheet = workbook.active
                sheet.title = target_sheet_name
            else:
                sheet = workbook.create_sheet(title=target_sheet_name)

            builder_map[logical_key](sheet)
            template_sheet_map.append({
                'logical_key': logical_key,
                'sheet_name': target_sheet_name,
            })

        self._reorder_workbook_sheets(workbook, [item['sheet_name'] for item in template_sheet_map])

        calculation = getattr(workbook, 'calculation', None)
        if calculation is not None:
            calculation.fullCalcOnLoad = True

        return workbook, template_sheet_map

    def _get_template_sheet_builder_map(self):
        return {
            'vat_control': self._build_template_vat_control_sheet,
            'prepayment': self._build_template_prepayment_sheet,
            'client_details': self._build_template_client_details_sheet,
            'summary_sheet': self._build_template_summary_sheet,
            'sofp': self._build_template_sofp_sheet,
            'soci': self._build_template_soci_sheet,
            'soce': self._build_template_soce_sheet,
            'socf': self._build_template_socf_sheet,
            'share_capital': self._build_template_share_capital_sheet,
            'accruals': self._build_template_accruals_sheet,
        }

    def _load_formatted_template_workbook(self):
        template_path = self._resolve_template_workbook_path()
        try:
            workbook = load_workbook(template_path, data_only=False, keep_links=False)
        except Exception as exc:
            error_message = str(exc) or repr(exc) or exc.__class__.__name__
            raise ValidationError(
                _('Unable to load template workbook "%(path)s": %(error)s', path=template_path, error=error_message)
            ) from exc

        source_sheet_names = [source_sheet_name for _key, source_sheet_name, _target in self._TEMPLATE_SHEET_SOURCE_PLAN]
        missing_sheets = [name for name in source_sheet_names if name not in workbook.sheetnames]
        if missing_sheets:
            raise ValidationError(
                _(
                    'Template workbook "%(path)s" is missing required sheets: %(sheets)s',
                    path=template_path,
                    sheets=', '.join(missing_sheets),
                )
            )

        for sheet in list(workbook.worksheets):
            if sheet.title not in source_sheet_names:
                workbook.remove(sheet)

        rename_map = {
            source_sheet_name: target_sheet_name
            for _logical_key, source_sheet_name, target_sheet_name in self._TEMPLATE_SHEET_SOURCE_PLAN
            if source_sheet_name != target_sheet_name
        }
        temporary_names = {}
        for source_sheet_name, target_sheet_name in rename_map.items():
            if source_sheet_name.lower() == target_sheet_name.lower():
                temporary_name = '__tmp_%s__' % source_sheet_name
                while temporary_name in workbook.sheetnames:
                    temporary_name = '%s_' % temporary_name
                workbook[source_sheet_name].title = temporary_name
                temporary_names[source_sheet_name] = temporary_name

        for source_sheet_name, target_sheet_name in rename_map.items():
            current_sheet_name = temporary_names.get(source_sheet_name, source_sheet_name)
            workbook[current_sheet_name].title = target_sheet_name

        template_sheet_map = []
        for logical_key, _source_sheet_name, target_sheet_name in self._TEMPLATE_SHEET_SOURCE_PLAN:
            template_sheet_map.append({
                'logical_key': logical_key,
                'sheet_name': target_sheet_name,
            })

        self._rewrite_workbook_sheet_references(workbook, rename_map)
        self._reorder_workbook_sheets(workbook, [item['sheet_name'] for item in template_sheet_map])
        return workbook, template_sheet_map

    def _rewrite_workbook_sheet_references(self, workbook, rename_map):
        if not rename_map:
            return

        for sheet in workbook.worksheets:
            for cell in sheet._cells.values():
                value = cell.value
                if isinstance(value, str) and value.startswith('='):
                    cell.value = self._replace_formula_sheet_references(value, rename_map)

        defined_names = getattr(workbook.defined_names, 'definedName', [])
        for defined_name in defined_names:
            attr_text = getattr(defined_name, 'attr_text', None)
            if isinstance(attr_text, str):
                defined_name.attr_text = self._replace_formula_sheet_references(attr_text, rename_map)

    def _replace_formula_sheet_references(self, expression, rename_map):
        updated_expression = expression
        for old_name, new_name in rename_map.items():
            old_quoted = "'%s'!" % old_name.replace("'", "''")
            new_quoted = "'%s'!" % new_name.replace("'", "''")
            updated_expression = updated_expression.replace(old_quoted, new_quoted)

            if self._sheet_name_is_unquoted_token(old_name) and self._sheet_name_is_unquoted_token(new_name):
                pattern = re.compile(r'(?<![A-Za-z0-9_.])%s!' % re.escape(old_name))
                updated_expression = pattern.sub('%s!' % new_name, updated_expression)
        return updated_expression

    def _sheet_name_is_unquoted_token(self, sheet_name):
        return bool(re.match(r'^[A-Za-z_][A-Za-z0-9_.]*$', sheet_name or ''))

    def _reorder_workbook_sheets(self, workbook, ordered_sheet_names):
        ordered_sheets = [workbook[name] for name in ordered_sheet_names if name in workbook.sheetnames]
        if len(ordered_sheets) != len(workbook.worksheets):
            remaining = [sheet for sheet in workbook.worksheets if sheet.title not in ordered_sheet_names]
            ordered_sheets.extend(remaining)
        workbook._sheets = ordered_sheets

    def _prepare_trial_balance_payload(self):
        report = self._get_report_or_raise('account_reports.trial_balance_report')
        year_end_date = self._get_effective_year_end_date()
        # Trial balance export must be flat (no account-group hierarchy) and fully expanded to leaves.
        tb_overrides = {
            'hierarchy': False,
            'unfold_all': True,
            'unfolded_lines': [],
        }
        options = self._build_report_options(
            report,
            date_mode='range',
            date_from=self.date_from,
            date_to=year_end_date,
            overrides=tb_overrides,
            aged=False,
        )
        payload = self._build_native_report_sheet_payload(
            report=report,
            options=options,
            sheet_name='Trial Balance',
            empty_message=_('Trial Balance XLSX export returned no file content.'),
        )
        payload['tb_maps'] = self._prepare_trial_balance_support_maps(self._get_reporting_periods())
        return payload

    def _prepare_trial_balance_map(self, *, date_from, date_to, sheet_name, empty_message):
        if not date_to:
            return {}
        report = self._get_report_or_raise('account_reports.trial_balance_report')
        tb_overrides = {
            'hierarchy': False,
            'unfold_all': True,
            'unfolded_lines': [],
        }
        # Opening snapshots are requested with no start date. In that case, force
        # single-date mode and provide a concrete date_from to avoid passing a
        # boolean date into account_reports audit option initialization.
        if date_from:
            date_mode = 'range'
            effective_date_from = date_from
        else:
            date_mode = 'single'
            effective_date_from = date_to
        options = self._build_report_options(
            report,
            date_mode=date_mode,
            date_from=effective_date_from,
            date_to=date_to,
            overrides=tb_overrides,
            aged=False,
        )
        payload = self._build_native_report_sheet_payload(
            report=report,
            options=options,
            sheet_name=sheet_name,
            empty_message=empty_message,
        )
        source_sheet = payload.get('native_source_sheet')
        _layout, rows, _total_row = self._extract_tb_data_rows_from_worksheet(source_sheet)
        result = {}
        for row in rows:
            code = row.get('code')
            if not code:
                continue
            result[code] = {
                'debit': float(row.get('debit') or 0.0),
                'credit': float(row.get('credit') or 0.0),
                'balance': float(row.get('balance') or 0.0),
            }
        return result

    def _prepare_trial_balance_support_maps(self, periods):
        prior_period_map = self._prepare_trial_balance_map(
            date_from=periods.get('prior_date_start'),
            date_to=periods.get('prior_date_end'),
            sheet_name='Trial Balance (Prior)',
            empty_message=_('Prior Trial Balance XLSX export returned no file content.'),
        )
        opening_current_map = self._prepare_trial_balance_map(
            date_from=False,
            date_to=periods.get('opening_date_end'),
            sheet_name='Trial Balance (Opening Current)',
            empty_message=_('Opening current Trial Balance XLSX export returned no file content.'),
        )
        opening_prior_map = self._prepare_trial_balance_map(
            date_from=False,
            date_to=periods.get('prior_opening_date_end'),
            sheet_name='Trial Balance (Opening Prior)',
            empty_message=_('Opening prior Trial Balance XLSX export returned no file content.'),
        )
        return {
            'periods': periods,
            'prior_period': prior_period_map,
            'opening_current': opening_current_map,
            'opening_prior': opening_prior_map,
        }

    def _extract_trial_balance_amounts(self, line_columns, normalized_labels):
        value_map = {}
        numeric_values = []
        for idx, col in enumerate(line_columns):
            value = col.get('no_format')
            if not isinstance(value, (int, float)):
                continue
            numeric_value = float(value)
            label = normalized_labels[idx] if idx < len(normalized_labels) else f'col{idx}'
            value_map[label] = numeric_value
            numeric_values.append(numeric_value)

        def get_first(*keys):
            for key in keys:
                if key in value_map:
                    return value_map[key]
            return None

        debit = get_first('debit', 'perioddebit', 'currentdebit')
        credit = get_first('credit', 'periodcredit', 'currentcredit')

        ending_debit = get_first('endingdebit', 'closingdebit')
        ending_credit = get_first('endingcredit', 'closingcredit')
        balance = get_first('balance', 'endingbalance', 'totalbalance')

        if balance is None and ending_debit is not None and ending_credit is not None:
            balance = ending_debit - ending_credit

        if debit is None or credit is None:
            if len(numeric_values) >= 2:
                debit = numeric_values[-2]
                credit = numeric_values[-1]
            else:
                debit = debit or 0.0
                credit = credit or 0.0

        if balance is None:
            balance = debit - credit

        return {
            'debit': float(debit or 0.0),
            'credit': float(credit or 0.0),
            'balance': float(balance or 0.0),
        }

    @staticmethod
    def _normalize_account_code(code):
        return ''.join(ch for ch in (code or '') if ch.isdigit())

    @staticmethod
    def _coerce_numeric(value):
        if isinstance(value, bool):
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            cleaned = value.replace(',', '').strip()
            if not cleaned:
                return 0.0
            try:
                return float(cleaned)
            except ValueError:
                return 0.0
        return 0.0

    def _detect_trial_balance_layout(self, worksheet):
        max_scan_rows = min(worksheet.max_row, 120)
        max_scan_cols = min(worksheet.max_column, 20)
        for row_idx in range(1, max_scan_rows + 1):
            columns = {}
            balance_cols = []
            for col_idx in range(1, max_scan_cols + 1):
                value = worksheet.cell(row=row_idx, column=col_idx).value
                if not isinstance(value, str):
                    continue
                key = self._normalize_key(value)
                if key in ('code', 'accountcode'):
                    columns.setdefault('code_col', col_idx)
                    continue
                if key in ('account', 'accountname', 'name'):
                    columns.setdefault('account_col', col_idx)
                    continue
                if key in ('debit', 'perioddebit', 'currentdebit'):
                    columns.setdefault('debit_col', col_idx)
                    continue
                if key in ('credit', 'periodcredit', 'currentcredit'):
                    columns.setdefault('credit_col', col_idx)
                    continue
                if key in ('balance', 'endingbalance', 'totalbalance', 'closingbalance'):
                    balance_cols.append(col_idx)

            code_like_col = columns.get('code_col') or columns.get('account_col')
            if code_like_col and (
                balance_cols or (columns.get('debit_col') and columns.get('credit_col'))
            ):
                if not columns.get('code_col'):
                    columns['code_col'] = code_like_col
                if balance_cols:
                    debit_col = columns.get('debit_col')
                    credit_col = columns.get('credit_col')
                    opening_candidates = [col for col in balance_cols if debit_col and col < debit_col]
                    ending_candidates = [col for col in balance_cols if credit_col and col > credit_col]
                    columns['opening_balance_col'] = (
                        opening_candidates[0]
                        if opening_candidates
                        else balance_cols[0]
                    )
                    columns['balance_col'] = (
                        ending_candidates[-1]
                        if ending_candidates
                        else balance_cols[-1]
                    )
                columns['header_row'] = row_idx
                return columns
        return {}

    def _extract_tb_data_rows_from_worksheet(self, worksheet):
        layout = self._detect_trial_balance_layout(worksheet)
        if not layout:
            return {}, [], None

        rows = []
        total_row = None
        header_row = layout['header_row']
        code_col = layout.get('code_col')
        account_col = layout.get('account_col')
        debit_col = layout.get('debit_col')
        credit_col = layout.get('credit_col')
        opening_balance_col = layout.get('opening_balance_col')
        balance_col = layout.get('balance_col')

        for row_idx in range(header_row + 1, worksheet.max_row + 1):
            raw_code = worksheet.cell(row=row_idx, column=code_col).value
            if isinstance(raw_code, str) and self._normalize_key(raw_code) in ('total', 'grandtotal'):
                total_row = row_idx
                break

            raw_code_text = str(raw_code or '').strip()
            normalized_code = self._normalize_account_code(raw_code_text)
            if len(normalized_code) < 6:
                continue

            account_name = ''
            if account_col and account_col != code_col:
                account_name = str(worksheet.cell(row=row_idx, column=account_col).value or '').strip()

            debit_value = self._coerce_numeric(
                worksheet.cell(row=row_idx, column=debit_col).value if debit_col else 0.0
            )
            credit_value = self._coerce_numeric(
                worksheet.cell(row=row_idx, column=credit_col).value if credit_col else 0.0
            )
            opening_balance_value = self._coerce_numeric(
                worksheet.cell(row=row_idx, column=opening_balance_col).value if opening_balance_col else 0.0
            )
            if balance_col:
                balance_value = self._coerce_numeric(
                    worksheet.cell(row=row_idx, column=balance_col).value
                )
            else:
                balance_value = debit_value - credit_value

            rows.append({
                'row': row_idx,
                'code': normalized_code,
                'code_raw': raw_code_text,
                'name': account_name,
                'opening_balance': opening_balance_value,
                'debit': debit_value,
                'credit': credit_value,
                'balance': balance_value,
            })

        return layout, rows, total_row

    def _copy_cell_style(self, source_cell, target_cell):
        if not source_cell or not source_cell.has_style:
            return
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)

    def _copy_row_style(self, sheet, source_row, target_row, max_col=None):
        if not sheet or source_row <= 0 or target_row <= 0:
            return
        max_col = max_col or min(sheet.max_column, 10)
        for col_idx in range(1, max_col + 1):
            self._copy_cell_style(
                sheet.cell(row=source_row, column=col_idx),
                sheet.cell(row=target_row, column=col_idx),
            )
        source_row_dim = sheet.row_dimensions.get(source_row)
        if source_row_dim and source_row_dim.height is not None:
            sheet.row_dimensions[target_row].height = source_row_dim.height

    def _find_row_by_label(self, sheet, label, *, start_row=1, end_row=None, occurrence=1):
        if sheet is None or not label:
            return None
        end_row = end_row or sheet.max_row
        expected = self._normalize_key(label)
        seen = 0
        for row_idx in range(start_row, end_row + 1):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if self._normalize_key(cell_value or '') != expected:
                continue
            seen += 1
            if seen == occurrence:
                return row_idx
        return None

    def _find_row_by_keys(self, sheet, keys, *, start_row=1, end_row=None, occurrence=1):
        if sheet is None:
            return None
        normalized = {self._normalize_key(key) for key in (keys or []) if key}
        if not normalized:
            return None
        end_row = end_row or sheet.max_row
        seen = 0
        for row_idx in range(start_row, end_row + 1):
            cell_value = self._normalize_key(sheet.cell(row=row_idx, column=1).value or '')
            if cell_value not in normalized:
                continue
            seen += 1
            if seen == occurrence:
                return row_idx
        return None

    def _find_cell_position(self, sheet, keys, *, start_row=1, end_row=None, start_col=1, end_col=None, occurrence=1):
        if sheet is None:
            return (None, None)
        if isinstance(keys, str):
            keys = [keys]
        normalized = {self._normalize_key(key) for key in (keys or []) if key}
        if not normalized:
            return (None, None)
        end_row = end_row or sheet.max_row
        end_col = end_col or sheet.max_column
        seen = 0
        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                cell_value = self._normalize_key(sheet.cell(row=row_idx, column=col_idx).value or '')
                if cell_value not in normalized:
                    continue
                seen += 1
                if seen == occurrence:
                    return (row_idx, col_idx)
        return (None, None)

    def _find_column_by_label(self, sheet, label, *, start_row=1, end_row=None, start_col=1, end_col=None, occurrence=1):
        _row_idx, col_idx = self._find_cell_position(
            sheet,
            label,
            start_row=start_row,
            end_row=end_row,
            start_col=start_col,
            end_col=end_col,
            occurrence=occurrence,
        )
        return col_idx

    def _find_column_by_keys(self, sheet, keys, *, start_row=1, end_row=None, start_col=1, end_col=None, occurrence=1):
        _row_idx, col_idx = self._find_cell_position(
            sheet,
            keys,
            start_row=start_row,
            end_row=end_row,
            start_col=start_col,
            end_col=end_col,
            occurrence=occurrence,
        )
        return col_idx

    def _set_detail_label(self, sheet, row_idx, label, *, indent=1):
        cell = sheet.cell(row=row_idx, column=1, value=label)
        base_alignment = cell.alignment or Alignment(horizontal='left', vertical='center')
        cell.alignment = Alignment(
            horizontal=base_alignment.horizontal,
            vertical=base_alignment.vertical,
            text_rotation=base_alignment.text_rotation,
            wrap_text=base_alignment.wrap_text,
            shrink_to_fit=base_alignment.shrink_to_fit,
            indent=indent,
        )

    def _ensure_sofp_hierarchy_rows(self, sheet):
        if sheet is None:
            return

        prepayment_row = self._find_row_by_label(sheet, 'Prepayment')
        if prepayment_row and self._find_row_by_label(sheet, 'Advances', start_row=prepayment_row + 1, end_row=prepayment_row + 4) is None:
            sheet.insert_rows(prepayment_row + 1, 3)
            detail_style_row = prepayment_row + 4
            labels = ['Advances', 'Prepaid expenses', 'Other receivables']
            for offset, label in enumerate(labels, start=1):
                target_row = prepayment_row + offset
                self._copy_row_style(sheet, detail_style_row, target_row, max_col=5)
                self._set_detail_label(sheet, target_row, label, indent=1)

        current_liabilities_row = self._find_row_by_label(sheet, 'Current liabilities')
        if current_liabilities_row and self._find_row_by_label(
            sheet,
            'Bank overdraft',
            start_row=current_liabilities_row + 1,
            end_row=current_liabilities_row + 8,
        ) is None:
            sheet.insert_rows(current_liabilities_row + 1, 7)
            detail_style_row = current_liabilities_row + 8
            labels = [
                'Bank overdraft',
                'Short-term loan',
                'Credit card payable',
                'Lease liability',
                'Trade payables',
                'Other payables',
                'Corporate tax liability',
            ]
            for offset, label in enumerate(labels, start=1):
                target_row = current_liabilities_row + offset
                self._copy_row_style(sheet, detail_style_row, target_row, max_col=5)
                self._set_detail_label(sheet, target_row, label, indent=1)

    def _ensure_soci_hierarchy_rows(self, sheet):
        if sheet is None:
            return []

        operating_heading_row = self._find_row_by_label(sheet, 'Operating expenses')
        total_operating_row = self._find_row_by_label(sheet, 'Total operating expenses')
        if not operating_heading_row or not total_operating_row or total_operating_row <= operating_heading_row:
            return []

        op_lines = [
            {'label': 'Director salary', 'metric': 'soci_director_salary', 'indent': 0},
            {'label': 'Salaries, wages and benefits', 'metric': None, 'indent': 0},
            {'label': 'Office Staff Salaries', 'metric': 'soci_salary_office', 'indent': 1},
            {'label': 'Coaching Staff Salaries', 'metric': 'soci_salary_coaching', 'indent': 1},
            {'label': 'Employee Benefits & Allowances', 'metric': 'soci_salary_benefits', 'indent': 1},
            {'label': 'Bonus & Incentives', 'metric': 'soci_salary_bonus', 'indent': 1},
            {'label': 'Staff Welfare', 'metric': 'soci_salary_welfare', 'indent': 1},
            {'label': 'Advertising', 'metric': 'soci_advertising', 'indent': 0},
            {'label': 'Audit and accounting', 'metric': None, 'indent': 0},
            {'label': 'Audit Fee', 'metric': 'soci_audit_fee', 'indent': 1},
            {'label': 'Accounting & Bookkeeping Fee', 'metric': 'soci_accounting_fee', 'indent': 1},
            {'label': 'Depreciation and amortization', 'metric': None, 'indent': 0},
            {'label': 'Depreciation Expense', 'metric': 'soci_depreciation_expense', 'indent': 1},
            {'label': 'Amortization Expense', 'metric': 'soci_amortization_expense', 'indent': 1},
            {'label': 'Government fees', 'metric': None, 'indent': 0},
            {'label': 'Legal & Government Fee', 'metric': 'soci_legal_gov_fee', 'indent': 1},
            {'label': 'Trade License', 'metric': 'soci_trade_license', 'indent': 1},
            {'label': 'Establishment Card', 'metric': 'soci_establishment_card', 'indent': 1},
            {'label': 'Visa Fee', 'metric': 'soci_visa_fee', 'indent': 1},
            {'label': 'Insurance', 'metric': 'soci_insurance', 'indent': 0},
            {'label': 'Office expense', 'metric': 'soci_office_expense', 'indent': 0},
            {'label': 'Bank charges', 'metric': 'soci_bank_charges', 'indent': 0},
            {'label': 'Exchange loss', 'metric': 'soci_exchange_loss', 'indent': 0},
            {'label': 'Other expenses', 'metric': 'soci_other_expenses', 'indent': 0},
        ]

        existing_line_count = max(total_operating_row - operating_heading_row - 1, 0)
        desired_line_count = len(op_lines)
        if desired_line_count > existing_line_count:
            sheet.insert_rows(total_operating_row, desired_line_count - existing_line_count)
            total_operating_row += (desired_line_count - existing_line_count)

        detail_style_row = operating_heading_row + 1
        for offset, line in enumerate(op_lines, start=1):
            row_idx = operating_heading_row + offset
            self._copy_row_style(sheet, detail_style_row, row_idx, max_col=6)
            self._set_detail_label(sheet, row_idx, line['label'], indent=line['indent'])

        return op_lines

    def _find_row_containing_text(self, sheet, text, *, start_row=1, end_row=None, occurrence=1):
        if sheet is None or not text:
            return None
        end_row = end_row or sheet.max_row
        expected = self._normalize_key(text)
        seen = 0
        for row_idx in range(start_row, end_row + 1):
            cell_value = self._normalize_key(sheet.cell(row=row_idx, column=1).value or '')
            if expected not in cell_value:
                continue
            seen += 1
            if seen == occurrence:
                return row_idx
        return None

    def _get_soci_net_profit_row(self, workbook):
        soci_sheet = self._get_sheet_if_exists(workbook, 'SOCI')
        if soci_sheet is None:
            return 31
        return self._find_row_by_keys(soci_sheet, ['Net profit / (loss)', 'Net profit']) or 31

    def _refresh_summary_sheet_links(self, workbook):
        summary_sheet = self._get_sheet_if_exists(workbook, 'Summary Sheet')
        sofp_sheet = self._get_sheet_if_exists(workbook, 'SOFP')
        soci_sheet = self._get_sheet_if_exists(workbook, 'SOCI')
        if summary_sheet is None or sofp_sheet is None or soci_sheet is None:
            return

        summary_row_total_revenue = self._find_row_by_label(summary_sheet, 'Total Revenue')
        summary_row_total_expenses = self._find_row_by_label(summary_sheet, 'Total Expenses (Excluding Director Salary)')
        summary_row_director_salary = self._find_row_by_label(summary_sheet, 'Director Salary')
        summary_row_profit_loss = self._find_row_by_label(summary_sheet, 'Profit / (Loss)')
        summary_row_total_assets = self._find_row_by_label(summary_sheet, 'Total Assets')
        summary_row_total_liabilities = self._find_row_by_label(summary_sheet, 'Total Liabilities')
        summary_row_total_equity = self._find_row_by_label(summary_sheet, 'Total Equity')
        summary_row_related_revenue = self._find_row_by_label(summary_sheet, 'Related Party Revenue')
        summary_row_related_loan = self._find_row_by_label(summary_sheet, 'Related Party Loan')
        summary_row_interest_income = self._find_row_by_label(summary_sheet, 'Interest Income')

        soci_row_revenue = self._find_row_by_label(soci_sheet, 'Revenue')
        soci_row_revenue_related = self._find_row_by_label(soci_sheet, 'Revenue - related party')
        soci_row_director_salary = self._find_row_by_label(soci_sheet, 'Director salary')
        soci_row_total_operating_expenses = self._find_row_by_label(soci_sheet, 'Total operating expenses')
        soci_row_other_income = self._find_row_by_label(soci_sheet, 'Other income')
        soci_row_net_profit = self._get_soci_net_profit_row(workbook)

        sofp_row_total_assets = self._find_row_by_label(sofp_sheet, 'Total assets')
        sofp_row_total_liabilities = self._find_row_by_label(sofp_sheet, 'Total Liabilities')
        sofp_row_total_equity = self._find_row_by_label(sofp_sheet, 'Total Equity')
        sofp_row_related_party_loan = self._find_row_by_label(sofp_sheet, 'Loan from related party')

        if summary_row_total_revenue and soci_row_revenue:
            summary_sheet[f'B{summary_row_total_revenue}'] = f'=SOCI!B{soci_row_revenue}'
        if summary_row_total_expenses and soci_row_total_operating_expenses and soci_row_director_salary:
            summary_sheet[f'B{summary_row_total_expenses}'] = (
                f'=SOCI!B{soci_row_total_operating_expenses}-SOCI!B{soci_row_director_salary}'
            )
        if summary_row_director_salary and soci_row_director_salary:
            summary_sheet[f'B{summary_row_director_salary}'] = f'=SOCI!B{soci_row_director_salary}'
        if summary_row_profit_loss and soci_row_net_profit:
            summary_sheet[f'B{summary_row_profit_loss}'] = f'=SOCI!B{soci_row_net_profit}'
        if summary_row_total_assets and sofp_row_total_assets:
            summary_sheet[f'B{summary_row_total_assets}'] = f'=SOFP!B{sofp_row_total_assets}'
        if summary_row_total_liabilities and sofp_row_total_liabilities:
            summary_sheet[f'B{summary_row_total_liabilities}'] = f'=SOFP!B{sofp_row_total_liabilities}'
        if summary_row_total_equity and sofp_row_total_equity:
            summary_sheet[f'B{summary_row_total_equity}'] = f'=SOFP!B{sofp_row_total_equity}'
        if summary_row_related_revenue and soci_row_revenue_related:
            summary_sheet[f'B{summary_row_related_revenue}'] = f'=SOCI!B{soci_row_revenue_related}'
        if summary_row_related_loan and sofp_row_related_party_loan:
            summary_sheet[f'B{summary_row_related_loan}'] = f'=SOFP!B{sofp_row_related_party_loan}'
        if summary_row_interest_income and soci_row_other_income:
            summary_sheet[f'B{summary_row_interest_income}'] = f'=SOCI!B{soci_row_other_income}'

    def _ensure_soce_dividend_rows(self, sheet, metric_refs=None):
        if sheet is None or not self._tb_context_has_account_code(metric_refs, '31010202'):
            return
        if self._find_row_by_label(sheet, 'Dividend paid', occurrence=1) and self._find_row_by_label(
            sheet, 'Dividend paid', occurrence=2
        ):
            return

        def insert_dividend_row(before_row):
            if not before_row:
                return
            sheet.insert_rows(before_row, 1)
            source_row = max(before_row - 1, 1)
            self._copy_row_style(sheet, source_row, before_row, max_col=6)
            sheet.cell(row=before_row, column=1, value='Dividend paid')
            for col_idx in range(2, 6):
                sheet.cell(row=before_row, column=col_idx, value=None)
            sheet.cell(row=before_row, column=6, value=f'=SUM(B{before_row}:E{before_row})')

        prior_close_row = (
            self._find_row_by_label(sheet, 'Balance as at end of period')
            or self._find_row_containing_text(sheet, 'Balance as at', occurrence=2)
        )
        if self._find_row_by_label(sheet, 'Dividend paid', occurrence=1) is None and prior_close_row:
            insert_dividend_row(prior_close_row)

        current_close_row = (
            self._find_row_by_label(sheet, 'Balance c/f')
            or self._find_row_containing_text(sheet, 'Balance as at', occurrence=3)
        )
        if self._find_row_by_label(sheet, 'Dividend paid', occurrence=2) is None and current_close_row:
            insert_dividend_row(current_close_row)

    def _ensure_socf_dividend_row(self, sheet, metric_refs=None):
        if sheet is None or not self._tb_context_has_account_code(metric_refs, '31010202'):
            return
        if self._find_row_by_label(sheet, 'Dividend paid'):
            return

        insert_before_row = self._find_row_by_label(sheet, 'Owner current account') or self._find_row_by_keys(
            sheet,
            ['Net cash generated from financing activities', 'Net cash (used in) financing activities'],
            start_row=1,
            end_row=sheet.max_row,
        )
        if not insert_before_row:
            return

        sheet.insert_rows(insert_before_row, 1)
        source_row = min(insert_before_row + 1, sheet.max_row)
        self._copy_row_style(sheet, source_row, insert_before_row, max_col=3)
        sheet.cell(row=insert_before_row, column=1, value='Dividend paid')
        sheet.cell(row=insert_before_row, column=2, value=None)
        sheet.cell(row=insert_before_row, column=3, value=None)

    def _stage_g_link_statement_sheets_to_trial_balance(self, stage_c_context, trial_balance_payload):
        workbook = stage_c_context.get('workbook')
        trial_balance_sheet_name = stage_c_context.get('trial_balance_sheet')
        if not workbook or not trial_balance_sheet_name or trial_balance_sheet_name not in workbook.sheetnames:
            return

        tb_sheet = workbook[trial_balance_sheet_name]
        layout, rows, total_row = self._extract_tb_data_rows_from_worksheet(tb_sheet)
        if not layout or not rows:
            return

        tb_maps = (trial_balance_payload or {}).get('tb_maps') or {}
        metric_refs = self._prepare_tb_formula_links(
            tb_sheet=tb_sheet,
            layout=layout,
            rows=rows,
            total_row=total_row,
            tb_maps=tb_maps,
        )
        if not metric_refs:
            return

        self._link_soci_to_tb_metrics(workbook, metric_refs, tb_sheet.title)
        self._link_soce_to_tb_metrics(workbook, metric_refs, tb_sheet.title)
        self._link_sofp_to_tb_metrics(workbook, metric_refs, tb_sheet.title)
        self._link_socf_to_tb_metrics(workbook, metric_refs, tb_sheet.title)
        self._refresh_summary_sheet_links(workbook)

        calculation = getattr(workbook, 'calculation', None)
        if calculation is not None:
            calculation.fullCalcOnLoad = True

    def _prepare_tb_formula_links(self, *, tb_sheet, layout, rows, total_row, tb_maps):
        header_row = layout['header_row']
        current_debit_col = layout.get('debit_col')
        current_credit_col = layout.get('credit_col')
        current_balance_col = layout.get('balance_col')
        if not current_debit_col or not current_credit_col:
            return {}

        periods = (tb_maps or {}).get('periods') or {}
        prior_period_map = (tb_maps or {}).get('prior_period') or {}
        opening_current_map = (tb_maps or {}).get('opening_current') or {}
        opening_prior_map = (tb_maps or {}).get('opening_prior') or {}
        current_bs_mode = periods.get('balance_sheet_date_mode') or 'end_only'
        prior_bs_mode = periods.get('prior_balance_sheet_date_mode') or 'end_only'

        next_col = max(tb_sheet.max_column, current_balance_col or 0) + 1
        if not current_balance_col:
            current_balance_col = next_col
            next_col += 1

        prior_debit_col = next_col
        prior_credit_col = next_col + 1
        prior_balance_col = next_col + 2
        opening_current_col = next_col + 3
        opening_prior_col = next_col + 4
        closing_current_col = next_col + 5
        closing_prior_col = next_col + 6
        norm_code_col = next_col + 7
        helper_label_col = next_col + 9
        helper_current_col = next_col + 10
        helper_prior_col = next_col + 11

        header_style_source = tb_sheet.cell(row=header_row, column=current_debit_col)
        number_style_source = tb_sheet.cell(row=rows[0]['row'], column=current_debit_col)
        number_format = number_style_source.number_format or '#,##0.00'

        headers = [
            (current_balance_col, 'Balance') if layout.get('balance_col') is None else None,
            (prior_debit_col, 'Prior Debit'),
            (prior_credit_col, 'Prior Credit'),
            (prior_balance_col, 'Prior Balance'),
            (opening_current_col, 'Opening Current'),
            (opening_prior_col, 'Opening Prior'),
            (closing_current_col, 'Closing Current'),
            (closing_prior_col, 'Closing Prior'),
            (norm_code_col, 'Norm Code'),
            (helper_label_col, 'TB Link Metric'),
            (helper_current_col, 'Current'),
            (helper_prior_col, 'Prior'),
        ]
        for item in headers:
            if not item:
                continue
            col_idx, label = item
            header_cell = tb_sheet.cell(row=header_row, column=col_idx, value=label)
            self._copy_cell_style(header_style_source, header_cell)

        data_start_row = min(row['row'] for row in rows)
        data_end_row = max(row['row'] for row in rows)

        current_debit_letter = get_column_letter(current_debit_col)
        current_credit_letter = get_column_letter(current_credit_col)
        current_balance_letter = get_column_letter(current_balance_col)
        prior_debit_letter = get_column_letter(prior_debit_col)
        prior_credit_letter = get_column_letter(prior_credit_col)
        prior_balance_letter = get_column_letter(prior_balance_col)
        opening_current_letter = get_column_letter(opening_current_col)
        opening_prior_letter = get_column_letter(opening_prior_col)
        closing_current_letter = get_column_letter(closing_current_col)
        closing_prior_letter = get_column_letter(closing_prior_col)
        norm_code_letter = get_column_letter(norm_code_col)

        for row in rows:
            row_idx = row['row']
            row_style_source = tb_sheet.cell(row=row_idx, column=current_debit_col)

            if layout.get('balance_col') is None:
                balance_cell = tb_sheet.cell(row=row_idx, column=current_balance_col)
                balance_cell.value = f"={current_debit_letter}{row_idx}-{current_credit_letter}{row_idx}"
                self._copy_cell_style(row_style_source, balance_cell)
                balance_cell.number_format = number_format

            prior_vals = prior_period_map.get(row['code']) or {}
            for col_idx, value in (
                (prior_debit_col, float(prior_vals.get('debit') or 0.0)),
                (prior_credit_col, float(prior_vals.get('credit') or 0.0)),
                (prior_balance_col, float(prior_vals.get('balance') or 0.0)),
            ):
                cell = tb_sheet.cell(row=row_idx, column=col_idx, value=value)
                self._copy_cell_style(row_style_source, cell)
                cell.number_format = number_format

            opening_current_value = float(row.get('opening_balance') or 0.0)
            if not layout.get('opening_balance_col'):
                opening_current_value = 0.0 if current_bs_mode == 'range' else float(
                    (opening_current_map.get(row['code']) or {}).get('balance') or 0.0
                )
            opening_prior_value = 0.0 if prior_bs_mode == 'range' else float(
                (opening_prior_map.get(row['code']) or {}).get('balance') or 0.0
            )

            opening_current_cell = tb_sheet.cell(
                row=row_idx, column=opening_current_col, value=opening_current_value
            )
            opening_prior_cell = tb_sheet.cell(
                row=row_idx, column=opening_prior_col, value=opening_prior_value
            )
            self._copy_cell_style(row_style_source, opening_current_cell)
            self._copy_cell_style(row_style_source, opening_prior_cell)
            opening_current_cell.number_format = number_format
            opening_prior_cell.number_format = number_format

            closing_current_cell = tb_sheet.cell(row=row_idx, column=closing_current_col)
            closing_prior_cell = tb_sheet.cell(row=row_idx, column=closing_prior_col)
            closing_current_cell.value = f"={current_balance_letter}{row_idx}"
            closing_prior_cell.value = f"={prior_balance_letter}{row_idx}"
            self._copy_cell_style(row_style_source, closing_current_cell)
            self._copy_cell_style(row_style_source, closing_prior_cell)
            closing_current_cell.number_format = number_format
            closing_prior_cell.number_format = number_format

            norm_cell = tb_sheet.cell(row=row_idx, column=norm_code_col, value=row['code'])
            self._copy_cell_style(row_style_source, norm_cell)

        if total_row:
            for letter, col_idx in (
                (prior_debit_letter, prior_debit_col),
                (prior_credit_letter, prior_credit_col),
                (prior_balance_letter, prior_balance_col),
                (opening_current_letter, opening_current_col),
                (opening_prior_letter, opening_prior_col),
                (closing_current_letter, closing_current_col),
                (closing_prior_letter, closing_prior_col),
            ):
                total_cell = tb_sheet.cell(row=total_row, column=col_idx)
                total_cell.value = f"=SUM({letter}{data_start_row}:{letter}{data_end_row})"
                total_cell.number_format = number_format
                self._copy_cell_style(tb_sheet.cell(row=total_row, column=current_debit_col), total_cell)

        code_range = f"${norm_code_letter}${data_start_row}:${norm_code_letter}${data_end_row}"
        current_debit_range = f"${current_debit_letter}${data_start_row}:${current_debit_letter}${data_end_row}"
        current_credit_range = f"${current_credit_letter}${data_start_row}:${current_credit_letter}${data_end_row}"
        current_balance_range = f"${current_balance_letter}${data_start_row}:${current_balance_letter}${data_end_row}"
        prior_debit_range = f"${prior_debit_letter}${data_start_row}:${prior_debit_letter}${data_end_row}"
        prior_credit_range = f"${prior_credit_letter}${data_start_row}:${prior_credit_letter}${data_end_row}"
        prior_balance_range = f"${prior_balance_letter}${data_start_row}:${prior_balance_letter}${data_end_row}"
        opening_current_range = f"${opening_current_letter}${data_start_row}:${opening_current_letter}${data_end_row}"
        opening_prior_range = f"${opening_prior_letter}${data_start_row}:${opening_prior_letter}${data_end_row}"

        def sumifs_expr(value_range, criteria):
            escaped = criteria.replace('"', '""')
            return f'SUMIFS({value_range},{code_range},"{escaped}")'

        def bal_prefix_expr(prefix, is_prior=False):
            value_range = prior_balance_range if is_prior else current_balance_range
            return sumifs_expr(value_range, f'{prefix}*')

        def bal_exact_expr(code, is_prior=False):
            value_range = prior_balance_range if is_prior else current_balance_range
            return sumifs_expr(value_range, code)

        def opening_prefix_expr(prefix, is_prior=False):
            value_range = opening_prior_range if is_prior else opening_current_range
            return sumifs_expr(value_range, f'{prefix}*')

        def opening_exact_expr(code, is_prior=False):
            value_range = opening_prior_range if is_prior else opening_current_range
            return sumifs_expr(value_range, code)

        def movement_prefix_expr(prefix, is_prior=False):
            return f"({bal_prefix_expr(prefix, is_prior)}-{opening_prefix_expr(prefix, is_prior)})"

        def movement_exact_expr(code, is_prior=False):
            return f"({bal_exact_expr(code, is_prior)}-{opening_exact_expr(code, is_prior)})"

        def expr_sum(*parts):
            valid = [part for part in parts if part]
            if not valid:
                return '0'
            return '+'.join(f"({part})" for part in valid)

        def sum_balance_prefixes(prefixes, is_prior=False):
            return expr_sum(*(bal_prefix_expr(prefix, is_prior) for prefix in prefixes))

        def sum_movement_prefixes(prefixes, is_prior=False):
            return expr_sum(*(movement_prefix_expr(prefix, is_prior) for prefix in prefixes))

        metric_refs = {}
        tb_link_rows = []
        helper_row = header_row + 1
        tb_sheet_name_escaped = tb_sheet.title.replace("'", "''")
        helper_header_style = tb_sheet.cell(row=header_row, column=helper_label_col)
        helper_value_style = tb_sheet.cell(row=rows[0]['row'], column=current_debit_col)

        def add_metric(metric_key, current_expr, prior_expr):
            nonlocal helper_row
            label_cell = tb_sheet.cell(row=helper_row, column=helper_label_col, value=metric_key)
            self._copy_cell_style(helper_header_style, label_cell)
            current_cell = tb_sheet.cell(row=helper_row, column=helper_current_col, value=f"={current_expr}")
            prior_cell = tb_sheet.cell(row=helper_row, column=helper_prior_col, value=f"={prior_expr}")
            self._copy_cell_style(helper_value_style, current_cell)
            self._copy_cell_style(helper_value_style, prior_cell)
            current_cell.number_format = number_format
            prior_cell.number_format = number_format

            current_ref = f"='{tb_sheet_name_escaped}'!${get_column_letter(helper_current_col)}${helper_row}"
            prior_ref = f"='{tb_sheet_name_escaped}'!${get_column_letter(helper_prior_col)}${helper_row}"
            metric_refs[metric_key] = {
                'current': current_ref,
                'prior': prior_ref,
            }
            helper_row += 1

        for row in rows:
            row_idx = row.get('row')
            prior_vals = prior_period_map.get(row['code']) or {}
            tb_link_rows.append({
                'code': row.get('code'),
                'code_raw': row.get('code_raw') or row.get('code') or '',
                'name': row.get('name') or '',
                'row_index': row_idx,
                'current_debit_ref': (
                    f"'{tb_sheet_name_escaped}'!${current_debit_letter}${row_idx}" if row_idx else None
                ),
                'current_credit_ref': (
                    f"'{tb_sheet_name_escaped}'!${current_credit_letter}${row_idx}" if row_idx else None
                ),
                'current_balance_ref': (
                    f"'{tb_sheet_name_escaped}'!${current_balance_letter}${row_idx}" if row_idx else None
                ),
                'prior_debit_ref': (
                    f"'{tb_sheet_name_escaped}'!${prior_debit_letter}${row_idx}" if row_idx else None
                ),
                'prior_credit_ref': (
                    f"'{tb_sheet_name_escaped}'!${prior_credit_letter}${row_idx}" if row_idx else None
                ),
                'prior_balance_ref': (
                    f"'{tb_sheet_name_escaped}'!${prior_balance_letter}${row_idx}" if row_idx else None
                ),
                'current_opening_ref': (
                    f"'{tb_sheet_name_escaped}'!${opening_current_letter}${row_idx}" if row_idx else None
                ),
                'prior_opening_ref': (
                    f"'{tb_sheet_name_escaped}'!${opening_prior_letter}${row_idx}" if row_idx else None
                ),
                'current_closing': float(row.get('balance') or 0.0),
                'prior_closing': float(prior_vals.get('balance') or 0.0),
            })

        add_metric(
            'sofp_non_current_other',
            sum_balance_prefixes(
                ['110201', '110202', '110203', '110204', '110205', '110206', '110301', '110401', '110501', '110601', '110701', '110702', '110703', '110801', '110901']
            ),
            sum_balance_prefixes(
                ['110201', '110202', '110203', '110204', '110205', '110206', '110301', '110401', '110501', '110601', '110701', '110702', '110703', '110801', '110901'],
                True,
            ),
        )
        add_metric('sofp_ppe', bal_prefix_expr('1101'), bal_prefix_expr('1101', True))
        add_metric(
            'sofp_accounts_receivable',
            sum_balance_prefixes(['120201', '120202']),
            sum_balance_prefixes(['120201', '120202'], True),
        )
        add_metric('sofp_vat_recoverable', bal_prefix_expr('120304'), bal_prefix_expr('120304', True))
        add_metric(
            'sofp_prepayment',
            sum_balance_prefixes(['120301', '120302', '120303']),
            sum_balance_prefixes(['120301', '120302', '120303'], True),
        )
        add_metric('sofp_prepayment_advances', bal_prefix_expr('120301'), bal_prefix_expr('120301', True))
        add_metric('sofp_prepayment_prepaid_expenses', bal_prefix_expr('120302'), bal_prefix_expr('120302', True))
        add_metric('sofp_prepayment_other_receivables', bal_prefix_expr('120303'), bal_prefix_expr('120303', True))
        add_metric(
            'sofp_cash_bank',
            f"({sum_balance_prefixes(['120401', '120402', '120601'])}-{bal_exact_expr('12040101')})",
            f"({sum_balance_prefixes(['120401', '120402', '120601'], True)}-{bal_exact_expr('12040101', True)})",
        )
        add_metric(
            'share_capital_close',
            f"(-{bal_exact_expr('31010101')})",
            f"(-{bal_exact_expr('31010101', True)})",
        )
        add_metric(
            'owner_current_account_close',
            f"(-{bal_exact_expr('12040101')})",
            f"(-{bal_exact_expr('12040101', True)})",
        )
        add_metric(
            'retained_earnings_open',
            f"((-{bal_exact_expr('31010203')})-(-{movement_exact_expr('31010203')}))",
            f"((-{bal_exact_expr('31010203', True)})-(-{movement_exact_expr('31010203', True)}))",
        )
        add_metric(
            'statutory_reserve_close',
            f"(-{bal_exact_expr('31010301')})",
            f"(-{bal_exact_expr('31010301', True)})",
        )
        add_metric(
            'share_capital_movement',
            f"(-{movement_exact_expr('31010101')})",
            f"(-{movement_exact_expr('31010101', True)})",
        )
        add_metric(
            'owner_current_account_movement',
            f"(-{movement_exact_expr('12040101')})",
            f"(-{movement_exact_expr('12040101', True)})",
        )
        add_metric(
            'statutory_transfer',
            f"(-{movement_exact_expr('31010301')})",
            f"(-{movement_exact_expr('31010301', True)})",
        )
        add_metric(
            'dividend_paid',
            f"(-{movement_exact_expr('31010202')})",
            f"(-{movement_exact_expr('31010202', True)})",
        )
        add_metric(
            'sofp_non_current_liabilities_total',
            f"(-{sum_balance_prefixes(['210101', '210102', '210103', '210201', '210301', '210401'])})",
            f"(-{sum_balance_prefixes(['210101', '210102', '210103', '210201', '210301', '210401'], True)})",
        )
        add_metric(
            'sofp_current_liabilities_accrual',
            f"(-{bal_prefix_expr('220303')})",
            f"(-{bal_prefix_expr('220303', True)})",
        )
        add_metric(
            'sofp_current_liabilities_vat',
            f"(-{bal_prefix_expr('220302')})",
            f"(-{bal_prefix_expr('220302', True)})",
        )
        add_metric(
            'sofp_current_liabilities_other',
            f"(-{sum_balance_prefixes(['220101', '220102', '220103', '220104', '220201', '220301', '220401'])})",
            f"(-{sum_balance_prefixes(['220101', '220102', '220103', '220104', '220201', '220301', '220401'], True)})",
        )
        add_metric('sofp_cl_bank_overdraft', f"(-{bal_prefix_expr('220101')})", f"(-{bal_prefix_expr('220101', True)})")
        add_metric('sofp_cl_short_term_loan', f"(-{bal_prefix_expr('220102')})", f"(-{bal_prefix_expr('220102', True)})")
        add_metric('sofp_cl_credit_card_payable', f"(-{bal_prefix_expr('220103')})", f"(-{bal_prefix_expr('220103', True)})")
        add_metric('sofp_cl_lease_liability', f"(-{bal_prefix_expr('220104')})", f"(-{bal_prefix_expr('220104', True)})")
        add_metric('sofp_cl_trade_payables', f"(-{bal_prefix_expr('220201')})", f"(-{bal_prefix_expr('220201', True)})")
        add_metric('sofp_cl_other_payables', f"(-{bal_prefix_expr('220301')})", f"(-{bal_prefix_expr('220301', True)})")
        add_metric('sofp_cl_corporate_tax_liability', f"(-{bal_prefix_expr('220401')})", f"(-{bal_prefix_expr('220401', True)})")

        add_metric('soci_revenue', f"(-{bal_prefix_expr('410101')})", f"(-{bal_prefix_expr('410101', True)})")
        add_metric('soci_revenue_related', f"(-{bal_prefix_expr('410201')})", f"(-{bal_prefix_expr('410201', True)})")
        add_metric(
            'soci_direct_cost',
            sum_balance_prefixes(['510101', '510102', '510103', '510104']),
            sum_balance_prefixes(['510101', '510102', '510103', '510104'], True),
        )
        add_metric('soci_director_salary', bal_prefix_expr('510701'), bal_prefix_expr('510701', True))
        add_metric('soci_salary_office', bal_prefix_expr('510801'), bal_prefix_expr('510801', True))
        add_metric('soci_salary_coaching', bal_prefix_expr('510802'), bal_prefix_expr('510802', True))
        add_metric('soci_salary_benefits', bal_prefix_expr('510803'), bal_prefix_expr('510803', True))
        add_metric('soci_salary_bonus', bal_prefix_expr('510804'), bal_prefix_expr('510804', True))
        add_metric('soci_salary_welfare', bal_prefix_expr('510805'), bal_prefix_expr('510805', True))
        add_metric(
            'soci_salaries',
            sum_balance_prefixes(['510801', '510802', '510803', '510804', '510805']),
            sum_balance_prefixes(['510801', '510802', '510803', '510804', '510805'], True),
        )
        add_metric('soci_advertising', bal_prefix_expr('510201'), bal_prefix_expr('510201', True))
        add_metric('soci_audit_fee', bal_prefix_expr('510901'), bal_prefix_expr('510901', True))
        add_metric('soci_accounting_fee', bal_prefix_expr('510902'), bal_prefix_expr('510902', True))
        add_metric(
            'soci_audit_accounting',
            sum_balance_prefixes(['510901', '510902']),
            sum_balance_prefixes(['510901', '510902'], True),
        )
        add_metric('soci_depreciation_expense', bal_prefix_expr('511401'), bal_prefix_expr('511401', True))
        add_metric('soci_amortization_expense', bal_prefix_expr('511402'), bal_prefix_expr('511402', True))
        add_metric(
            'soci_depreciation',
            sum_balance_prefixes(['511401', '511402']),
            sum_balance_prefixes(['511401', '511402'], True),
        )
        add_metric('soci_legal_gov_fee', bal_prefix_expr('510601'), bal_prefix_expr('510601', True))
        add_metric('soci_trade_license', bal_prefix_expr('512601'), bal_prefix_expr('512601', True))
        add_metric('soci_establishment_card', bal_prefix_expr('512602'), bal_prefix_expr('512602', True))
        add_metric('soci_visa_fee', bal_prefix_expr('512603'), bal_prefix_expr('512603', True))
        add_metric(
            'soci_government_fees',
            sum_balance_prefixes(['510601', '512601', '512602', '512603']),
            sum_balance_prefixes(['510601', '512601', '512602', '512603'], True),
        )
        add_metric('soci_insurance', bal_prefix_expr('512501'), bal_prefix_expr('512501', True))
        add_metric(
            'soci_office_expense',
            sum_balance_prefixes(['510401', '511001', '511002', '511101', '511201', '511202', '511501', '511701', '511801', '512801']),
            sum_balance_prefixes(['510401', '511001', '511002', '511101', '511201', '511202', '511501', '511701', '511801', '512801'], True),
        )
        add_metric('soci_bank_charges', bal_prefix_expr('512301'), bal_prefix_expr('512301', True))
        add_metric('soci_exchange_loss', bal_prefix_expr('512201'), bal_prefix_expr('512201', True))
        add_metric(
            'soci_other_expenses',
            sum_balance_prefixes(['510301', '510501', '511301', '511302', '511601', '511901', '512001', '512101', '512202', '512401', '512701']),
            sum_balance_prefixes(['510301', '510501', '511301', '511302', '511601', '511901', '512001', '512101', '512202', '512401', '512701'], True),
        )
        add_metric(
            'soci_other_income',
            f"ABS({sum_balance_prefixes(['410301', '410302', '410303', '410304', '410305', '410306', '410307', '410308', '410309', '410310'])})",
            f"ABS({sum_balance_prefixes(['410301', '410302', '410303', '410304', '410305', '410306', '410307', '410308', '410309', '410310'], True)})",
        )
        add_metric(
            'soci_investment_gain_loss',
            f"(-{bal_prefix_expr('5201')})",
            f"(-{bal_prefix_expr('5201', True)})",
        )

        add_metric(
            'socf_depreciation_movement',
            movement_prefix_expr('5114'),
            movement_prefix_expr('5114', True),
        )
        add_metric(
            'socf_eosb_adjustment',
            f"MAX(-{movement_exact_expr('21040101')},0)",
            f"MAX(-{movement_exact_expr('21040101', True)},0)",
        )
        add_metric(
            'socf_change_current_assets',
            f"(-{sum_movement_prefixes(['120101', '120201', '120202', '120301', '120302', '120303', '120304', '120501', '120502'])})",
            f"(-{sum_movement_prefixes(['120101', '120201', '120202', '120301', '120302', '120303', '120304', '120501', '120502'], True)})",
        )
        add_metric(
            'socf_change_current_liabilities',
            f"(-{sum_movement_prefixes(self._SOCF_CURRENT_LIABILITY_PREFIXES)})",
            f"(-{sum_movement_prefixes(self._SOCF_CURRENT_LIABILITY_PREFIXES, True)})",
        )
        add_metric(
            'socf_property_investing',
            f"(-{movement_prefix_expr('1101')})",
            f"(-{movement_prefix_expr('1101', True)})",
        )
        add_metric(
            'socf_security_deposit',
            f"(-{movement_exact_expr(self._SOCF_SECURITY_DEPOSIT_ACCOUNT_CODE)})",
            f"(-{movement_exact_expr(self._SOCF_SECURITY_DEPOSIT_ACCOUNT_CODE, True)})",
        )
        add_metric(
            'socf_related_party_loan',
            f"(-{movement_prefix_expr(self._SOCF_RELATED_PARTY_LOAN_PREFIX)})",
            f"(-{movement_prefix_expr(self._SOCF_RELATED_PARTY_LOAN_PREFIX, True)})",
        )
        add_metric(
            'socf_corporate_tax_paid',
            (
                f"((-{bal_prefix_expr('220401')})-(-{opening_prefix_expr('220401')})"
                f"-({expr_sum(*(movement_exact_expr(code) for code in self._SOCF_CT_EXPENSE_ACCOUNT_CODES))}))"
            ),
            (
                f"((-{bal_prefix_expr('220401', True)})-(-{opening_prefix_expr('220401', True)})"
                f"-({expr_sum(*(movement_exact_expr(code, True) for code in self._SOCF_CT_EXPENSE_ACCOUNT_CODES))}))"
            ),
        )
        add_metric(
            'socf_eosb_paid',
            f"MIN(-{movement_exact_expr(self._SOCF_EOSB_LIABILITY_CODE)},0)",
            f"MIN(-{movement_exact_expr(self._SOCF_EOSB_LIABILITY_CODE, True)},0)",
        )
        add_metric(
            'socf_cash_equivalent_opening',
            f"({opening_prefix_expr('1204')}-{opening_exact_expr('12040101')}-{opening_prefix_expr('1206')})",
            f"({opening_prefix_expr('1204', True)}-{opening_exact_expr('12040101', True)}-{opening_prefix_expr('1206', True)})",
        )

        for col_idx in (
            opening_current_col,
            opening_prior_col,
            closing_current_col,
            closing_prior_col,
            norm_code_col,
            helper_label_col,
            helper_current_col,
            helper_prior_col,
        ):
            tb_sheet.column_dimensions[get_column_letter(col_idx)].hidden = True

        metric_refs['__tb_link_context__'] = {
            'sheet_name': tb_sheet.title,
            'code_range_ref': f"'{tb_sheet_name_escaped}'!{code_range}",
            'current_balance_range_ref': f"'{tb_sheet_name_escaped}'!{current_balance_range}",
            'prior_balance_range_ref': f"'{tb_sheet_name_escaped}'!{prior_balance_range}",
            'present_codes': {
                str(code)
                for code in (
                    list(prior_period_map.keys())
                    + [row.get('code') for row in rows if row.get('code')]
                )
                if code
            },
            'rows': tb_link_rows,
        }

        return metric_refs

    def _metric_ref(self, metric_refs, key, period):
        return metric_refs.get(key, {}).get(period, '0')

    def _metric_expr(self, metric_refs, key, period):
        ref = self._metric_ref(metric_refs, key, period)
        if isinstance(ref, str) and ref.startswith('='):
            return ref[1:]
        return ref

    def _tb_context_has_account_code(self, metric_refs, *codes):
        tb_context = (metric_refs or {}).get('__tb_link_context__') or {}
        present_codes = tb_context.get('present_codes') or set()
        normalized_codes = {str(code).strip() for code in (codes or []) if code}
        if not normalized_codes:
            return False
        if present_codes.intersection(normalized_codes):
            return True
        for row in tb_context.get('rows') or []:
            row_code = str(row.get('code') or row.get('code_raw') or '').strip()
            if row_code in normalized_codes:
                return True
        return False

    def _tb_formula_expr(self, formula):
        if isinstance(formula, str) and formula.startswith('='):
            return formula[1:]
        return formula or '0'

    def _tb_match_context_rows(self, tb_context, criteria):
        if not tb_context or criteria is None:
            return []

        token = str(criteria).strip()
        if not token:
            return []

        is_prefix = token.endswith('*')
        prefix = token[:-1] if is_prefix else token
        matches = []
        for row in tb_context.get('rows') or []:
            code = str(row.get('code') or row.get('code_raw') or '').strip()
            if not code:
                continue
            if is_prefix:
                if not code.startswith(prefix):
                    continue
            elif code != token:
                continue
            matches.append(row)
        return matches

    def _tb_row_reference_terms(self, row_matches, ref_key):
        indexed_rows = []
        direct_refs = []

        for row in row_matches or []:
            ref = row.get(ref_key)
            if not ref:
                continue
            row_idx = row.get('row_index')
            if isinstance(row_idx, int):
                indexed_rows.append((row_idx, ref))
            else:
                direct_refs.append(ref)

        indexed_rows.sort(key=lambda item: item[0])
        terms = []
        if indexed_rows:
            start_idx, start_ref = indexed_rows[0]
            prev_idx, prev_ref = indexed_rows[0]
            for row_idx, ref in indexed_rows[1:]:
                if row_idx == prev_idx + 1:
                    prev_idx = row_idx
                    prev_ref = ref
                    continue

                terms.append(start_ref if start_idx == prev_idx else f'{start_ref}:{prev_ref}')
                start_idx = row_idx
                start_ref = ref
                prev_idx = row_idx
                prev_ref = ref

            terms.append(start_ref if start_idx == prev_idx else f'{start_ref}:{prev_ref}')

        terms.extend(direct_refs)
        return terms

    def _tb_terms_expr(self, terms):
        if not terms:
            return '0'
        if len(terms) == 1:
            term = terms[0]
            return term if ':' not in term else f'SUM({term})'
        return f"SUM({','.join(terms)})"

    def _tb_row_sum_formula_by_ref(self, tb_context, criteria, ref_key, *, negate=False, absolute=False):
        row_matches = self._tb_match_context_rows(tb_context, criteria)
        if not row_matches:
            expr = '0'
        else:
            terms = self._tb_row_reference_terms(row_matches, ref_key)
            expr = self._tb_terms_expr(terms)

        if len(expr) > 7800:
            return None

        if absolute:
            expr = f'ABS({expr})'
        if negate:
            expr = f'-({expr})'
        return f'={expr}'

    def _tb_row_sum_formula(self, tb_context, criteria, *, is_prior=False, negate=False, absolute=False):
        ref_key = 'prior_balance_ref' if is_prior else 'current_balance_ref'
        return self._tb_row_sum_formula_by_ref(
            tb_context,
            criteria,
            ref_key,
            negate=negate,
            absolute=absolute,
        )

    def _tb_sumifs_formula(self, tb_context, criteria, *, is_prior=False, negate=False, absolute=False):
        if not tb_context or criteria is None:
            return None
        value_range = tb_context.get('prior_balance_range_ref') if is_prior else tb_context.get('current_balance_range_ref')
        code_range = tb_context.get('code_range_ref')
        if not value_range or not code_range:
            return None

        escaped_criteria = str(criteria).replace('"', '""')
        expr = f'SUMIFS({value_range},{code_range},"{escaped_criteria}")'
        if absolute:
            expr = f'ABS({expr})'
        if negate:
            expr = f'-({expr})'
        return f'={expr}'

    def _tb_prefix_balance_formula(self, tb_context, prefix, *, is_prior=False, negate=False, absolute=False):
        if not prefix:
            return None
        direct_formula = self._tb_row_sum_formula(
            tb_context,
            f'{prefix}*',
            is_prior=is_prior,
            negate=negate,
            absolute=absolute,
        )
        if direct_formula:
            return direct_formula
        return self._tb_sumifs_formula(
            tb_context,
            f'{prefix}*',
            is_prior=is_prior,
            negate=negate,
            absolute=absolute,
        )

    def _tb_sum_prefix_balance_formula(self, tb_context, prefixes, *, is_prior=False, negate=False, absolute=False):
        if not tb_context:
            return None
        terms = []
        for prefix in prefixes or []:
            term_formula = self._tb_prefix_balance_formula(tb_context, prefix, is_prior=is_prior)
            if not term_formula:
                continue
            terms.append(f'({self._tb_formula_expr(term_formula)})')

        expr = '+'.join(terms) if terms else '0'
        if absolute:
            expr = f'ABS({expr})'
        if negate:
            expr = f'-({expr})'
        return f'={expr}'

    def _tb_exact_balance_formula(self, tb_context, code, *, is_prior=False, negate=False):
        if not code:
            return None
        direct_formula = self._tb_row_sum_formula(
            tb_context,
            str(code),
            is_prior=is_prior,
            negate=negate,
        )
        if direct_formula:
            return direct_formula
        return self._tb_sumifs_formula(tb_context, str(code), is_prior=is_prior, negate=negate)

    def _tb_prefix_movement_formula(self, tb_context, prefix, *, is_prior=False, negate=False):
        if not prefix:
            return None

        criteria = f'{prefix}*'
        debit_ref_key = 'prior_debit_ref' if is_prior else 'current_debit_ref'
        credit_ref_key = 'prior_credit_ref' if is_prior else 'current_credit_ref'
        debit_formula = self._tb_row_sum_formula_by_ref(tb_context, criteria, debit_ref_key)
        credit_formula = self._tb_row_sum_formula_by_ref(tb_context, criteria, credit_ref_key)
        if not debit_formula or not credit_formula:
            return None

        expr = f"({self._tb_formula_expr(debit_formula)})-({self._tb_formula_expr(credit_formula)})"
        if negate:
            expr = f'-({expr})'
        return f'={expr}'

    def _tb_exact_movement_formula(self, tb_context, code, *, is_prior=False, negate=False):
        if not code:
            return None

        criteria = str(code)
        debit_ref_key = 'prior_debit_ref' if is_prior else 'current_debit_ref'
        credit_ref_key = 'prior_credit_ref' if is_prior else 'current_credit_ref'
        debit_formula = self._tb_row_sum_formula_by_ref(tb_context, criteria, debit_ref_key)
        credit_formula = self._tb_row_sum_formula_by_ref(tb_context, criteria, credit_ref_key)
        if not debit_formula or not credit_formula:
            return None

        expr = f"({self._tb_formula_expr(debit_formula)})-({self._tb_formula_expr(credit_formula)})"
        if negate:
            expr = f'-({expr})'
        return f'={expr}'

    def _tb_sum_prefix_movement_formula(self, tb_context, prefixes, *, is_prior=False, negate=False):
        if not tb_context:
            return None

        terms = []
        for prefix in prefixes or []:
            formula = self._tb_prefix_movement_formula(tb_context, prefix, is_prior=is_prior)
            if not formula:
                continue
            terms.append(f"({self._tb_formula_expr(formula)})")

        expr = '+'.join(terms) if terms else '0'
        if negate:
            expr = f'-({expr})'
        return f'={expr}'

    def _tb_prefix_opening_formula(self, tb_context, prefix, *, is_prior=False, negate=False):
        if not prefix:
            return None

        ref_key = 'prior_opening_ref' if is_prior else 'current_opening_ref'
        return self._tb_row_sum_formula_by_ref(tb_context, f'{prefix}*', ref_key, negate=negate)

    def _tb_exact_opening_formula(self, tb_context, code, *, is_prior=False, negate=False):
        if not code:
            return None

        ref_key = 'prior_opening_ref' if is_prior else 'current_opening_ref'
        return self._tb_row_sum_formula_by_ref(tb_context, str(code), ref_key, negate=negate)

    def _tb_socf_metric_formula(self, tb_context, metric_key, *, is_prior=False):
        if not tb_context:
            return None

        if metric_key == 'socf_depreciation_movement':
            return self._tb_prefix_movement_formula(tb_context, '5114', is_prior=is_prior)

        if metric_key == 'socf_eosb_adjustment':
            eosb_movement = self._tb_exact_movement_formula(tb_context, '21040101', is_prior=is_prior, negate=True)
            if not eosb_movement:
                return None
            return f"=MAX({self._tb_formula_expr(eosb_movement)},0)"

        if metric_key == 'socf_change_current_assets':
            return self._tb_sum_prefix_movement_formula(
                tb_context,
                ['120101', '120201', '120202', '120301', '120302', '120303', '120304', '120501', '120502'],
                is_prior=is_prior,
                negate=True,
            )

        if metric_key == 'socf_change_current_liabilities':
            return self._tb_sum_prefix_movement_formula(
                tb_context,
                self._SOCF_CURRENT_LIABILITY_PREFIXES,
                is_prior=is_prior,
                negate=True,
            )

        if metric_key == 'socf_property_investing':
            return self._tb_prefix_movement_formula(tb_context, '1101', is_prior=is_prior, negate=True)

        if metric_key == 'socf_security_deposit':
            return self._tb_exact_movement_formula(
                tb_context,
                self._SOCF_SECURITY_DEPOSIT_ACCOUNT_CODE,
                is_prior=is_prior,
                negate=True,
            )

        if metric_key == 'socf_related_party_loan':
            return self._tb_prefix_movement_formula(
                tb_context,
                self._SOCF_RELATED_PARTY_LOAN_PREFIX,
                is_prior=is_prior,
                negate=True,
            )

        if metric_key == 'socf_corporate_tax_paid':
            tax_liability_movement = self._tb_prefix_movement_formula(
                tb_context,
                '220401',
                is_prior=is_prior,
                negate=True,
            )
            if not tax_liability_movement:
                return None
            expense_terms = []
            for code in self._SOCF_CT_EXPENSE_ACCOUNT_CODES:
                expense_formula = self._tb_exact_movement_formula(tb_context, code, is_prior=is_prior)
                if not expense_formula:
                    continue
                expense_terms.append(f"({self._tb_formula_expr(expense_formula)})")
            expense_expr = '+'.join(expense_terms) if expense_terms else '0'
            return f"=({self._tb_formula_expr(tax_liability_movement)})-({expense_expr})"

        if metric_key == 'socf_eosb_paid':
            eosb_movement = self._tb_exact_movement_formula(
                tb_context,
                self._SOCF_EOSB_LIABILITY_CODE,
                is_prior=is_prior,
                negate=True,
            )
            if not eosb_movement:
                return None
            return f"=MIN({self._tb_formula_expr(eosb_movement)},0)"

        if metric_key == 'socf_cash_equivalent_opening':
            opening_1204 = self._tb_prefix_opening_formula(tb_context, '1204', is_prior=is_prior)
            opening_12040101 = self._tb_exact_opening_formula(tb_context, '12040101', is_prior=is_prior)
            opening_1206 = self._tb_prefix_opening_formula(tb_context, '1206', is_prior=is_prior)
            if not opening_1204 or not opening_12040101 or not opening_1206:
                return None
            return (
                f"=({self._tb_formula_expr(opening_1204)})"
                f"-({self._tb_formula_expr(opening_12040101)})"
                f"-({self._tb_formula_expr(opening_1206)})"
            )

        if metric_key == 'share_capital_movement':
            return self._tb_exact_movement_formula(tb_context, '31010101', is_prior=is_prior, negate=True)

        if metric_key == 'owner_current_account_movement':
            return self._tb_exact_movement_formula(tb_context, '12040101', is_prior=is_prior, negate=True)

        if metric_key == 'dividend_paid':
            return self._tb_exact_movement_formula(tb_context, '31010202', is_prior=is_prior, negate=True)

        return None

    def _tb_sofp_cash_bank_formula(self, tb_context, *, is_prior=False):
        total_cash_formula = self._tb_sum_prefix_balance_formula(
            tb_context,
            ['120401', '120402', '120601'],
            is_prior=is_prior,
        )
        owner_current_formula = self._tb_exact_balance_formula(tb_context, '12040101', is_prior=is_prior)
        if not total_cash_formula or not owner_current_formula:
            return None
        return f"=({self._tb_formula_expr(total_cash_formula)})-({self._tb_formula_expr(owner_current_formula)})"

    def _tb_soci_metric_formula(self, tb_context, metric_key, *, is_prior=False):
        metric_specs = {
            'soci_revenue': {'prefixes': ['410101'], 'negate': True},
            'soci_revenue_related': {'prefixes': ['410201'], 'negate': True},
            'soci_direct_cost': {'prefixes': ['510101', '510102', '510103', '510104']},
            'soci_director_salary': {'prefixes': ['510701']},
            'soci_salary_office': {'prefixes': ['510801']},
            'soci_salary_coaching': {'prefixes': ['510802']},
            'soci_salary_benefits': {'prefixes': ['510803']},
            'soci_salary_bonus': {'prefixes': ['510804']},
            'soci_salary_welfare': {'prefixes': ['510805']},
            'soci_advertising': {'prefixes': ['510201']},
            'soci_audit_fee': {'prefixes': ['510901']},
            'soci_accounting_fee': {'prefixes': ['510902']},
            'soci_depreciation_expense': {'prefixes': ['511401']},
            'soci_amortization_expense': {'prefixes': ['511402']},
            'soci_legal_gov_fee': {'prefixes': ['510601']},
            'soci_trade_license': {'prefixes': ['512601']},
            'soci_establishment_card': {'prefixes': ['512602']},
            'soci_visa_fee': {'prefixes': ['512603']},
            'soci_insurance': {'prefixes': ['512501']},
            'soci_office_expense': {
                'prefixes': ['510401', '511001', '511002', '511101', '511201', '511202', '511501', '511701', '511801', '512801'],
            },
            'soci_bank_charges': {'prefixes': ['512301']},
            'soci_exchange_loss': {'prefixes': ['512201']},
            'soci_other_expenses': {
                'prefixes': ['510301', '510501', '511301', '511302', '511601', '511901', '512001', '512101', '512202', '512401', '512701'],
            },
            'soci_other_income': {
                'prefixes': ['410301', '410302', '410303', '410304', '410305', '410306', '410307', '410308', '410309', '410310'],
                'absolute': True,
            },
            'soci_investment_gain_loss': {'prefixes': ['5201'], 'negate': True},
        }

        spec = metric_specs.get(metric_key)
        if not spec:
            return None

        return self._tb_sum_prefix_balance_formula(
            tb_context,
            spec.get('prefixes') or [],
            is_prior=is_prior,
            negate=bool(spec.get('negate')),
            absolute=bool(spec.get('absolute')),
        )

    def _build_sofp_live_link_lines(self, metric_refs, show_prior):
        tb_context = metric_refs.get('__tb_link_context__') or {}
        tb_rows = tb_context.get('rows') or []
        if not tb_context or not tb_rows:
            return {}

        lines_by_section = {
            'non_current_assets': [],
            'current_assets': [],
            'non_current_liabilities': [],
            'current_liabilities': [],
        }

        for row in tb_rows:
            code = row.get('code') or ''
            section = self._ar_sofp_account_section(code)
            if section not in lines_by_section:
                continue

            current_value = float(row.get('current_closing') or 0.0)
            prior_value = float(row.get('prior_closing') or 0.0)
            negate = section in ('non_current_liabilities', 'current_liabilities')
            if negate:
                current_value = -current_value
                prior_value = -prior_value
            if not (current_value or (show_prior and prior_value)):
                continue

            lines_by_section[section].append({
                'label': self._ar_format_sofp_account_label(row),
                'current_formula': self._tb_exact_balance_formula(tb_context, code, negate=negate),
                'prior_formula': self._tb_exact_balance_formula(tb_context, code, is_prior=True, negate=negate),
            })

        cash_current_formula = self._tb_sofp_cash_bank_formula(tb_context, is_prior=False)
        cash_prior_formula = self._tb_sofp_cash_bank_formula(tb_context, is_prior=True)
        cash_line = {
            'label': 'Cash and bank balances',
            'current_formula': cash_current_formula or self._metric_ref(metric_refs, 'sofp_cash_bank', 'current'),
            'prior_formula': cash_prior_formula or self._metric_ref(metric_refs, 'sofp_cash_bank', 'prior'),
        }
        lines_by_section['current_assets'].append(cash_line)
        return lines_by_section

    def _get_sofp_link_rows(self, sheet):
        if sheet is None:
            return {}
        return {
            'non_current_assets': self._find_row_by_label(sheet, 'Non-current assets'),
            'total_non_current_assets': self._find_row_by_label(sheet, 'Total non-current assets'),
            'current_assets': self._find_row_by_label(sheet, 'Current assets'),
            'total_current_assets': self._find_row_by_label(sheet, 'Total current assets'),
            'total_assets': self._find_row_by_label(sheet, 'Total assets'),
            'equity': self._find_row_by_label(sheet, 'Equity'),
            'total_equity': self._find_row_by_keys(sheet, ['Total Equity', 'Total equity']),
            'non_current_liabilities': self._find_row_by_label(sheet, 'Non-current liabilities'),
            'current_liabilities': self._find_row_by_label(sheet, 'Current liabilities'),
            'total_liabilities': self._find_row_by_keys(sheet, ['Total Liabilities', 'Total liabilities']),
            'total_equity_liabilities': self._find_row_by_keys(
                sheet,
                ['Total Equity and Liabilities', 'Total equity and liabilities'],
            ),
        }

    def _build_sofp_equity_formula_lines(self, workbook, sofp_sheet=None):
        soce_sheet = self._get_sheet_if_exists(workbook, 'SOCE')
        if soce_sheet is None:
            return []

        soce_layout = self._get_soce_link_layout(soce_sheet)
        cols = soce_layout.get('cols') or {}
        rows = soce_layout.get('rows') or {}
        prior_close_row = rows.get('prior_close')
        current_close_row = rows.get('current_close')
        if not prior_close_row or not current_close_row:
            return []

        def _ref(col_key, row_idx):
            col_idx = cols.get(col_key)
            if not col_idx or not row_idx:
                return None
            return f'=SOCE!{get_column_letter(col_idx)}{row_idx}'

        line_specs = [
            {
                'default_label': 'Share capital',
                'col_key': 'share_capital',
                'aliases': ('Share capital', 'Share Capital'),
            },
            {
                'default_label': 'Retained earnings',
                'col_key': 'retained_earnings',
                'aliases': ('Retained earnings', 'Retained earning'),
            },
            {
                'default_label': 'Owner current account',
                'col_key': 'owner_current_account',
                'aliases': ('Owner current account', "Owner's current account"),
            },
            {
                'default_label': 'Statutory reserves',
                'col_key': 'statutory_reserves',
                'aliases': ('Statutory reserves', 'Statutory reserve'),
            },
        ]
        existing_lines = []
        if sofp_sheet is not None:
            sofp_rows = self._get_sofp_link_rows(sofp_sheet)
            start_row = sofp_rows.get('equity')
            end_row = sofp_rows.get('total_equity')
            if start_row and end_row and end_row > start_row:
                for row_idx in range(start_row + 1, end_row):
                    row_label = sofp_sheet.cell(row=row_idx, column=1).value
                    normalized_label = self._normalize_key(row_label or '')
                    if not normalized_label:
                        continue
                    for spec in line_specs:
                        aliases = {self._normalize_key(alias) for alias in spec['aliases']}
                        if normalized_label not in aliases:
                            continue
                        existing_lines.append((row_label or spec['default_label'], spec['col_key']))
                        break

        lines = existing_lines or [
            (spec['default_label'], spec['col_key'])
            for spec in line_specs[:3]
        ]
        return [
            {
                'label': label,
                'current_formula': _ref(col_key, current_close_row),
                'prior_formula': _ref(col_key, prior_close_row),
            }
            for label, col_key in lines
            if cols.get(col_key)
        ]

    def _link_specific_sofp_equity_rows_to_soce(self, sofp_sheet, workbook, *, show_prior):
        if sofp_sheet is None:
            return
        soce_sheet = self._get_sheet_if_exists(workbook, 'SOCE')
        if soce_sheet is None:
            return

        soce_layout = self._get_soce_link_layout(soce_sheet)
        cols = soce_layout.get('cols') or {}
        rows = soce_layout.get('rows') or {}
        prior_close_row = rows.get('prior_close')
        current_close_row = rows.get('current_close')
        if not prior_close_row or not current_close_row:
            return

        mappings = (
            ('share_capital', ['Share capital', 'Share Capital']),
            ('retained_earnings', ['Retained earnings', 'Retained earning']),
            ('owner_current_account', ['Owner current account', "Owner's current account"]),
        )
        for col_key, labels in mappings:
            soce_col = cols.get(col_key)
            if not soce_col:
                continue
            sofp_row = self._find_row_by_keys(sofp_sheet, labels)
            if not sofp_row:
                continue
            soce_col_letter = get_column_letter(soce_col)
            sofp_sheet[f'B{sofp_row}'] = f'=SOCE!{soce_col_letter}{current_close_row}'
            if show_prior:
                sofp_sheet[f'C{sofp_row}'] = f'=SOCE!{soce_col_letter}{prior_close_row}'
            else:
                sofp_sheet[f'C{sofp_row}'] = None

    def _get_soce_link_layout(self, sheet):
        if sheet is None:
            return {'rows': {}, 'cols': {}}

        header_end_row = min(sheet.max_row, 10)
        balance_rows = []
        for row_idx in range(1, sheet.max_row + 1):
            normalized_value = self._normalize_key(sheet.cell(row=row_idx, column=1).value or '')
            if 'balanceasat' in normalized_value or normalized_value == 'balancecf':
                balance_rows.append(row_idx)

        explicit_prior_close = (
            self._find_row_by_label(sheet, 'Balance as at end of period')
            or self._find_row_containing_text(sheet, 'Balance as at', occurrence=2)
        )
        explicit_current_close = (
            self._find_row_by_label(sheet, 'Balance c/f')
            or self._find_row_containing_text(sheet, 'Balance as at', occurrence=3)
        )
        prior_close_row = explicit_prior_close
        current_close_row = explicit_current_close
        if not prior_close_row and len(balance_rows) >= 2:
            prior_close_row = balance_rows[1]
        if not current_close_row:
            if len(balance_rows) >= 3:
                current_close_row = balance_rows[2]
            elif len(balance_rows) >= 2:
                current_close_row = balance_rows[-1]

        opening_balance_row = self._find_row_by_label(sheet, 'Balance as at start of period')
        if not opening_balance_row and balance_rows:
            opening_balance_row = balance_rows[0]

        cols = {
            'share_capital': self._find_column_by_label(sheet, 'Share capital', start_row=1, end_row=header_end_row),
            'owner_current_account': self._find_column_by_keys(
                sheet,
                ["Owner's current account", 'Owner current account'],
                start_row=1,
                end_row=header_end_row,
            ),
            'retained_earnings': self._find_column_by_label(
                sheet, 'Retained earnings', start_row=1, end_row=header_end_row
            ),
            'statutory_reserves': self._find_column_by_keys(
                sheet, ['Statutory reserves', 'Statutory reserve'], start_row=1, end_row=header_end_row
            ),
            'total': self._find_column_by_label(sheet, 'Total', start_row=1, end_row=header_end_row),
        }
        rows = {
            'opening_balance': opening_balance_row,
            'prior_share_movement': self._find_row_by_label(sheet, 'Paid up capital', occurrence=1),
            'prior_owner_movement': self._find_row_by_keys(
                sheet,
                ["Movement in owner's current account", 'Movement in owner current account', "Net movement in owner's current account"],
                start_row=1,
                end_row=sheet.max_row,
            ),
            'prior_net_profit': self._find_row_by_keys(
                sheet, ['Net profit / (loss)', 'Net profit', 'Net profit for the period'], start_row=1, end_row=sheet.max_row
            ),
            'prior_transfer': self._find_row_by_keys(
                sheet, ['Transfer to reserve', 'Transfer to statutory reserve'], start_row=1, end_row=sheet.max_row
            ),
            'prior_dividend': self._find_row_by_label(sheet, 'Dividend paid', occurrence=1),
            'prior_close': prior_close_row,
            'current_share_movement': self._find_row_by_label(sheet, 'Paid up capital', occurrence=2),
            'current_owner_movement': self._find_row_by_keys(
                sheet,
                ["Movement in owner's current account", 'Movement in owner current account', "Net movement in owner's current account"],
                start_row=(prior_close_row or 1) + 1,
                end_row=sheet.max_row,
            ),
            'current_net_profit': self._find_row_by_keys(
                sheet,
                ['Net profit / (loss)', 'Net profit', 'Net profit for the period'],
                start_row=(prior_close_row or 1) + 1,
                end_row=sheet.max_row,
            ),
            'current_transfer': self._find_row_by_keys(
                sheet,
                ['Transfer to reserve', 'Transfer to statutory reserve'],
                start_row=(prior_close_row or 1) + 1,
                end_row=sheet.max_row,
                occurrence=1,
            ),
            'current_dividend': self._find_row_by_label(sheet, 'Dividend paid', occurrence=2),
            'current_close': current_close_row,
        }
        return {'rows': rows, 'cols': cols}

    def _get_socf_link_rows(self, sheet):
        if sheet is None:
            return {}
        return {
            'net_profit': self._find_row_by_keys(sheet, ['Net profit for the year', 'Net profit for the period']),
            'operating_before_wc': self._find_row_by_keys(
                sheet,
                [
                    'Operating cash flows before changes in working capital',
                    'Operating cash flows before working capital changes',
                ],
            ),
            'change_current_assets': self._find_row_by_keys(
                sheet,
                ['(Increase) / decrease in current assets', 'Increase / decrease in current assets'],
            ),
            'change_current_liabilities': self._find_row_by_keys(
                sheet,
                ['Increase / (decrease) in current liabilities', 'Increase / decrease in current liabilities'],
            ),
            'net_operations': self._find_row_by_keys(
                sheet,
                ['Net cash generated from operations', 'Net cash (used in) operations'],
            ),
            'property': self._find_row_by_label(sheet, 'Property, plant and equipment'),
            'security_deposit': self._find_row_by_label(sheet, 'Security deposit'),
            'net_investing': self._find_row_by_keys(
                sheet,
                ['Net cash generated from investing activities', 'Net cash (used in) investing activities'],
            ),
            'paid_up_capital': self._find_row_by_label(sheet, 'Paid up capital'),
            'dividend_paid': self._find_row_by_label(sheet, 'Dividend paid'),
            'owner_current_account': self._find_row_by_label(sheet, 'Owner current account'),
            'related_party_loan': self._find_row_by_label(sheet, 'Loan from related party'),
            'net_financing': self._find_row_by_keys(
                sheet,
                ['Net cash generated from financing activities', 'Net cash (used in) financing activities'],
            ),
            'net_increase': self._find_row_by_keys(
                sheet,
                [
                    'Net increase in cash and cash equivalents',
                    'Net increase / (decrease) in cash and cash equivalents',
                ],
            ),
            'cash_beginning': self._find_row_by_keys(
                sheet,
                ['Cash and cash equivalents, beginning of the period', 'Cash and cash equivalents at the beginning'],
            ),
            'cash_end': self._find_row_by_keys(
                sheet,
                ['Cash and cash equivalents, end of the period', 'Cash and cash equivalents at the end'],
            ),
        }

    def _stage_g_hide_prior_year_live_link_columns(self, stage_c_context, selected_sheet_names):
        periods = self._get_reporting_periods()
        if periods.get('show_prior_year'):
            return

        workbook = stage_c_context.get('workbook')
        if not workbook:
            return

        for sheet_name in ('SOFP', 'SOCI', 'SOCF'):
            if sheet_name not in selected_sheet_names or sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            sheet['C6'] = None
            for row_idx in range(7, sheet.max_row + 1):
                sheet[f'C{row_idx}'] = None

    def _link_sofp_to_tb_metrics(self, workbook, metric_refs, tb_sheet_name):
        sheet = self._get_sheet_if_exists(workbook, 'SOFP')
        if sheet is None:
            return
        show_prior = bool(self._get_reporting_periods().get('show_prior_year'))

        rows = self._get_sofp_link_rows(sheet)
        if not all(rows.values()):
            return

        sofp_lines = self._build_sofp_live_link_lines(metric_refs, show_prior)
        equity_lines = self._build_sofp_equity_formula_lines(workbook, sofp_sheet=sheet)

        non_current_asset_rows = self._ar_ensure_detail_rows(
            sheet, 'Non-current assets', ['Total non-current assets'], len(sofp_lines.get('non_current_assets', []))
        )
        self._write_dynamic_formula_rows(
            sheet, non_current_asset_rows, sofp_lines.get('non_current_assets', []), show_prior
        )

        current_asset_rows = self._ar_ensure_detail_rows(
            sheet, 'Current assets', ['Total current assets'], len(sofp_lines.get('current_assets', []))
        )
        self._write_dynamic_formula_rows(
            sheet, current_asset_rows, sofp_lines.get('current_assets', []), show_prior
        )

        equity_rows = self._ar_ensure_detail_rows(
            sheet, 'Equity', ['Total Equity', 'Total equity'], len(equity_lines)
        )
        self._write_dynamic_formula_rows(sheet, equity_rows, equity_lines, show_prior)

        non_current_liability_rows = self._ar_ensure_detail_rows(
            sheet, 'Non-current liabilities', ['Current liabilities'], len(sofp_lines.get('non_current_liabilities', []))
        )
        self._write_dynamic_formula_rows(
            sheet, non_current_liability_rows, sofp_lines.get('non_current_liabilities', []), show_prior
        )

        current_liability_rows = self._ar_ensure_detail_rows(
            sheet, 'Current liabilities', ['Total Liabilities'], len(sofp_lines.get('current_liabilities', []))
        )
        self._write_dynamic_formula_rows(
            sheet, current_liability_rows, sofp_lines.get('current_liabilities', []), show_prior
        )

        rows = self._get_sofp_link_rows(sheet)
        if not all(rows.values()):
            return

        self._link_specific_sofp_equity_rows_to_soce(sheet, workbook, show_prior=show_prior)

        total_non_current_row = rows['total_non_current_assets']
        total_current_assets_row = rows['total_current_assets']
        total_assets_row = rows['total_assets']
        total_equity_row = rows['total_equity']
        total_liabilities_row = rows['total_liabilities']
        total_equity_liabilities_row = rows['total_equity_liabilities']

        sheet[f'B{total_non_current_row}'] = f"=SUM(B{rows['non_current_assets']}:B{total_non_current_row - 1})"
        sheet[f'C{total_non_current_row}'] = f"=SUM(C{rows['non_current_assets']}:C{total_non_current_row - 1})"
        sheet[f'B{total_current_assets_row}'] = f"=SUM(B{rows['current_assets']}:B{total_current_assets_row - 1})"
        sheet[f'C{total_current_assets_row}'] = f"=SUM(C{rows['current_assets']}:C{total_current_assets_row - 1})"
        sheet[f'B{total_assets_row}'] = f"=B{total_current_assets_row}+B{total_non_current_row}"
        sheet[f'C{total_assets_row}'] = f"=C{total_current_assets_row}+C{total_non_current_row}"
        sheet[f'B{total_equity_row}'] = f"=SUM(B{rows['equity']}:B{total_equity_row - 1})"
        sheet[f'C{total_equity_row}'] = f"=SUM(C{rows['equity']}:C{total_equity_row - 1})"
        sheet[f'B{total_liabilities_row}'] = f"=SUM(B{rows['non_current_liabilities']}:B{total_liabilities_row - 1})"
        sheet[f'C{total_liabilities_row}'] = f"=SUM(C{rows['non_current_liabilities']}:C{total_liabilities_row - 1})"
        sheet[f'B{total_equity_liabilities_row}'] = f"=B{total_liabilities_row}+B{total_equity_row}"
        sheet[f'C{total_equity_liabilities_row}'] = f"=C{total_liabilities_row}+C{total_equity_row}"

        diff_row = total_equity_liabilities_row + 2
        if diff_row <= sheet.max_row:
            sheet[f'B{diff_row}'] = f"=B{total_equity_liabilities_row}-B{total_assets_row}"

    def _link_soci_to_tb_metrics(self, workbook, metric_refs, tb_sheet_name):
        sheet = self._get_sheet_if_exists(workbook, 'SOCI')
        if sheet is None:
            return

        operating_lines = self._ensure_soci_hierarchy_rows(sheet)
        if not operating_lines:
            return

        row_revenue = self._find_row_by_label(sheet, 'Revenue')
        row_revenue_related = self._find_row_by_label(sheet, 'Revenue - related party')
        row_direct_cost = self._find_row_by_label(sheet, 'Direct cost')
        row_gross_profit = self._find_row_by_label(sheet, 'Gross profit')
        row_operating_heading = self._find_row_by_label(sheet, 'Operating expenses')
        row_total_operating = self._find_row_by_label(sheet, 'Total operating expenses')
        row_investment_gain_loss = self._find_row_by_label(sheet, 'Gain / (loss) on investment')
        row_other_income = self._find_row_by_label(sheet, 'Other income')
        row_net_profit = self._find_row_by_keys(sheet, ['Net profit / (loss)', 'Net profit'])
        if not all((
            row_revenue,
            row_revenue_related,
            row_direct_cost,
            row_gross_profit,
            row_operating_heading,
            row_total_operating,
            row_other_income,
            row_net_profit,
        )):
            return

        tb_context = metric_refs.get('__tb_link_context__') or {}

        def soci_formula(metric_key, period):
            direct_formula = self._tb_soci_metric_formula(tb_context, metric_key, is_prior=(period == 'prior'))
            if direct_formula:
                return direct_formula
            return self._metric_ref(metric_refs, metric_key, period)

        def soci_expr(metric_key, period):
            return self._tb_formula_expr(soci_formula(metric_key, period))

        sheet[f'B{row_revenue}'] = soci_formula('soci_revenue', 'current')
        sheet[f'C{row_revenue}'] = soci_formula('soci_revenue', 'prior')
        sheet[f'B{row_revenue_related}'] = soci_formula('soci_revenue_related', 'current')
        sheet[f'C{row_revenue_related}'] = soci_formula('soci_revenue_related', 'prior')
        sheet[f'B{row_direct_cost}'] = soci_formula('soci_direct_cost', 'current')
        sheet[f'C{row_direct_cost}'] = soci_formula('soci_direct_cost', 'prior')

        sheet[f'B{row_operating_heading}'] = None
        sheet[f'C{row_operating_heading}'] = None
        for offset, line in enumerate(operating_lines, start=1):
            row_idx = row_operating_heading + offset
            if not line.get('metric'):
                sheet[f'B{row_idx}'] = None
                sheet[f'C{row_idx}'] = None
                continue
            sheet[f'B{row_idx}'] = soci_formula(line['metric'], 'current')
            sheet[f'C{row_idx}'] = soci_formula(line['metric'], 'prior')

        if row_investment_gain_loss:
            sheet[f'B{row_other_income}'] = soci_formula('soci_other_income', 'current')
            sheet[f'C{row_other_income}'] = soci_formula('soci_other_income', 'prior')
            sheet[f'B{row_investment_gain_loss}'] = soci_formula('soci_investment_gain_loss', 'current')
            sheet[f'C{row_investment_gain_loss}'] = soci_formula('soci_investment_gain_loss', 'prior')
        else:
            other_income_current = soci_expr('soci_other_income', 'current')
            other_income_prior = soci_expr('soci_other_income', 'prior')
            investment_current = soci_expr('soci_investment_gain_loss', 'current')
            investment_prior = soci_expr('soci_investment_gain_loss', 'prior')
            sheet[f'B{row_other_income}'] = f'=({other_income_current})+({investment_current})'
            sheet[f'C{row_other_income}'] = f'=({other_income_prior})+({investment_prior})'

        sheet[f'B{row_gross_profit}'] = f"=B{row_revenue}+B{row_revenue_related}-B{row_direct_cost}"
        sheet[f'C{row_gross_profit}'] = f"=C{row_revenue}+C{row_revenue_related}-C{row_direct_cost}"
        sheet[f'B{row_total_operating}'] = f"=SUM(B{row_operating_heading + 1}:B{row_total_operating - 1})"
        sheet[f'C{row_total_operating}'] = f"=SUM(C{row_operating_heading + 1}:C{row_total_operating - 1})"
        if row_investment_gain_loss:
            sheet[f'B{row_net_profit}'] = (
                f"=B{row_gross_profit}-B{row_total_operating}+B{row_investment_gain_loss}+B{row_other_income}"
            )
            sheet[f'C{row_net_profit}'] = (
                f"=C{row_gross_profit}-C{row_total_operating}+C{row_investment_gain_loss}+C{row_other_income}"
            )
        else:
            sheet[f'B{row_net_profit}'] = f"=B{row_gross_profit}-B{row_total_operating}+B{row_other_income}"
            sheet[f'C{row_net_profit}'] = f"=C{row_gross_profit}-C{row_total_operating}+C{row_other_income}"

    def _link_soce_to_tb_metrics(self, workbook, metric_refs, tb_sheet_name):
        sheet = self._get_sheet_if_exists(workbook, 'SOCE')
        if sheet is None:
            return

        self._ensure_soce_dividend_rows(sheet, metric_refs=metric_refs)
        soci_net_profit_row = self._get_soci_net_profit_row(workbook)
        soce_layout = self._get_soce_link_layout(sheet)
        rows = soce_layout.get('rows') or {}
        cols = soce_layout.get('cols') or {}
        value_col_keys = ('share_capital', 'owner_current_account', 'retained_earnings', 'statutory_reserves')

        required_rows = (
            'opening_balance',
            'prior_share_movement',
            'prior_owner_movement',
            'prior_net_profit',
            'prior_transfer',
            'prior_close',
            'current_share_movement',
            'current_owner_movement',
            'current_net_profit',
            'current_transfer',
            'current_close',
        )
        if not all(rows.get(key) for key in required_rows):
            return
        if not all(cols.get(key) for key in value_col_keys):
            return

        def set_formula(row_key, col_key, formula):
            row_idx = rows.get(row_key)
            col_idx = cols.get(col_key)
            if not row_idx or not col_idx:
                return
            sheet.cell(row=row_idx, column=col_idx, value=formula)

        def col_letter(col_key):
            return get_column_letter(cols[col_key])

        share_close_prior = self._metric_expr(metric_refs, 'share_capital_close', 'prior')
        share_mov_prior = self._metric_expr(metric_refs, 'share_capital_movement', 'prior')
        owner_close_prior = self._metric_expr(metric_refs, 'owner_current_account_close', 'prior')
        owner_mov_prior = self._metric_expr(metric_refs, 'owner_current_account_movement', 'prior')
        retained_open_prior_ref = self._metric_ref(metric_refs, 'retained_earnings_open', 'prior')
        statutory_close_prior = self._metric_expr(metric_refs, 'statutory_reserve_close', 'prior')
        statutory_transfer_prior = self._metric_expr(metric_refs, 'statutory_transfer', 'prior')
        statutory_transfer_current = self._metric_expr(metric_refs, 'statutory_transfer', 'current')
        prior_dividend_row = rows.get('prior_dividend')
        current_dividend_row = rows.get('current_dividend')

        set_formula('opening_balance', 'share_capital', f"={share_close_prior}-{share_mov_prior}")
        set_formula('opening_balance', 'owner_current_account', f"={owner_close_prior}-{owner_mov_prior}")
        set_formula('opening_balance', 'retained_earnings', retained_open_prior_ref)
        set_formula('opening_balance', 'statutory_reserves', f"={statutory_close_prior}-{statutory_transfer_prior}")

        set_formula('prior_share_movement', 'share_capital', self._metric_ref(metric_refs, 'share_capital_movement', 'prior'))
        set_formula(
            'prior_owner_movement',
            'owner_current_account',
            self._metric_ref(metric_refs, 'owner_current_account_movement', 'prior'),
        )
        set_formula('prior_net_profit', 'retained_earnings', f'=SOCI!C{soci_net_profit_row}')
        set_formula('prior_transfer', 'retained_earnings', f"=-{statutory_transfer_prior}")
        set_formula('prior_transfer', 'statutory_reserves', self._metric_ref(metric_refs, 'statutory_transfer', 'prior'))
        if prior_dividend_row:
            set_formula('prior_dividend', 'retained_earnings', self._metric_ref(metric_refs, 'dividend_paid', 'prior'))

        set_formula('current_share_movement', 'share_capital', self._metric_ref(metric_refs, 'share_capital_movement', 'current'))
        set_formula(
            'current_owner_movement',
            'owner_current_account',
            self._metric_ref(metric_refs, 'owner_current_account_movement', 'current'),
        )
        set_formula('current_net_profit', 'retained_earnings', f'=SOCI!B{soci_net_profit_row}')
        set_formula('current_transfer', 'retained_earnings', f"=-{statutory_transfer_current}")
        set_formula('current_transfer', 'statutory_reserves', self._metric_ref(metric_refs, 'statutory_transfer', 'current'))
        if current_dividend_row:
            set_formula('current_dividend', 'retained_earnings', self._metric_ref(metric_refs, 'dividend_paid', 'current'))

        for col_key in value_col_keys:
            letter = col_letter(col_key)
            prior_close_end_row = prior_dividend_row or (rows['prior_close'] - 1)
            current_close_end_row = current_dividend_row or (rows['current_close'] - 1)
            sheet.cell(
                row=rows['prior_close'],
                column=cols[col_key],
                value=f'=SUM({letter}{rows["opening_balance"]}:{letter}{prior_close_end_row})',
            )
            sheet.cell(
                row=rows['current_close'],
                column=cols[col_key],
                value=f'=SUM({letter}{rows["prior_close"]}:{letter}{current_close_end_row})',
            )

        total_col = cols.get('total')
        if total_col:
            first_value_col = get_column_letter(min(cols[key] for key in value_col_keys))
            last_value_col = get_column_letter(max(cols[key] for key in value_col_keys))
            total_rows = [
                'opening_balance',
                'prior_share_movement',
                'prior_owner_movement',
                'prior_net_profit',
                'prior_transfer',
                'prior_close',
                'current_share_movement',
                'current_owner_movement',
                'current_net_profit',
                'current_transfer',
                'current_close',
            ]
            if prior_dividend_row:
                total_rows.append('prior_dividend')
            if current_dividend_row:
                total_rows.append('current_dividend')
            for row_key in total_rows:
                row_idx = rows.get(row_key)
                if not row_idx:
                    continue
                sheet.cell(row=row_idx, column=total_col, value=f'=SUM({first_value_col}{row_idx}:{last_value_col}{row_idx})')

    def _link_socf_to_tb_metrics(self, workbook, metric_refs, tb_sheet_name):
        sheet = self._get_sheet_if_exists(workbook, 'SOCF')
        if sheet is None:
            return

        self._ensure_socf_dividend_row(sheet, metric_refs=metric_refs)
        soci_net_profit_row = self._get_soci_net_profit_row(workbook)
        rows = self._get_socf_link_rows(sheet)
        required_rows = (
            'net_profit',
            'operating_before_wc',
            'change_current_assets',
            'change_current_liabilities',
            'net_operations',
            'property',
            'net_investing',
            'paid_up_capital',
            'owner_current_account',
            'net_financing',
            'net_increase',
            'cash_beginning',
            'cash_end',
        )
        if not all(rows.get(key) for key in required_rows):
            return

        def set_formula(row_key, col_letter, formula):
            row_idx = rows.get(row_key)
            if not row_idx:
                return
            sheet[f'{col_letter}{row_idx}'] = formula

        sofp_sheet = self._get_sheet_if_exists(workbook, 'SOFP')
        sofp_rows = self._get_sofp_link_rows(sofp_sheet) if sofp_sheet else {}
        sofp_cash_row = self._find_row_by_label(sofp_sheet, 'Cash and bank balances') if sofp_sheet else None
        sofp_property_row = self._find_row_by_keys(
            sofp_sheet,
            ['Property, plant and equipment', 'Property plant and equipment'],
        ) if sofp_sheet else None
        sofp_security_deposit_row = self._find_row_by_label(sofp_sheet, 'Security deposit') if sofp_sheet else None
        sofp_related_party_loan_row = self._find_row_by_label(sofp_sheet, 'Loan from related party') if sofp_sheet else None
        sofp_owner_current_account_row = self._find_row_by_label(sofp_sheet, 'Owner current account') if sofp_sheet else None
        sofp_share_capital_row = self._find_row_by_keys(
            sofp_sheet,
            ['Share capital', 'Share Capital'],
        ) if sofp_sheet else None

        def _sofp_row_expr(row_idx, period):
            if not row_idx:
                return None
            col_letter = 'C' if period == 'prior' else 'B'
            return f'SOFP!{col_letter}{row_idx}'

        def _sofp_delta_formula(row_idx, period, *, positive_when_increase):
            if period != 'current' or not row_idx:
                return None
            current_expr = _sofp_row_expr(row_idx, 'current')
            prior_expr = _sofp_row_expr(row_idx, 'prior')
            if not current_expr or not prior_expr:
                return None
            if positive_when_increase:
                return f'=({current_expr})-({prior_expr})'
            return f'=({prior_expr})-({current_expr})'

        def _sofp_current_assets_formula(period):
            if period != 'current':
                return None
            total_current_assets_row = sofp_rows.get('total_current_assets')
            if not total_current_assets_row:
                return None

            current_expr = _sofp_row_expr(total_current_assets_row, 'current')
            prior_expr = _sofp_row_expr(total_current_assets_row, 'prior')
            if sofp_cash_row:
                current_expr = f'({current_expr})-({_sofp_row_expr(sofp_cash_row, "current")})'
                prior_expr = f'({prior_expr})-({_sofp_row_expr(sofp_cash_row, "prior")})'
            return f'=({prior_expr})-({current_expr})'

        def _sofp_current_liabilities_formula(period):
            if period != 'current':
                return None
            current_liabilities_row = sofp_rows.get('current_liabilities')
            total_liabilities_row = sofp_rows.get('total_liabilities')
            if not current_liabilities_row or not total_liabilities_row or total_liabilities_row <= (current_liabilities_row + 1):
                return None

            current_expr = f'SUM(SOFP!B{current_liabilities_row + 1}:SOFP!B{total_liabilities_row - 1})'
            prior_expr = f'SUM(SOFP!C{current_liabilities_row + 1}:SOFP!C{total_liabilities_row - 1})'
            for label in ('Short-term loan', 'Lease liability', 'Corporate tax liability'):
                exclude_row = self._find_row_by_label(sofp_sheet, label) if sofp_sheet else None
                if not exclude_row:
                    continue
                if not (current_liabilities_row < exclude_row < total_liabilities_row):
                    continue
                current_expr = f'({current_expr})-({_sofp_row_expr(exclude_row, "current")})'
                prior_expr = f'({prior_expr})-({_sofp_row_expr(exclude_row, "prior")})'
            return f'=({current_expr})-({prior_expr})'

        def socf_sofp_formula(metric_key, period):
            if not sofp_sheet:
                return None

            if metric_key == 'socf_change_current_assets':
                return _sofp_current_assets_formula(period)
            if metric_key == 'socf_change_current_liabilities':
                return _sofp_current_liabilities_formula(period)
            if metric_key == 'socf_property_investing':
                return _sofp_delta_formula(sofp_property_row, period, positive_when_increase=False)
            if metric_key == 'socf_security_deposit':
                return _sofp_delta_formula(sofp_security_deposit_row, period, positive_when_increase=False)
            if metric_key == 'socf_related_party_loan':
                return _sofp_delta_formula(sofp_related_party_loan_row, period, positive_when_increase=True)
            if metric_key == 'share_capital_movement':
                return _sofp_delta_formula(sofp_share_capital_row, period, positive_when_increase=True)
            if metric_key == 'owner_current_account_movement':
                return _sofp_delta_formula(sofp_owner_current_account_row, period, positive_when_increase=True)

            return None

        tb_context = metric_refs.get('__tb_link_context__') or {}

        def socf_formula(metric_key, period):
            sofp_formula = socf_sofp_formula(metric_key, period)
            if sofp_formula:
                return sofp_formula
            direct_formula = self._tb_socf_metric_formula(tb_context, metric_key, is_prior=(period == 'prior'))
            if direct_formula:
                return direct_formula
            return self._metric_ref(metric_refs, metric_key, period)

        def socf_expr(metric_key, period):
            return self._tb_formula_expr(socf_formula(metric_key, period))

        def metric_term(metric_key, period):
            return f'({socf_expr(metric_key, period)})'

        dep_current = socf_expr('socf_depreciation_movement', 'current')
        dep_prior = socf_expr('socf_depreciation_movement', 'prior')
        eosb_adj_current = socf_expr('socf_eosb_adjustment', 'current')
        eosb_adj_prior = socf_expr('socf_eosb_adjustment', 'prior')

        set_formula('net_profit', 'B', f'=SOCI!B{soci_net_profit_row}')
        set_formula('net_profit', 'C', f'=SOCI!C{soci_net_profit_row}')
        set_formula(
            'operating_before_wc',
            'B',
            f"=B{rows['net_profit']}+{dep_current}+{eosb_adj_current}",
        )
        set_formula(
            'operating_before_wc',
            'C',
            f"=C{rows['net_profit']}+{dep_prior}+{eosb_adj_prior}",
        )
        set_formula('change_current_assets', 'B', socf_formula('socf_change_current_assets', 'current'))
        set_formula('change_current_assets', 'C', socf_formula('socf_change_current_assets', 'prior'))
        set_formula(
            'change_current_liabilities',
            'B',
            socf_formula('socf_change_current_liabilities', 'current'),
        )
        set_formula(
            'change_current_liabilities',
            'C',
            socf_formula('socf_change_current_liabilities', 'prior'),
        )
        set_formula('property', 'B', socf_formula('socf_property_investing', 'current'))
        set_formula('property', 'C', socf_formula('socf_property_investing', 'prior'))
        if rows.get('security_deposit'):
            set_formula('security_deposit', 'B', socf_formula('socf_security_deposit', 'current'))
            set_formula('security_deposit', 'C', socf_formula('socf_security_deposit', 'prior'))
        set_formula('paid_up_capital', 'B', socf_formula('share_capital_movement', 'current'))
        set_formula('paid_up_capital', 'C', socf_formula('share_capital_movement', 'prior'))
        if rows.get('dividend_paid'):
            set_formula('dividend_paid', 'B', socf_formula('dividend_paid', 'current'))
            set_formula('dividend_paid', 'C', socf_formula('dividend_paid', 'prior'))
        set_formula(
            'owner_current_account',
            'B',
            socf_formula('owner_current_account_movement', 'current'),
        )
        set_formula(
            'owner_current_account',
            'C',
            socf_formula('owner_current_account_movement', 'prior'),
        )
        if rows.get('related_party_loan'):
            set_formula('related_party_loan', 'B', socf_formula('socf_related_party_loan', 'current'))
            set_formula('related_party_loan', 'C', socf_formula('socf_related_party_loan', 'prior'))
        set_formula(
            'net_operations',
            'B',
            f'=SUM(B{rows["operating_before_wc"]}:B{rows["change_current_liabilities"]})',
        )
        set_formula(
            'net_operations',
            'C',
            f'=SUM(C{rows["operating_before_wc"]}:C{rows["change_current_liabilities"]})',
        )
        net_investing_current_terms = [f'B{rows["property"]}']
        net_investing_prior_terms = [f'C{rows["property"]}']
        if rows.get('security_deposit'):
            net_investing_current_terms.append(f'B{rows["security_deposit"]}')
            net_investing_prior_terms.append(f'C{rows["security_deposit"]}')
        else:
            net_investing_current_terms.append(metric_term('socf_security_deposit', 'current'))
            net_investing_prior_terms.append(metric_term('socf_security_deposit', 'prior'))
        set_formula('net_investing', 'B', '=' + '+'.join(net_investing_current_terms))
        set_formula('net_investing', 'C', '=' + '+'.join(net_investing_prior_terms))

        net_financing_current_terms = [f'B{rows["paid_up_capital"]}']
        net_financing_prior_terms = [f'C{rows["paid_up_capital"]}']
        if rows.get('dividend_paid'):
            net_financing_current_terms.append(f'B{rows["dividend_paid"]}')
            net_financing_prior_terms.append(f'C{rows["dividend_paid"]}')
        else:
            net_financing_current_terms.append(metric_term('dividend_paid', 'current'))
            net_financing_prior_terms.append(metric_term('dividend_paid', 'prior'))
        net_financing_current_terms.append(f'B{rows["owner_current_account"]}')
        net_financing_prior_terms.append(f'C{rows["owner_current_account"]}')
        if rows.get('related_party_loan'):
            net_financing_current_terms.append(f'B{rows["related_party_loan"]}')
            net_financing_prior_terms.append(f'C{rows["related_party_loan"]}')
        else:
            net_financing_current_terms.append(metric_term('socf_related_party_loan', 'current'))
            net_financing_prior_terms.append(metric_term('socf_related_party_loan', 'prior'))
        set_formula(
            'net_financing',
            'B',
            '=' + '+'.join(net_financing_current_terms),
        )
        set_formula(
            'net_financing',
            'C',
            '=' + '+'.join(net_financing_prior_terms),
        )
        set_formula(
            'net_increase',
            'B',
            f'=B{rows["net_financing"]}+B{rows["net_investing"]}+B{rows["net_operations"]}',
        )
        set_formula(
            'net_increase',
            'C',
            f'=C{rows["net_financing"]}+C{rows["net_investing"]}+C{rows["net_operations"]}',
        )
        set_formula('cash_beginning', 'B', f'=C{rows["cash_end"]}')
        set_formula('cash_beginning', 'C', socf_formula('socf_cash_equivalent_opening', 'prior'))
        set_formula('cash_end', 'B', f'=SUM(B{rows["net_increase"]}:B{rows["cash_beginning"]})')
        set_formula('cash_end', 'C', f'=SUM(C{rows["net_increase"]}:C{rows["cash_beginning"]})')

    def _get_report_or_raise(self, xmlid):
        report = self.env.ref(xmlid, raise_if_not_found=False)
        if not report:
            raise ValidationError(_('Accounting report not found: %s') % xmlid)
        return report.with_context(allowed_company_ids=self.company_ids.ids)

    def _build_report_options(self, report, *, date_mode, date_from, date_to, overrides=None, aged=False):
        previous_options = {
            'selected_variant_id': report.id,
            'date': {
                'filter': 'custom',
                'mode': date_mode,
                'date_from': fields.Date.to_string(date_from) if date_from else False,
                'date_to': fields.Date.to_string(date_to),
            },
            'all_entries': self.include_draft_entries,
            'unfold_all': self.unfold_all,
            'hide_0_lines': self.hide_zero_lines,
        }
        if aged:
            previous_options.update({
                'aging_based_on': self.aging_based_on,
                'aging_interval': self.aging_interval,
                'show_currency': self.show_currency,
                'show_account': self.show_account,
            })

        options = report.get_options(previous_options=previous_options)

        if overrides:
            options = self._deep_merge_dict(options, overrides)

        options['report_id'] = report.id
        return options

    def _deep_merge_dict(self, base_dict, patch_dict):
        result = dict(base_dict)
        for key, value in patch_dict.items():
            if isinstance(value, dict) and isinstance(result.get(key), dict):
                result[key] = self._deep_merge_dict(result[key], value)
            else:
                result[key] = value
        return result

    def _build_native_report_sheet_payload(self, *, report, options, sheet_name, empty_message):
        export_data = report.export_to_xlsx(options)
        file_content = export_data.get('file_content')
        if not file_content:
            raise ValidationError(empty_message)

        try:
            source_workbook = load_workbook(io.BytesIO(file_content), data_only=False)
        except Exception as exc:
            report_name = report.display_name or report.name or report._name
            raise ValidationError(
                _('Unable to read native XLSX export for %(report)s: %(error)s', report=report_name, error=str(exc) or repr(exc))
            ) from exc

        if not source_workbook.sheetnames:
            raise ValidationError(_('Native XLSX export did not contain any sheet.'))

        return {
            'sheet_name': sheet_name,
            'native_source_sheet': source_workbook[source_workbook.sheetnames[0]],
        }

    def _prepare_general_ledger_payload(self, overrides):
        report = self._get_report_or_raise('account_reports.general_ledger_report')
        year_end_date = self._get_effective_year_end_date()
        options = self._build_report_options(
            report,
            date_mode='range',
            date_from=self.date_from,
            date_to=year_end_date,
            overrides=overrides,
            aged=False,
        )
        return self._build_native_report_sheet_payload(
            report=report,
            options=options,
            sheet_name='General Ledger',
            empty_message=_('General Ledger XLSX export returned no file content.'),
        )

    def _prepare_aged_payload(self, *, xmlid, sheet_name, title, overrides):
        report = self._get_report_or_raise(xmlid)
        year_end_date = self._get_effective_year_end_date()
        options = self._build_report_options(
            report,
            date_mode='single',
            date_from=year_end_date,
            date_to=year_end_date,
            overrides=overrides,
            aged=True,
        )
        return self._build_native_report_sheet_payload(
            report=report,
            options=options,
            sheet_name=sheet_name,
            empty_message=_('%s XLSX export returned no file content.') % title,
        )

    def _prepare_invoice_bill_payload(self, *, is_customer):
        report_xmlid = 'audit_excel_export.audit_customer_invoices_report' if is_customer else 'audit_excel_export.audit_vendor_bills_report'
        title = 'Customer Invoices' if is_customer else 'Vendor Bills'
        party_label = _('Customer') if is_customer else _('Vendor')
        year_end_date = self._get_effective_year_end_date()

        report = self._get_report_or_raise(report_xmlid)
        overrides = {
            'invoice_bill_report_kind': 'customer' if is_customer else 'vendor',
            'invoice_bill_scope': self.invoice_bill_scope,
            'include_refunds': self.include_refunds,
            'include_dynamic_columns': self.include_dynamic_columns,
        }
        options = self._build_report_options(
            report,
            date_mode='range',
            date_from=self.date_from,
            date_to=year_end_date,
            overrides=overrides,
            aged=False,
        )
        moves = self._get_invoice_bill_moves(report=report, options=options)

        if is_customer:
            return self._prepare_customer_invoice_flat_payload(title=title, moves=moves)

        summary_headers = [
            _('Invoice No'),
            _('Invoice Date'),
            _('Accounting Date'),
            _('Currency'),
            party_label,
            _('Due Date'),
            _('Conversion Rate'),
        ]
        line_headers = [
            _('Label'),
            _('Account'),
            _('Quantity'),
            _('Price'),
            _('Taxes'),
            _('VAT Amount'),
            _('Amount'),
            _('Currency'),
        ]

        blocks = []
        for move in moves:
            move_name = move.name or move.payment_reference or move.ref or '/'
            move_currency = move.currency_id or move.company_currency_id
            line_currency = move_currency.display_name if move_currency else ''
            conversion_rate = self._compute_move_conversion_rate(move)

            lines = []
            for move_line in move.invoice_line_ids.sorted(lambda line: (line.sequence, line.id)):
                line_is_amount_line = move_line.display_type not in {'line_section', 'line_subsection', 'line_note'}
                line_subtotal = move_line.price_subtotal if line_is_amount_line else None
                line_total = move_line.price_total if line_is_amount_line else None
                line_vat_amount = (line_total - line_subtotal) if line_is_amount_line else None
                lines.append([
                    move_line.name or move_line.product_id.display_name or '',
                    move_line.account_id.display_name or '',
                    move_line.quantity if line_is_amount_line else None,
                    move_line.price_unit if line_is_amount_line else None,
                    ', '.join(move_line.tax_ids.mapped('name')),
                    line_vat_amount,
                    line_total,
                    line_currency,
                ])

            blocks.append({
                'title': move_name,
                'summary_values': [
                    move_name,
                    move.invoice_date,
                    move.date,
                    line_currency,
                    move.partner_id.display_name or '',
                    move.invoice_date_due,
                    conversion_rate,
                ],
                'line_rows': lines,
            })

        return {
            'sheet_mode': 'invoice_bill_blocks',
            'sheet_name': title,
            'title': title,
            'summary_headers': summary_headers,
            'line_headers': line_headers,
            'blocks': blocks,
        }

    def _prepare_customer_invoice_flat_payload(self, *, title, moves):
        headers = [
            _('Invoice/Bill Date'),
            _('Number'),
            _('Invoice Partner Display Name'),
            _('Invoice lines/Label'),
            _('Currency'),
            _('Currency/Inverse Rate'),
            _('Invoice lines/Quantity'),
            _('Invoice lines/Unit Price'),
            _('Invoice lines/Subtotal'),
            _('Amount paid'),
            _('Payments/Date'),
            _('Amount Due'),
            _('Amount Due Signed'),
            _('Total'),
            _('Total Signed'),
        ]
        column_types = {
            0: 'date',
            5: 'number',
            6: 'number',
            7: 'number',
            8: 'number',
            9: 'number',
            11: 'number',
            12: 'number',
            13: 'number',
            14: 'number',
        }

        rows = []
        for move in moves:
            move_name = move.name or move.payment_reference or move.ref or '/'
            move_currency = move.currency_id or move.company_currency_id
            currency_label = move_currency.display_name if move_currency else ''
            inverse_rate = self._compute_move_inverse_rate(move)
            payment_dates = ', '.join(self._get_move_payment_dates(move))
            amount_paid = float(move.amount_total or 0.0) - float(move.amount_residual or 0.0)

            invoice_lines = move.invoice_line_ids.sorted(lambda line: (line.sequence, line.id))
            if not invoice_lines:
                invoice_lines = self.env['account.move.line']

            rows.append({
                'row_kind': 'invoice',
                'values': [
                    move.invoice_date,
                    move_name,
                    move.partner_id.display_name or '',
                    _('Invoice'),
                    currency_label,
                    inverse_rate,
                    None,
                    None,
                    None,
                    amount_paid,
                    payment_dates,
                    move.amount_residual,
                    move.amount_residual_signed,
                    move.amount_total,
                    move.amount_total_signed,
                ],
            })

            for line_index, move_line in enumerate(invoice_lines or [False], start=1):
                line_is_amount_line = (
                    move_line.display_type not in {'line_section', 'line_subsection', 'line_note'}
                    if move_line else False
                )
                line_label = (move_line.name or move_line.product_id.display_name or '') if move_line else ''
                rows.append({
                    'row_kind': 'line',
                    'label_indent': 1,
                    'values': [
                        None,
                        None,
                        None,
                        f"{line_index}. {line_label}".strip(),
                        None,
                        None,
                        move_line.quantity if move_line and line_is_amount_line else None,
                        move_line.price_unit if move_line and line_is_amount_line else None,
                        move_line.price_subtotal if move_line and line_is_amount_line else None,
                        None,
                        None,
                        None,
                        None,
                        None,
                        None,
                    ],
                })

        return {
            'sheet_mode': 'invoice_bill_flat',
            'sheet_name': title,
            'title': title,
            'headers': headers,
            'rows': rows,
            'column_types': column_types,
        }

    def _prepare_empty_sheet_payload(self, sheet_name):
        return {
            'sheet_mode': 'empty',
            'sheet_name': sheet_name,
        }

    def _get_invoice_bill_moves(self, *, report, options):
        handler = self.env['audit.invoice.bill.report.handler']
        return handler._query_moves(report, options)

    def _compute_move_conversion_rate(self, move):
        move_currency = move.currency_id
        company_currency = move.company_currency_id
        if not move_currency or not company_currency:
            return None
        if move_currency == company_currency:
            return 1.0

        currency_rate = getattr(move, 'currency_rate', 0.0) or 0.0
        if currency_rate:
            return float(currency_rate)

        total_currency = float(move.amount_total or 0.0)
        total_company = float(move.amount_total_signed or 0.0)
        if total_currency:
            return abs(total_company / total_currency)
        return None

    def _compute_move_inverse_rate(self, move):
        move_currency = move.currency_id
        company_currency = move.company_currency_id
        if not move_currency or not company_currency:
            return None
        if move_currency == company_currency:
            return 1.0

        total_currency = float(move.amount_total or 0.0)
        total_company = abs(float(move.amount_total_signed or 0.0))
        if total_currency:
            return total_company / total_currency

        currency_rate = getattr(move, 'currency_rate', 0.0) or 0.0
        if currency_rate:
            return 1.0 / float(currency_rate)
        return None

    def _get_move_payment_dates(self, move):
        move.invalidate_recordset(['invoice_payments_widget'])
        widget = move.invoice_payments_widget or {}
        if not isinstance(widget, dict):
            return []

        payment_dates = []
        for payment_line in widget.get('content') or []:
            raw_date = payment_line.get('date')
            if not raw_date:
                continue
            try:
                normalized = fields.Date.to_string(fields.Date.to_date(raw_date))
            except Exception:
                normalized = str(raw_date)
            if normalized not in payment_dates:
                payment_dates.append(normalized)
        return payment_dates

    def _build_sheet_metadata(self, sheet_label, options=None):
        year_end_date = self._get_effective_year_end_date()

        metadata = [
            ('Report', sheet_label),
            ('Companies', ', '.join(self.company_ids.mapped('name'))),
            ('Date Range', f"{fields.Date.to_string(self.date_from)} to {fields.Date.to_string(year_end_date)}"),
            ('Year End Date', fields.Date.to_string(year_end_date)),
            ('Include Draft Entries', 'Yes' if self.include_draft_entries else 'No'),
            ('Unfold All', 'Yes' if self.unfold_all else 'No'),
            ('Hide Zero Lines', 'Yes' if self.hide_zero_lines else 'No'),
            ('Generated At', fields.Datetime.to_string(fields.Datetime.now())),
            ('Generated By', self.env.user.name),
        ]

        if options and options.get('warnings'):
            metadata.append(('Warnings', ', '.join(sorted(options['warnings'].keys()))))

        return metadata

    def _get_plain_styles(self):
        thin_side = Side(style='thin', color='D9D9D9')
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        return {
            'title_font': Font(size=13, bold=True),
            'section_fill': PatternFill(fill_type='solid', fgColor='F2F2F2'),
            'header_font': Font(bold=True),
            'meta_key_font': Font(bold=True),
            'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'left': Alignment(horizontal='left', vertical='center', wrap_text=False),
            'number': Alignment(horizontal='right', vertical='center'),
            'border': thin_border,
        }

    def _write_invoice_bill_blocks_sheet(
        self,
        workbook,
        used_sheet_names,
        *,
        insert_index,
        sheet_name,
        title,
        summary_headers,
        line_headers,
        blocks,
    ):
        styles = self._get_plain_styles()
        sheet = workbook.create_sheet(title=self._get_unique_sheet_name(sheet_name, used_sheet_names), index=insert_index)

        total_columns = len(line_headers) or 8
        row = 1
        col_widths = {}
        first_line_header_row = None
        default_col_widths = {
            1: 34,
            2: 28,
            3: 12,
            4: 14,
            5: 24,
            6: 14,
            7: 14,
            8: 14,
        }

        def track_width(col_idx, value):
            text = '' if value is None else str(value)
            col_widths[col_idx] = max(col_widths.get(col_idx, 0), len(text))

        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
        title_cell = sheet.cell(row=row, column=1, value=title)
        title_cell.font = styles['title_font']
        title_cell.alignment = styles['left']
        track_width(1, title)
        row += 2

        if not blocks:
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
            no_data_cell = sheet.cell(row=row, column=1, value=_('No data'))
            no_data_cell.alignment = styles['left']
            no_data_cell.border = styles['border']
            track_width(1, _('No data'))
        else:
            for block in blocks:
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
                block_title_cell = sheet.cell(row=row, column=1, value=block.get('title') or '/')
                block_title_cell.font = styles['header_font']
                block_title_cell.fill = styles['section_fill']
                block_title_cell.alignment = styles['left']
                block_title_cell.border = styles['border']
                track_width(1, block.get('title') or '/')
                row += 1

                summary_count = len(summary_headers)
                for col_idx in range(1, summary_count + 1):
                    header_value = summary_headers[col_idx - 1]
                    header_cell = sheet.cell(row=row, column=col_idx, value=header_value)
                    header_cell.font = styles['header_font']
                    header_cell.fill = styles['section_fill']
                    header_cell.alignment = styles['center']
                    header_cell.border = styles['border']
                    track_width(col_idx, header_value)
                row += 1

                summary_values = block.get('summary_values') or []
                for col_idx in range(1, summary_count + 1):
                    cell_value = summary_values[col_idx - 1] if col_idx - 1 < len(summary_values) else None
                    data_cell = sheet.cell(row=row, column=col_idx, value=cell_value)
                    data_cell.border = styles['border']
                    if col_idx in (2, 3, 6):
                        if isinstance(cell_value, datetime.datetime):
                            data_cell.number_format = 'yyyy-mm-dd'
                            data_cell.alignment = styles['center']
                        elif isinstance(cell_value, datetime.date):
                            data_cell.value = datetime.datetime.combine(cell_value, datetime.time.min)
                            data_cell.number_format = 'yyyy-mm-dd'
                            data_cell.alignment = styles['center']
                        else:
                            data_cell.alignment = styles['left']
                    elif col_idx == 7 and isinstance(cell_value, (int, float)):
                        data_cell.number_format = '#,##0.000000'
                        data_cell.alignment = styles['number']
                    else:
                        data_cell.alignment = styles['left']
                    track_width(col_idx, cell_value)
                row += 2

                header_row = row
                if first_line_header_row is None:
                    first_line_header_row = header_row
                for col_idx, header in enumerate(line_headers, start=1):
                    header_cell = sheet.cell(row=row, column=col_idx, value=header)
                    header_cell.font = styles['header_font']
                    header_cell.fill = styles['section_fill']
                    header_cell.alignment = styles['center']
                    header_cell.border = styles['border']
                    track_width(col_idx, header)
                row += 1

                line_rows = block.get('line_rows') or []
                if not line_rows:
                    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_columns)
                    no_line_cell = sheet.cell(row=row, column=1, value=_('No data'))
                    no_line_cell.alignment = styles['left']
                    no_line_cell.border = styles['border']
                    track_width(1, _('No data'))
                    row += 1
                else:
                    for line_values in line_rows:
                        for col_idx, cell_value in enumerate(line_values, start=1):
                            data_cell = sheet.cell(row=row, column=col_idx, value=cell_value)
                            data_cell.border = styles['border']
                            if col_idx in (3, 4, 6, 7) and isinstance(cell_value, (int, float)):
                                data_cell.number_format = '#,##0.00'
                                data_cell.alignment = styles['number']
                            else:
                                data_cell.alignment = styles['left']
                            track_width(col_idx, cell_value)
                        row += 1

                row += 2

        if first_line_header_row is not None:
            sheet.freeze_panes = f"A{first_line_header_row + 1}"

        for col_idx in range(1, total_columns + 1):
            base_width = default_col_widths.get(col_idx, 12)
            width = max(base_width, col_widths.get(col_idx, 0) + 2)
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(width, 56)

    def _write_invoice_bill_flat_sheet(
        self,
        workbook,
        used_sheet_names,
        *,
        insert_index,
        sheet_name,
        title,
        headers,
        rows,
        column_types=None,
    ):
        sheet = workbook.create_sheet(title=self._get_unique_sheet_name(sheet_name, used_sheet_names), index=insert_index)

        total_columns = len(headers) or 1
        label_col_idx = 4
        row = 1
        col_widths = {}
        column_types = column_types or {}
        default_col_widths = {
            1: 16,
            2: 18,
            3: 26,
            4: 36,
            5: 12,
            6: 14,
            7: 12,
            8: 13,
            9: 14,
            10: 13,
            11: 14,
            12: 13,
            13: 13,
            14: 13,
            15: 13,
        }

        def track_width(col_idx, value):
            text = '' if value is None else str(value)
            col_widths[col_idx] = max(col_widths.get(col_idx, 0), len(text))

        header_row = row
        for col_idx, header in enumerate(headers, start=1):
            header_cell = sheet.cell(row=row, column=col_idx, value=header)
            header_cell.font = Font(name='Calibri', size=11, bold=True)
            header_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            track_width(col_idx, header)
        row += 1

        if not rows:
            no_data_cell = sheet.cell(row=row, column=1, value=_('No data'))
            no_data_cell.font = Font(name='Calibri', size=11)
            no_data_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            track_width(1, _('No data'))
        else:
            for row_data in rows:
                if isinstance(row_data, dict):
                    line_values = row_data.get('values') or []
                    row_kind = row_data.get('row_kind')
                    label_indent = int(row_data.get('label_indent') or 0)
                else:
                    line_values = row_data
                    row_kind = False
                    label_indent = 0
                for col_idx, cell_value in enumerate(line_values, start=1):
                    data_cell = sheet.cell(row=row, column=col_idx)
                    col_type = column_types.get(col_idx - 1)

                    if col_type == 'date':
                        if isinstance(cell_value, datetime.datetime):
                            data_cell.value = cell_value
                            data_cell.number_format = 'yyyy-mm-dd'
                            data_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        elif isinstance(cell_value, datetime.date):
                            data_cell.value = datetime.datetime.combine(cell_value, datetime.time.min)
                            data_cell.number_format = 'yyyy-mm-dd'
                            data_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        else:
                            data_cell.value = cell_value or ''
                            data_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    elif col_type == 'number' and isinstance(cell_value, (int, float)):
                        data_cell.value = float(cell_value)
                        data_cell.number_format = '#,##0.00'
                        data_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    else:
                        data_cell.value = cell_value or ''
                        data_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                    if row_kind == 'invoice':
                        data_cell.font = Font(name='Calibri', size=11, bold=True)
                    elif label_indent and col_idx == label_col_idx:
                        data_cell.font = Font(name='Calibri', size=11)
                        data_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=label_indent)
                    else:
                        data_cell.font = Font(name='Calibri', size=11)

                    track_width(col_idx, cell_value)
                row += 1

        for col_idx in range(1, total_columns + 1):
            base_width = default_col_widths.get(col_idx, 12)
            width = max(base_width, col_widths.get(col_idx, 0) + 2)
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(width, 56)

    def _write_empty_sheet(self, workbook, used_sheet_names, *, insert_index, sheet_name):
        workbook.create_sheet(
            title=self._get_unique_sheet_name(sheet_name, used_sheet_names),
            index=insert_index,
        )

    def _write_plain_sheet(self, workbook, used_sheet_names, *, insert_index, sheet_name, title, metadata, sections):
        styles = self._get_plain_styles()
        sheet = workbook.create_sheet(title=self._get_unique_sheet_name(sheet_name, used_sheet_names), index=insert_index)

        max_columns = max([2] + [len(section.get('headers', [])) for section in sections])
        row = 1
        col_widths = {}

        def track_width(col_idx, value):
            text = '' if value is None else str(value)
            col_widths[col_idx] = max(col_widths.get(col_idx, 0), len(text))

        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_columns)
        title_cell = sheet.cell(row=row, column=1, value=title)
        title_cell.font = styles['title_font']
        title_cell.alignment = styles['left']
        track_width(1, title)
        row += 1

        for key, value in metadata:
            key_cell = sheet.cell(row=row, column=1, value=key)
            key_cell.font = styles['meta_key_font']
            key_cell.alignment = styles['left']
            key_cell.border = styles['border']

            sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=max_columns)
            value_cell = sheet.cell(row=row, column=2, value=value or '')
            value_cell.alignment = styles['left']
            value_cell.border = styles['border']

            track_width(1, key)
            track_width(2, value)
            row += 1

        row += 1
        first_header_row = None

        for section in sections:
            headers = section.get('headers', [])
            rows = section.get('rows', [])
            section_title = section.get('section_title') or _('Section')
            column_types = section.get('column_types', {})
            total_columns = section.get('total_columns', set())

            if not headers:
                continue

            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            section_cell = sheet.cell(row=row, column=1, value=section_title)
            section_cell.font = styles['header_font']
            section_cell.fill = styles['section_fill']
            section_cell.alignment = styles['left']
            section_cell.border = styles['border']
            track_width(1, section_title)
            row += 1

            header_row = row
            if first_header_row is None:
                first_header_row = header_row

            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=row, column=col_idx, value=header)
                cell.font = styles['header_font']
                cell.fill = styles['section_fill']
                cell.alignment = styles['center']
                cell.border = styles['border']
                track_width(col_idx, header)
            row += 1

            data_start_row = row
            if not rows:
                sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
                no_data_cell = sheet.cell(row=row, column=1, value=_('No data'))
                no_data_cell.alignment = styles['left']
                no_data_cell.border = styles['border']
                track_width(1, _('No data'))
                row += 1
                data_end_row = row - 1
            else:
                for line in rows:
                    for col_idx, cell_value in enumerate(line, start=1):
                        cell = sheet.cell(row=row, column=col_idx)
                        col_type = column_types.get(col_idx - 1)

                        if col_type == 'number' and isinstance(cell_value, (int, float)):
                            cell.value = float(cell_value)
                            cell.number_format = '#,##0.00'
                            cell.alignment = styles['number']
                        elif col_type == 'date':
                            if isinstance(cell_value, datetime.datetime):
                                cell.value = cell_value
                                cell.number_format = 'yyyy-mm-dd'
                                cell.alignment = styles['center']
                            elif isinstance(cell_value, datetime.date):
                                cell.value = datetime.datetime.combine(cell_value, datetime.time.min)
                                cell.number_format = 'yyyy-mm-dd'
                                cell.alignment = styles['center']
                            else:
                                cell.value = cell_value or ''
                                cell.alignment = styles['left']
                        elif isinstance(cell_value, (int, float)):
                            cell.value = float(cell_value)
                            cell.alignment = styles['number']
                        else:
                            cell.value = cell_value or ''
                            cell.alignment = styles['left']

                        cell.border = styles['border']
                        track_width(col_idx, cell_value)
                    row += 1
                data_end_row = row - 1

                if total_columns:
                    for col_idx in range(1, len(headers) + 1):
                        cell = sheet.cell(row=row, column=col_idx)
                        cell.border = styles['border']
                        cell.font = styles['header_font']
                        if col_idx == 1:
                            cell.value = _('Total')
                            cell.alignment = styles['left']
                            track_width(col_idx, _('Total'))
                        elif (col_idx - 1) in total_columns:
                            col_letter = get_column_letter(col_idx)
                            cell.value = f'=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})'
                            cell.number_format = '#,##0.00'
                            cell.alignment = styles['number']
                        else:
                            cell.value = ''
                            cell.alignment = styles['left']
                    row += 1

            sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(data_end_row, header_row)}"
            row += 1

        if first_header_row is not None:
            sheet.freeze_panes = f"A{first_header_row + 1}"

        for col_idx, width in col_widths.items():
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(width + 2, 10), 48)

    def _write_native_sheet_copy(self, workbook, used_sheet_names, *, insert_index, sheet_name, source_sheet):
        target_sheet = workbook.create_sheet(
            title=self._get_unique_sheet_name(sheet_name, used_sheet_names),
            index=insert_index,
        )

        for src_row in source_sheet.iter_rows(
            min_row=1,
            max_row=source_sheet.max_row,
            min_col=1,
            max_col=source_sheet.max_column,
        ):
            for src_cell in src_row:
                if src_cell.value is None and not src_cell.has_style and not src_cell.comment and not src_cell.hyperlink:
                    continue

                dst_cell = target_sheet.cell(
                    row=src_cell.row,
                    column=src_cell.column,
                    value=src_cell.value,
                )
                if src_cell.has_style:
                    dst_cell.font = copy(src_cell.font)
                    dst_cell.fill = copy(src_cell.fill)
                    dst_cell.border = copy(src_cell.border)
                    dst_cell.alignment = copy(src_cell.alignment)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.protection = copy(src_cell.protection)
                if src_cell.comment:
                    dst_cell.comment = copy(src_cell.comment)
                if src_cell.hyperlink:
                    dst_cell._hyperlink = copy(src_cell.hyperlink)

        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))

        for col_key, source_dimension in source_sheet.column_dimensions.items():
            target_dimension = target_sheet.column_dimensions[col_key]
            target_dimension.width = source_dimension.width
            target_dimension.hidden = source_dimension.hidden
            target_dimension.bestFit = source_dimension.bestFit
            target_dimension.outlineLevel = source_dimension.outlineLevel
            target_dimension.collapsed = source_dimension.collapsed

        for row_key, source_dimension in source_sheet.row_dimensions.items():
            target_dimension = target_sheet.row_dimensions[row_key]
            target_dimension.height = source_dimension.height
            target_dimension.hidden = source_dimension.hidden
            target_dimension.outlineLevel = source_dimension.outlineLevel
            target_dimension.collapsed = source_dimension.collapsed

        target_sheet.freeze_panes = source_sheet.freeze_panes
        if source_sheet.auto_filter and source_sheet.auto_filter.ref:
            target_sheet.auto_filter.ref = source_sheet.auto_filter.ref
        return target_sheet.title

    def _write_trial_balance_to_template_sheet(self, sheet, trial_balance_payload):
        styles = self._get_plain_styles()
        headers = trial_balance_payload['headers']
        rows = trial_balance_payload['rows']
        start_row = 6

        # Clear previous static values in the write area while preserving formulas.
        max_row = max(sheet.max_row, start_row + 1)
        for row in sheet.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    continue
                cell.value = None

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=start_row, column=col_idx, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['section_fill']
            cell.alignment = styles['center']
            cell.border = styles['border']

        write_row = start_row + 1
        if not rows:
            cell = sheet.cell(row=write_row, column=1, value=_('No data'))
            cell.alignment = styles['left']
            cell.border = styles['border']
            return

        for row_values in rows:
            for col_idx, value in enumerate(row_values, start=1):
                cell = sheet.cell(row=write_row, column=col_idx, value=value)
                cell.border = styles['border']
                if col_idx >= 3 and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = styles['number']
                elif col_idx == 1:
                    cell.alignment = styles['center']
                else:
                    cell.alignment = styles['left']
            write_row += 1

        total_row = write_row
        total_label = sheet.cell(row=total_row, column=1, value=_('Total'))
        total_label.font = styles['header_font']
        total_label.alignment = styles['left']
        total_label.border = styles['border']
        for col_idx in range(2, len(headers) + 1):
            cell = sheet.cell(row=total_row, column=col_idx)
            cell.font = styles['header_font']
            cell.border = styles['border']
            if col_idx >= 3:
                letter = get_column_letter(col_idx)
                cell.value = f'=SUM({letter}{start_row + 1}:{letter}{write_row - 1})'
                cell.number_format = '#,##0.00'
                cell.alignment = styles['number']
            else:
                cell.value = ''
                cell.alignment = styles['left']

        sheet.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{write_row - 1}"
        sheet.freeze_panes = f"A{start_row + 1}"
        sheet.column_dimensions['A'].width = max(sheet.column_dimensions['A'].width or 0, 16)
        sheet.column_dimensions['B'].width = max(sheet.column_dimensions['B'].width or 0, 32)
        sheet.column_dimensions['C'].width = max(sheet.column_dimensions['C'].width or 0, 14)
        sheet.column_dimensions['D'].width = max(sheet.column_dimensions['D'].width or 0, 14)
        sheet.column_dimensions['E'].width = max(sheet.column_dimensions['E'].width or 0, 14)

    def _cleanup_template_sheet_inputs(self, sheet):
        for cell in sheet._cells.values():
            value = cell.value
            if value is None:
                continue
            if isinstance(value, str) and value.startswith('='):
                continue
            if isinstance(value, (int, float, bool, datetime.date, datetime.datetime)):
                cell.value = None
                continue
            if isinstance(value, str) and value.strip().lower() in {'0', '0.0', '0.00', 'n/a', 'na'}:
                cell.value = None

    def _normalize_key(self, value):
        return ''.join(ch for ch in (value or '').strip().lower() if ch.isalnum())

    def _get_template_sheet_styles(self):
        thin_side = Side(style='thin', color='D9D9D9')
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        return {
            'title_font': Font(size=13, bold=True, color='FFFFFF'),
            'subtitle_font': Font(size=11, italic=True, color='595959'),
            'title_fill': PatternFill(fill_type='solid', fgColor='1F4E78'),
            'header_font': Font(bold=True, color='FFFFFF'),
            'header_fill': PatternFill(fill_type='solid', fgColor='2F5597'),
            'label_font': Font(bold=True),
            'border': thin_border,
            'left': Alignment(horizontal='left', vertical='center'),
            'center': Alignment(horizontal='center', vertical='center'),
            'right': Alignment(horizontal='right', vertical='center'),
        }

    def _init_template_sheet(self, sheet, *, title, subtitle_formula=None, with_client_ref=True):
        styles = self._get_template_sheet_styles()
        sheet.page_setup.orientation = 'landscape'
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.column_dimensions['A'].width = 42
        sheet.column_dimensions['B'].width = 18
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 18
        sheet.column_dimensions['E'].width = 18
        sheet.row_dimensions[2].height = 28
        sheet.merge_cells('A2:E2')
        title_cell = sheet['A2']
        title_cell.value = title
        title_cell.font = styles['title_font']
        title_cell.fill = styles['title_fill']
        title_cell.alignment = styles['left']
        title_cell.border = styles['border']
        if with_client_ref:
            sheet['A1'] = "='Client Details'!B4"
        if subtitle_formula:
            sheet.merge_cells('A4:E4')
            subtitle_cell = sheet['A4']
            subtitle_cell.value = subtitle_formula
            subtitle_cell.font = styles['subtitle_font']
            subtitle_cell.alignment = styles['left']

    def _apply_template_header_row(self, sheet, row, headers):
        styles = self._get_template_sheet_styles()
        for col_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=col_index, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['center']
            cell.border = styles['border']

    def _build_template_client_details_sheet(self, sheet):
        self._init_template_sheet(sheet, title='Client Details', with_client_ref=False)
        sheet['A1'] = 'Client Details'
        labels = [
            (4, 'Client Name'),
            (5, 'Period Start Date'),
            (6, 'Period End Date'),
            (8, 'Trade License Number'),
            (9, 'Freezone / Mainland'),
            (10, 'Company Formation Date'),
            (11, 'License Expiry Date'),
            (12, 'Business Activities'),
            (13, 'Owner Names'),
            (15, 'Corporate Tax Registration Number'),
            (16, 'VAT Registration Number'),
        ]
        styles = self._get_template_sheet_styles()
        for row, label in labels:
            cell = sheet.cell(row=row, column=1, value=label)
            cell.font = styles['label_font']
            cell.alignment = styles['left']
            cell.border = styles['border']
            sheet.cell(row=row, column=2, value='')
            sheet.cell(row=row, column=2).border = styles['border']
        sheet['A18'] = 'Client Name Copy'
        sheet['B18'] = '=IF(B4="","",B4)'

    def _build_template_summary_sheet(self, sheet):
        self._init_template_sheet(
            sheet,
            title='Summary Sheet',
            subtitle_formula='=IF(OR(\'Client Details\'!B5="",\'Client Details\'!B6=""),"",TEXT(\'Client Details\'!B5,"DD mmmm YYYY")&" - "&TEXT(\'Client Details\'!B6,"DD mmmm YYYY"))',
        )
        self._apply_template_header_row(sheet, 6, ['Description', 'Amount (AED)'])
        rows = [
            (7, 'Total Revenue', '=SOCI!B8'),
            (8, 'Total Expenses (Excluding Director Salary)', '=SOCI!B27'),
            (9, 'Director Salary', '=SOCI!B16'),
            (10, 'Profit / (Loss)', '=SOCI!B31'),
            (11, 'Total Assets', '=SOFP!B17'),
            (12, 'Total Liabilities', '=SOFP!B30'),
            (13, 'Total Equity', '=SOFP!B24'),
            (15, 'Related Party Revenue', '=SOCI!B9'),
            (16, 'Related Party Loan', '=SOFP!B26'),
            (17, 'Interest Income', '=SOCI!B29'),
            (18, 'Dividend Income', '=0'),
        ]
        styles = self._get_template_sheet_styles()
        for row, label, formula in rows:
            sheet.cell(row=row, column=1, value=label).border = styles['border']
            value_cell = sheet.cell(row=row, column=2, value=formula)
            value_cell.number_format = '#,##0.00'
            value_cell.alignment = styles['right']
            value_cell.border = styles['border']

    def _build_template_sofp_sheet(self, sheet):
        self._init_template_sheet(
            sheet,
            title='SOFP',
            subtitle_formula='=IF(\'Client Details\'!B6="","","As at " & TEXT(\'Client Details\'!B6,"DD mmmm YYYY"))',
        )
        self._apply_template_header_row(sheet, 6, ['', 'Current Year', 'Prior Year'])
        sheet['B6'] = '=IF(\'Client Details\'!B6="","",TEXT(\'Client Details\'!B6,"YYYY"))'
        sheet['C6'] = '=IF(B6="","",B6-1)'
        entries = [
            (7, 'Assets:'),
            (8, 'Non-current assets'),
            (9, 'Property, plant and equipment'),
            (10, 'Total non-current assets'),
            (11, 'Current assets'),
            (12, 'Accounts receivable'),
            (13, 'Prepayment'),
            (14, 'VAT recoverable'),
            (15, 'Cash and bank balances'),
            (16, 'Total current assets'),
            (17, 'Total assets'),
            (19, 'Equity and Liabilities:'),
            (20, 'Equity'),
            (21, 'Share capital'),
            (22, 'Retained earnings'),
            (23, 'Owner current account'),
            (24, 'Total Equity'),
            (25, 'Non-current liabilities'),
            (26, 'Loan from related party'),
            (27, 'Current liabilities'),
            (28, 'Audit and accounting accrual'),
            (29, 'VAT payable'),
            (30, 'Total Liabilities'),
            (31, 'Total Equity and Liabilities'),
            (33, 'Difference (Should be 0)'),
        ]
        formulas = {
            'B10': '=SUM(B8:B9)', 'C10': '=SUM(C8:C9)',
            'B16': '=SUM(B11:B15)', 'C16': '=SUM(C11:C15)',
            'B17': '=B16+B10', 'C17': '=C16+C10',
            'B24': '=SUM(B21:B23)', 'C24': '=SUM(C21:C23)',
            'B30': '=SUM(B25:B29)', 'C30': '=SUM(C25:C29)',
            'B31': '=B30+B24', 'C31': '=C30+C24',
            'B33': '=B31-B17',
        }
        styles = self._get_template_sheet_styles()
        for row, label in entries:
            sheet.cell(row=row, column=1, value=label).border = styles['border']
            sheet.cell(row=row, column=2).border = styles['border']
            sheet.cell(row=row, column=3).border = styles['border']
        for coord, formula in formulas.items():
            cell = sheet[coord]
            cell.value = formula
            cell.number_format = '#,##0.00'
            cell.alignment = styles['right']

    def _build_template_soci_sheet(self, sheet):
        self._init_template_sheet(
            sheet,
            title='SOCI',
            subtitle_formula='=IF(\'Client Details\'!B6="","","For the year ended " & TEXT(\'Client Details\'!B6,"DD mmmm YYYY"))',
        )
        self._apply_template_header_row(sheet, 6, ['', 'Current Year', 'Prior Year'])
        sheet['B6'] = '=IF(\'Client Details\'!B6="","",TEXT(\'Client Details\'!B6,"YYYY"))'
        sheet['C6'] = '=IF(B6="","",B6-1)'
        entries = [
            (8, 'Revenue'),
            (9, 'Revenue - related party'),
            (11, 'Direct cost'),
            (13, 'Gross profit'),
            (15, 'Operating expenses'),
            (16, 'Director salary'),
            (17, 'Salaries, wages and benefits'),
            (18, 'Advertising'),
            (19, 'Audit and accounting'),
            (20, 'Depreciation'),
            (21, 'Government fees'),
            (22, 'Insurance'),
            (23, 'Office expense'),
            (24, 'Bank charges'),
            (25, 'Exchange loss'),
            (26, 'Others'),
            (27, 'Total operating expenses'),
            (29, 'Gain / (loss) on investment'),
            (30, 'Other income'),
            (32, 'Net profit / (loss)'),
        ]
        formulas = {
            'B13': '=SUM(B9:B11)', 'C13': '=SUM(C9:C11)',
            'B27': '=SUM(B16:B26)', 'C27': '=SUM(C16:C26)',
            'B32': '=B13-B27+B29+B30', 'C32': '=C13-C27+C29+C30',
        }
        styles = self._get_template_sheet_styles()
        for row, label in entries:
            sheet.cell(row=row, column=1, value=label).border = styles['border']
            sheet.cell(row=row, column=2).border = styles['border']
            sheet.cell(row=row, column=3).border = styles['border']
        for coord, formula in formulas.items():
            cell = sheet[coord]
            cell.value = formula
            cell.number_format = '#,##0.00'
            cell.alignment = styles['right']

    def _build_template_soce_sheet(self, sheet):
        self._init_template_sheet(
            sheet,
            title='SOCE',
            subtitle_formula='=IF(\'Client Details\'!B6="","","For the year ended " & TEXT(\'Client Details\'!B6,"DD mmmm YYYY"))',
        )
        self._apply_template_header_row(
            sheet,
            6,
            ['', 'Share capital', "Owner's current account", 'Retained earnings', 'Statutory reserves', 'Total'],
        )
        entries = [
            (7, 'Balance as at start of period'),
            (8, 'Paid up capital'),
            (9, "Movement in owner's current account"),
            (10, 'Net profit / (loss)'),
            (11, 'Transfer to reserve'),
            (12, 'Dividend paid'),
            (13, 'Balance as at end of period'),
            (14, 'Paid up capital'),
            (15, "Movement in owner's current account"),
            (16, 'Net profit / (loss)'),
            (17, 'Transfer to reserve'),
            (18, 'Dividend paid'),
            (19, 'Balance c/f'),
        ]
        styles = self._get_template_sheet_styles()
        for row, label in entries:
            sheet.cell(row=row, column=1, value=label).border = styles['border']
            for col in range(2, 7):
                sheet.cell(row=row, column=col).border = styles['border']
        for row in (7, 8, 9, 10, 11, 12, 14, 15, 16, 17, 18):
            sheet.cell(row=row, column=6, value=f'=SUM(B{row}:E{row})')
        for col in ('B', 'C', 'D', 'E', 'F'):
            sheet[f'{col}13'] = f'=SUM({col}7:{col}12)'
            sheet[f'{col}19'] = f'=SUM({col}13:{col}18)'

    def _build_template_socf_sheet(self, sheet):
        self._init_template_sheet(
            sheet,
            title='SOCF',
            subtitle_formula='=IF(\'Client Details\'!B6="","","For the year ended " & TEXT(\'Client Details\'!B6,"DD mmmm YYYY"))',
        )
        self._apply_template_header_row(sheet, 6, ['', 'Current Year', 'Prior Year'])
        entries = [
            (8, 'Cash flows from operating activities'),
            (9, 'Net profit for the year'),
            (10, 'Operating cash flows before working capital changes'),
            (11, 'Changes in working capital'),
            (12, '(Increase) / decrease in current assets'),
            (13, 'Increase / (decrease) in current liabilities'),
            (14, 'Net cash generated from operations'),
            (15, 'Cash flows from investing activities'),
            (16, 'Property, plant and equipment'),
            (17, 'Security deposit'),
            (18, 'Net cash generated from investing activities'),
            (19, 'Cash flows from financing activities'),
            (20, 'Paid up capital'),
            (21, 'Dividend paid'),
            (22, 'Owner current account'),
            (23, 'Loan from related party'),
            (24, 'Net cash generated from financing activities'),
            (25, 'Net increase in cash and cash equivalents'),
            (26, 'Cash and cash equivalents, beginning of the period'),
            (27, 'Cash and cash equivalents, end of the period'),
        ]
        formulas = {
            'B9': '=SOCI!B31', 'C9': '=SOCI!C31',
            'B10': '=B9', 'C10': '=C9',
            'B14': '=SUM(B10:B13)', 'C14': '=SUM(C10:C13)',
            'B18': '=SUM(B16:B17)', 'C18': '=SUM(C16:C17)',
            'B24': '=SUM(B20:B23)', 'C24': '=SUM(C20:C23)',
            'B25': '=B24+B18+B14', 'C25': '=C24+C18+C14',
            'B26': '=C27',
            'B27': '=SUM(B25:B26)', 'C27': '=SUM(C25:C26)',
        }
        styles = self._get_template_sheet_styles()
        for row, label in entries:
            sheet.cell(row=row, column=1, value=label).border = styles['border']
            sheet.cell(row=row, column=2).border = styles['border']
            sheet.cell(row=row, column=3).border = styles['border']
        for coord, formula in formulas.items():
            cell = sheet[coord]
            cell.value = formula
            cell.number_format = '#,##0.00'
            cell.alignment = styles['right']

    def _build_template_prepayment_sheet(self, sheet):
        self._init_template_sheet(
            sheet,
            title='Prepayment',
            subtitle_formula='=IF(\'Client Details\'!B6="","","For the year ended " & TEXT(\'Client Details\'!B6,"DD mmmm YYYY"))',
        )
        rows = [
            (6, 'Renewed in Current Period'),
            (7, 'License'),
            (8, 'Total Amount (AED)'),
            (9, 'Date of Issuance'),
            (10, 'Date of Expiry'),
            (11, 'Total Days'),
            (12, 'Cost Per Day'),
            (13, 'Date of bookkeeping'),
            (14, 'Days Expired'),
            (15, 'Expense'),
            (16, 'Pre-Payment'),
            (18, 'Invoice Details'),
        ]
        styles = self._get_template_sheet_styles()
        for row, label in rows:
            sheet.cell(row=row, column=1, value=label).border = styles['border']
            sheet.cell(row=row, column=2).border = styles['border']

        formulas = {
            'B8': '=SUM(F23:F26)',
            'B11': '=IF(OR(B9="",B10=""),"",B10-B9+1)',
            'B12': '=IF(B11="","",B8/B11)',
            'B13': "='Client Details'!B6",
            'B14': '=IF(OR(B13="",B9=""),"",B13-B9+1)',
            'B15': '=IF(OR(B12="",B14=""),"",B12*B14)',
            'B16': '=IF(OR(B8="",B15=""),"",B8-B15)',
        }
        for coord, formula in formulas.items():
            sheet[coord] = formula

        self._apply_template_header_row(sheet, 19, ['Inv. Date', 'Description', 'Inv No', 'Gross', 'VAT', 'Net'])
        self._apply_template_header_row(sheet, 22, ['Description', 'Quantity', 'Rate', 'Taxable Amount', 'Tax', 'Amount'])

    def _build_template_vat_control_sheet(self, sheet):
        self._init_template_sheet(
            sheet,
            title='VAT Control',
            subtitle_formula='=IF(\'Client Details\'!B6="","","For the year ended " & TEXT(\'Client Details\'!B6,"DD mmmm YYYY"))',
        )
        self._apply_template_header_row(sheet, 6, ['Description', 'Dr.', 'Cr.', 'Balance'])
        rows = [
            (8, 'Balance b/f'),
            (10, 'Output VAT'),
            (12, 'Output VAT on uninvoiced sales'),
            (14, 'Input VAT on service invoice'),
            (16, 'Balance c/d'),
            (20, 'VAT reconciliation'),
            (22, 'VAT liability'),
            (24, 'Less: VAT adjustments'),
            (26, 'B/fwd'),
            (28, 'Current year'),
            (29, 'Output VAT on uninvoiced sales'),
            (33, 'VAT Summary'),
        ]
        styles = self._get_template_sheet_styles()
        for row, label in rows:
            sheet.cell(row=row, column=1, value=label).border = styles['border']
        formulas = {
            'B18': '=SUM(B7:B17)',
            'C18': '=SUM(C7:C17)',
            'D18': '=B18-C18',
            'B22': '=D18',
            'C22': '=SUM(B22)',
            'C27': '=SUM(B27)',
            'C31': '=SUM(C22:C29)',
            'A36': '=C10',
            'B36': '=B14',
            'C36': '=A36-B36',
            'E36': '=C36',
            'A37': '=SUM(A36:A36)',
            'B37': '=SUM(B36:B36)',
            'C37': '=SUM(C36:C36)',
            'D37': '=SUM(D36:D36)',
            'E37': '=SUM(E36:E36)',
        }
        for coord, formula in formulas.items():
            sheet[coord] = formula

    def _build_template_accruals_sheet(self, sheet):
        self._init_template_sheet(
            sheet,
            title='Accruals',
            subtitle_formula='=IF(\'Client Details\'!B6="","","For the year ended " & TEXT(\'Client Details\'!B6,"DD mmmm YYYY"))',
        )
        self._apply_template_header_row(sheet, 6, ['Description', 'Dr.', 'Cr.'])
        rows = [
            (7, 'B/F'),
            (8, '=IF(\'Client Details\'!B6="","","Audit fees " & TEXT(\'Client Details\'!B6,"DD mmmm YYYY"))'),
            (9, 'C/F'),
        ]
        styles = self._get_template_sheet_styles()
        for row, label in rows:
            sheet.cell(row=row, column=1, value=label).border = styles['border']
            sheet.cell(row=row, column=2).border = styles['border']
            sheet.cell(row=row, column=3).border = styles['border']
        sheet['B10'] = '=SUM(B7:B9)'
        sheet['C10'] = '=SUM(C7:C9)'

    def _build_template_share_capital_sheet(self, sheet):
        self._init_template_sheet(
            sheet,
            title='Share Capital',
            subtitle_formula='=IF(\'Client Details\'!B6="","","As at " & TEXT(\'Client Details\'!B6,"DD mmmm YYYY"))',
        )
        self._apply_template_header_row(sheet, 6, ['Owner Name', 'No. of Shares', 'Value per Share (AED)', 'Total Share Value (AED)'])
        styles = self._get_template_sheet_styles()
        for row in range(7, 11):
            for col in range(1, 5):
                sheet.cell(row=row, column=col).border = styles['border']
            sheet[f'D{row}'] = f'=B{row}*C{row}'
        sheet['B11'] = '=SUM(B7:B10)'
        sheet['C11'] = '=SUM(C7:C10)'
        sheet['D11'] = '=SUM(D7:D10)'

    def _build_template_trial_balance_sheet(self, sheet):
        self._init_template_sheet(
            sheet,
            title='Trial Balance',
            subtitle_formula='=IF(\'Client Details\'!B6="","","For the year ended " & TEXT(\'Client Details\'!B6,"DD mmmm YYYY"))',
        )
        styles = self._get_template_sheet_styles()
        note = sheet['A5']
        note.value = 'Figures are populated by export logic.'
        note.font = styles['subtitle_font']


    # ──────────────────────────────────────────────────────────────────────
    # Stage G2: Populate SOFP / SOCI / SOCE / SOCF from Audit Report data
    # ──────────────────────────────────────────────────────────────────────

    _AUDIT_REPORT_STATEMENT_SHEETS = ('SOFP', 'SOCI', 'SOCE', 'SOCF')

    def _stage_g2_populate_statements_from_audit_report(self, stage_c_context, selected_sheet_names):
        """Overwrite SOFP/SOCI/SOCE/SOCF cells with values from Audit Report.

        This stage runs after stage G (TB formula linking) and replaces the
        TB-anchored formulas with actual computed values from
        ``audit.report._get_report_data()``, ensuring the Excel output matches
        the audit report PDF.
        """
        workbook = stage_c_context.get('workbook')
        if not workbook:
            return

        selected_set = set(selected_sheet_names or [])
        if not selected_set.intersection(self._AUDIT_REPORT_STATEMENT_SHEETS):
            return  # none of the 4 sheets requested – nothing to do

        try:
            stmt_data = self._get_audit_report_statement_data()
        except Exception as exc:
            raise ValidationError(
                _(
                    'Unable to populate statement sheets from Audit Report data: %(error)s',
                    error=str(exc) or repr(exc) or exc.__class__.__name__,
                )
            ) from exc

        rd = stmt_data['rd']
        pt = stmt_data['period_totals']       # current-period prefix totals
        ppt = stmt_data['prior_period_totals']  # prior-period prefix totals

        if 'SOFP' in selected_set:
            sofp = self._get_sheet_if_exists(workbook, 'SOFP')
            if sofp is not None:
                self._ar_populate_sofp(sofp, rd)

        if 'SOCI' in selected_set:
            soci = self._get_sheet_if_exists(workbook, 'SOCI')
            if soci is not None:
                self._ar_populate_soci(soci, rd, pt, ppt)

        if 'SOCE' in selected_set:
            soce = self._get_sheet_if_exists(workbook, 'SOCE')
            if soce is not None:
                self._ar_populate_soce(soce, rd)

        if 'SOCF' in selected_set:
            socf = self._get_sheet_if_exists(workbook, 'SOCF')
            if socf is not None:
                self._ar_populate_socf(socf, rd)

        # Re-link summary sheet after overwriting statement values.
        self._refresh_summary_sheet_links(workbook)

    # -- helpers for building an audit.report wizard and extracting data ----

    def _get_audit_report_statement_data(self):
        """Create a shadow ``audit.report`` wizard and return statement data."""
        company = self._get_primary_company_for_template()
        periods = self._get_reporting_periods()

        ar_vals = {
            'company_id': company.id,
            'date_start': periods['date_start'],
            'date_end': periods['date_end'],
            'balance_sheet_date_mode': self.balance_sheet_date_mode,
            'prior_year_mode': self.prior_year_mode,
            'prior_balance_sheet_date_mode': self.prior_balance_sheet_date_mode or 'end_only',
            'prior_date_start': periods.get('prior_date_start') if periods.get('show_prior_year') else False,
            'prior_date_end': periods.get('prior_date_end') if periods.get('show_prior_year') else False,
            'audit_period_category': 'normal_2y' if periods.get('show_prior_year') else 'normal_1y',
            'soce_prior_opening_label_date': (
                periods.get('prior_date_start')
                or periods.get('prior_opening_date_end')
                or periods.get('date_start')
            ) if periods.get('show_prior_year') else False,
            'use_previous_settings': False,
        }
        ar_wizard = self.env['audit.report'].create(ar_vals)
        rd = ar_wizard._get_report_data()

        # Period-level prefix totals for SOCI expense breakdown now come from
        # the local native Odoo Trial Balance loader, matching audit.report's
        # amount basis without depending on its helper methods.
        period_rows = self._tb_fetch_movement_rows(
            company,
            periods['date_start'], periods['date_end'],
        )
        prior_period_rows = []
        if periods.get('show_prior_year'):
            prior_period_rows = self._tb_fetch_movement_rows(
                company,
                periods.get('prior_date_start'), periods.get('prior_date_end'),
            )

        return {
            'rd': rd,
            'period_totals': self._build_flat_prefix_totals(period_rows),
            'prior_period_totals': self._build_flat_prefix_totals(prior_period_rows),
        }

    @staticmethod
    def _build_flat_prefix_totals(rows):
        """Build a single dict mapping every prefix (2/4/6/8-digit) to its balance sum."""
        totals = {}
        for row in (rows or []):
            code = row.get('code') or ''
            balance = row.get('balance', 0.0)
            for length in (2, 4, 6, 8):
                if len(code) >= length:
                    key = code[:length]
                    totals[key] = totals.get(key, 0.0) + balance
        return totals

    @staticmethod
    def _subhead_balance(subheads_by_group, group_code, subhead_code):
        for entry in (subheads_by_group or {}).get(group_code, []):
            if entry.get('code') == subhead_code:
                return entry.get('balance', 0.0)
        return 0.0

    @staticmethod
    def _sum_subheads_excluding(subheads_by_group, group_code, exclude_codes):
        """Sum all subhead balances in *group_code* except those in *exclude_codes*."""
        exclude = set(exclude_codes or [])
        return sum(
            entry.get('balance', 0.0)
            for entry in (subheads_by_group or {}).get(group_code, [])
            if entry.get('code') not in exclude
        )

    def _ar_set_value(self, sheet, coord, value, number_format='#,##0.00'):
        """Write a numeric value to *coord* with consistent formatting."""
        cell = sheet[coord]
        cell.value = value
        cell.number_format = number_format

    def _ar_ensure_detail_rows(self, sheet, header_label, end_labels, required_count):
        """Return detail rows between *header_label* and *end_labels*.

        Rows are expanded (never shrunk) to fit *required_count* while preserving
        local styling copied from the section's first detail row.
        """
        header_row = self._find_row_by_label(sheet, header_label)
        if not header_row:
            return []

        if isinstance(end_labels, str):
            end_labels = [end_labels]
        end_row = self._find_row_by_keys(sheet, end_labels, start_row=header_row + 1)
        if not end_row or end_row <= header_row:
            return []

        first_detail_row = header_row + 1
        existing_count = max(end_row - first_detail_row, 0)
        if required_count > existing_count:
            insert_count = required_count - existing_count
            style_source_row = first_detail_row if existing_count > 0 else max(header_row, end_row - 1)
            sheet.insert_rows(end_row, insert_count)
            for offset in range(insert_count):
                self._copy_row_style(sheet, style_source_row, end_row + offset, max_col=5)

        end_row = self._find_row_by_keys(sheet, end_labels, start_row=header_row + 1)
        if not end_row or end_row <= header_row:
            return []
        return list(range(header_row + 1, end_row))

    def _ar_write_dynamic_rows(self, sheet, row_indexes, lines, show_prior, number_format='#,##0.00'):
        """Write dynamic lines into provided statement section rows."""
        for idx, row_idx in enumerate(row_indexes):
            if idx < len(lines):
                line = lines[idx]
                sheet.row_dimensions[row_idx].hidden = False
                sheet.cell(row=row_idx, column=1, value=line.get('label'))
                current_value = line.get('current', 0.0)
                if current_value is None:
                    sheet[f'B{row_idx}'] = None
                else:
                    self._ar_set_value(sheet, f'B{row_idx}', current_value, number_format)
                if show_prior:
                    prior_value = line.get('prev', 0.0)
                    if prior_value is None:
                        sheet[f'C{row_idx}'] = None
                    else:
                        self._ar_set_value(sheet, f'C{row_idx}', prior_value, number_format)
                else:
                    sheet[f'C{row_idx}'] = None
                continue

            sheet.row_dimensions[row_idx].hidden = True
            sheet[f'B{row_idx}'] = None
            sheet[f'C{row_idx}'] = None

    def _write_dynamic_formula_rows(self, sheet, row_indexes, lines, show_prior, number_format='#,##0.00'):
        """Write dynamic lines whose amounts are Excel formulas instead of static values."""
        for idx, row_idx in enumerate(row_indexes):
            if idx < len(lines):
                line = lines[idx]
                sheet.row_dimensions[row_idx].hidden = False
                sheet.cell(row=row_idx, column=1, value=line.get('label'))

                current_formula = line.get('current_formula')
                if current_formula is None:
                    sheet[f'B{row_idx}'] = None
                else:
                    sheet[f'B{row_idx}'] = current_formula
                    sheet[f'B{row_idx}'].number_format = number_format

                if show_prior:
                    prior_formula = line.get('prior_formula')
                    if prior_formula is None:
                        sheet[f'C{row_idx}'] = None
                    else:
                        sheet[f'C{row_idx}'] = prior_formula
                        sheet[f'C{row_idx}'].number_format = number_format
                else:
                    sheet[f'C{row_idx}'] = None
                continue

            sheet.row_dimensions[row_idx].hidden = True
            sheet[f'B{row_idx}'] = None
            sheet[f'C{row_idx}'] = None

    def _ar_ensure_rows_between(self, sheet, start_row, end_row, required_count, *, max_col=5):
        """Ensure there are at least *required_count* rows between two anchors."""
        if not sheet or not start_row or not end_row or end_row <= start_row:
            return []

        existing_count = max(end_row - start_row - 1, 0)
        insert_count = max(required_count - existing_count, 0)
        if insert_count:
            style_source_row = start_row + 1 if existing_count > 0 else start_row
            sheet.insert_rows(end_row, insert_count)
            for offset in range(insert_count):
                self._copy_row_style(sheet, style_source_row, end_row + offset, max_col=max_col)

        final_end_row = end_row + insert_count
        return list(range(start_row + 1, final_end_row))

    def _ar_get_note_section_lines(self, rd, section_label, show_prior):
        """Extract note-section lines into a generic label/current/prev list."""
        expected = self._normalize_key(section_label or '')
        for section in (rd.get('note_sections') or []):
            if self._normalize_key(section.get('label') or '') != expected:
                continue
            lines = []
            for line in (section.get('lines') or []):
                label = (line.get('name') or line.get('code') or '').strip()
                if not label:
                    continue
                current = abs(line.get('current', 0.0) or 0.0)
                prev = abs(line.get('prev', 0.0) or 0.0)
                if not (current or (show_prior and prev)):
                    continue
                lines.append({
                    'label': label,
                    'current': current,
                    'prev': prev,
                })
            return lines
        return []

    def _ar_fetch_closing_account_rows(self, company, date_end):
        """Return account-level closing balances as of *date_end*."""
        if not company or not date_end:
            return []
        try:
            rows = self._tb_fetch_rows_from_odoo_trial_balance(company, False, date_end)
            if rows is None:
                rows = self._tb_fetch_grouped_move_line_rows(company, False, date_end)
        except Exception as err:
            _logger.exception(
                "Falling back to grouped move-line closing rows for SOFP in %s id=%s due to: %s",
                getattr(self, '_name', self.__class__.__name__),
                getattr(self, 'id', False),
                err,
            )
            rows = self._tb_fetch_grouped_move_line_rows(company, False, date_end)
        return self._tb_project_account_rows(rows, balance_role='closing')

    @staticmethod
    def _ar_sofp_account_section(code):
        normalized_code = (code or '').strip()
        if not normalized_code:
            return False
        if normalized_code == '12040101':
            return 'equity'
        if normalized_code.startswith('11'):
            return 'non_current_assets'
        if normalized_code.startswith('1204') or normalized_code.startswith('1206'):
            return 'cash_bank'
        if normalized_code.startswith('1'):
            return 'current_assets'
        if normalized_code.startswith('21'):
            return 'non_current_liabilities'
        if normalized_code.startswith('22'):
            return 'current_liabilities'
        if normalized_code.startswith('3'):
            return 'equity'
        return False

    @staticmethod
    def _ar_format_sofp_account_label(row):
        code = (row.get('code_raw') or row.get('code') or '').strip()
        name = (row.get('name') or '').strip()
        return name or code or 'Account'

    def _ar_build_sofp_account_lines(self, show_prior):
        company = self._get_primary_company_for_template()
        periods = self._get_reporting_periods()
        current_rows = self._ar_fetch_closing_account_rows(company, periods.get('date_end'))
        prior_rows = []
        if show_prior:
            prior_rows = self._ar_fetch_closing_account_rows(company, periods.get('prior_date_end'))

        current_map = {
            (row.get('code') or ''): row
            for row in (current_rows or [])
            if row.get('code')
        }
        prior_map = {
            (row.get('code') or ''): row
            for row in (prior_rows or [])
            if row.get('code')
        }
        all_codes = sorted(set(current_map) | set(prior_map))

        lines_by_section = {
            'non_current_assets': [],
            'current_assets': [],
            'equity': [],
            'non_current_liabilities': [],
            'current_liabilities': [],
        }
        cash_bank_current = 0.0
        cash_bank_prior = 0.0

        for code in all_codes:
            current_row = current_map.get(code) or {}
            prior_row = prior_map.get(code) or {}
            current_value = self._tb_to_float(current_row.get('end_balance'))
            prior_value = self._tb_to_float(prior_row.get('end_balance'))

            section = self._ar_sofp_account_section(code)
            if not section:
                continue
            if section == 'cash_bank':
                cash_bank_current += current_value
                cash_bank_prior += prior_value
                continue

            if section in ('equity', 'non_current_liabilities', 'current_liabilities'):
                current_value = -current_value
                prior_value = -prior_value

            if not (current_value or (show_prior and prior_value)):
                continue

            lines_by_section[section].append({
                'label': self._ar_format_sofp_account_label(current_row or prior_row),
                'current': current_value,
                'prev': prior_value,
            })

        if cash_bank_current or (show_prior and cash_bank_prior):
            lines_by_section['current_assets'].append({
                'label': 'Cash and bank balances',
                'current': cash_bank_current,
                'prev': cash_bank_prior,
            })

        return lines_by_section

    # -- SOFP population ---------------------------------------------------

    def _ar_populate_sofp(self, sheet, rd):
        show_prior = bool(rd.get('show_prior_year'))
        fmt = '#,##0.00'

        def include_line(current, prev=0.0):
            return bool(current) or (show_prior and bool(prev))
        sofp_lines = self._ar_build_sofp_account_lines(show_prior)
        non_current_asset_lines = sofp_lines.get('non_current_assets', [])
        current_asset_lines = sofp_lines.get('current_assets', [])
        equity_lines = sofp_lines.get('equity', [])
        non_current_liability_lines = sofp_lines.get('non_current_liabilities', [])
        current_liability_lines = sofp_lines.get('current_liabilities', [])

        # Fallback to grouped totals if account-level extraction yielded no lines.
        if not (
            non_current_asset_lines
            or current_asset_lines
            or equity_lines
            or non_current_liability_lines
            or current_liability_lines
        ):
            cgt = rd.get('current_group_totals', {})
            pgt = rd.get('prev_group_totals', {})
            main_head_labels = rd.get('main_head_labels', {})
            non_current_codes = ('1101', '1102', '1103', '1104', '1105', '1106', '1107', '1108', '1109')
            current_codes = ('1201', '1202', '1203', '1204', '1206')
            non_current_liability_codes = ('2101', '2102', '2103', '2104')
            current_liability_codes = ('2201', '2202', '2203', '2204')

            for code in non_current_codes:
                current = cgt.get(code, 0.0)
                prev = pgt.get(code, 0.0)
                if include_line(current, prev):
                    non_current_asset_lines.append({
                        'label': main_head_labels.get(code, code),
                        'current': current,
                        'prev': prev,
                    })

            for code in current_codes:
                current = cgt.get(code, 0.0)
                prev = pgt.get(code, 0.0)
                if code == '1203':
                    current += cgt.get('1205', 0.0)
                    prev += pgt.get('1205', 0.0)
                if include_line(current, prev):
                    current_asset_lines.append({
                        'label': main_head_labels.get(code, code),
                        'current': current,
                        'prev': prev,
                    })

            share_capital_current = rd.get('share_capital_total', 0.0) or 0.0
            share_capital_prev = rd.get('prev_share_capital_total', 0.0) or 0.0
            if rd.get('show_shareholder_note'):
                equity_lines.append({
                    'label': 'Share capital',
                    'current': share_capital_current,
                    'prev': share_capital_prev,
                })

            retained_current = rd.get('retained_earnings_balance', 0.0) or 0.0
            retained_prev = rd.get('prev_retained_earnings_balance', 0.0) or 0.0
            if include_line(retained_current, retained_prev):
                equity_lines.append({
                    'label': 'Retained earnings',
                    'current': retained_current,
                    'prev': retained_prev,
                })

            owner_current = rd.get('owner_current_account_equity', 0.0) or 0.0
            owner_prev = rd.get('prev_owner_current_account_equity', 0.0) or 0.0
            if rd.get('show_owner_current_account_equity_row'):
                equity_lines.append({
                    'label': 'Owner current account',
                    'current': owner_current,
                    'prev': owner_prev,
                })

            statutory_current = rd.get('statutory_reserves_equity', 0.0) or 0.0
            statutory_prev = rd.get('prev_statutory_reserves_equity', 0.0) or 0.0
            if include_line(statutory_current, statutory_prev):
                equity_lines.append({
                    'label': 'Statutory reserves',
                    'current': statutory_current,
                    'prev': statutory_prev,
                })

            for code in non_current_liability_codes:
                current = -(cgt.get(code, 0.0))
                prev = -(pgt.get(code, 0.0))
                if include_line(current, prev):
                    non_current_liability_lines.append({
                        'label': main_head_labels.get(code, code),
                        'current': current,
                        'prev': prev,
                    })

            other_payables_current = -sum(cgt.get(code, 0.0) for code in current_liability_codes)
            other_payables_prev = -sum(pgt.get(code, 0.0) for code in current_liability_codes)
            if include_line(other_payables_current, other_payables_prev):
                current_liability_lines.append({
                    'label': 'Other payables',
                    'current': other_payables_current,
                    'prev': other_payables_prev,
                })

        non_current_asset_rows = self._ar_ensure_detail_rows(
            sheet, 'Non-current assets', ['Total non-current assets'], len(non_current_asset_lines)
        )
        self._ar_write_dynamic_rows(sheet, non_current_asset_rows, non_current_asset_lines, show_prior, fmt)

        current_asset_rows = self._ar_ensure_detail_rows(
            sheet, 'Current assets', ['Total current assets'], len(current_asset_lines)
        )
        self._ar_write_dynamic_rows(sheet, current_asset_rows, current_asset_lines, show_prior, fmt)

        equity_rows = self._ar_ensure_detail_rows(sheet, 'Equity', ['Total Equity', 'Total equity'], len(equity_lines))
        self._ar_write_dynamic_rows(sheet, equity_rows, equity_lines, show_prior, fmt)

        non_current_liability_rows = self._ar_ensure_detail_rows(
            sheet, 'Non-current liabilities', ['Current liabilities'], len(non_current_liability_lines)
        )
        self._ar_write_dynamic_rows(
            sheet, non_current_liability_rows, non_current_liability_lines, show_prior, fmt
        )

        current_liability_rows = self._ar_ensure_detail_rows(
            sheet, 'Current liabilities', ['Total Liabilities'], len(current_liability_lines)
        )
        self._ar_write_dynamic_rows(sheet, current_liability_rows, current_liability_lines, show_prior, fmt)

        # Totals – keep/restore template SUM formulas so the sheet is self-consistent.
        r_tnca = self._find_row_by_label(sheet, 'Total non-current assets')
        r_tca = self._find_row_by_label(sheet, 'Total current assets')
        r_ta = self._find_row_by_label(sheet, 'Total assets')
        r_te = self._find_row_by_keys(sheet, ['Total Equity', 'Total equity'])
        r_tl = self._find_row_by_label(sheet, 'Total Liabilities')
        r_tel = self._find_row_by_label(sheet, 'Total Equity and Liabilities')
        r_nca_hdr = self._find_row_by_label(sheet, 'Non-current assets')
        r_ca_hdr = self._find_row_by_label(sheet, 'Current assets')
        r_eq_hdr = self._find_row_by_label(sheet, 'Equity')
        r_ncl_hdr = self._find_row_by_label(sheet, 'Non-current liabilities')

        if r_tnca and r_nca_hdr:
            sheet[f'B{r_tnca}'] = f'=SUM(B{r_nca_hdr}:B{r_tnca - 1})'
            sheet[f'C{r_tnca}'] = f'=SUM(C{r_nca_hdr}:C{r_tnca - 1})'
        if r_tca and r_ca_hdr:
            sheet[f'B{r_tca}'] = f'=SUM(B{r_ca_hdr}:B{r_tca - 1})'
            sheet[f'C{r_tca}'] = f'=SUM(C{r_ca_hdr}:C{r_tca - 1})'
        if r_ta and r_tnca and r_tca:
            sheet[f'B{r_ta}'] = f'=B{r_tca}+B{r_tnca}'
            sheet[f'C{r_ta}'] = f'=C{r_tca}+C{r_tnca}'
        if r_te and r_eq_hdr:
            sheet[f'B{r_te}'] = f'=SUM(B{r_eq_hdr}:B{r_te - 1})'
            sheet[f'C{r_te}'] = f'=SUM(C{r_eq_hdr}:C{r_te - 1})'
        if r_tl and r_ncl_hdr:
            sheet[f'B{r_tl}'] = f'=SUM(B{r_ncl_hdr}:B{r_tl - 1})'
            sheet[f'C{r_tl}'] = f'=SUM(C{r_ncl_hdr}:C{r_tl - 1})'
        if r_tel and r_tl and r_te:
            sheet[f'B{r_tel}'] = f'=B{r_tl}+B{r_te}'
            sheet[f'C{r_tel}'] = f'=C{r_tl}+C{r_te}'

        # Difference row
        r_diff = self._find_row_by_keys(sheet, ['Difference (Should be 0)', 'Difference'])
        if r_diff and r_tel and r_ta:
            sheet[f'B{r_diff}'] = f'=B{r_tel}-B{r_ta}'

        if not show_prior:
            sheet['C6'] = None
            for row_idx in range(7, sheet.max_row + 1):
                sheet[f'C{row_idx}'] = None

    # -- SOCI population ---------------------------------------------------

    def _ar_populate_soci(self, sheet, rd, pt, ppt):
        """Populate Statement of Comprehensive Income from Audit Report data.

        *pt* / *ppt*: flat prefix-total dicts for current / prior period.
        """
        show_prior = bool(rd.get('show_prior_year'))
        fmt = '#,##0.00'

        r_rev = self._find_row_by_label(sheet, 'Revenue') or 8
        r_rev_rel = self._find_row_by_label(sheet, 'Revenue - related party') or 9
        r_dc = self._find_row_by_label(sheet, 'Direct cost') or 11
        r_gp = self._find_row_by_label(sheet, 'Gross profit') or 13
        r_op_hdr = self._find_row_by_label(sheet, 'Operating expenses') or 15

        # Revenue (4101 negated for positive display)
        self._ar_set_value(sheet, f'B{r_rev}', -(pt.get('4101', 0.0)), fmt)
        self._ar_set_value(sheet, f'C{r_rev}', -(ppt.get('4101', 0.0)), fmt)

        # Revenue – related party (4102)
        self._ar_set_value(sheet, f'B{r_rev_rel}', -(pt.get('4102', 0.0)), fmt)
        self._ar_set_value(sheet, f'C{r_rev_rel}', -(ppt.get('4102', 0.0)), fmt)

        # Direct cost (5101 group)
        self._ar_set_value(sheet, f'B{r_dc}', pt.get('5101', 0.0), fmt)
        self._ar_set_value(sheet, f'C{r_dc}', ppt.get('5101', 0.0), fmt)

        # Gross profit formula
        sheet[f'B{r_gp}'] = f'=B{r_rev}+B{r_rev_rel}-B{r_dc}'
        sheet[f'C{r_gp}'] = f'=C{r_rev}+C{r_rev_rel}-C{r_dc}'

        # Use Operating expenses note lines directly so SOCI shows full dynamic
        # expense account detail (same source as notes section).
        operating_expense_lines = self._ar_get_note_section_lines(rd, 'Operating expenses', show_prior)

        # Backward-compatible fallback if note lines are absent.
        if not operating_expense_lines:
            _SOCI_EXPENSE_MAP = {
                'Director salary': ['510701'],
                'Salaries, wages and benefits': ['5108'],
                'Office Staff Salaries': ['510801'],
                'Coaching Staff Salaries': ['510802'],
                'Employee Benefits & Allowances': ['510803'],
                'Bonus & Incentives': ['510804'],
                'Staff Welfare': ['510805'],
                'Advertising': ['510201'],
                'Audit and accounting': ['5109'],
                'Audit Fee': ['510901'],
                'Accounting & Bookkeeping Fee': ['510902'],
                'Depreciation and amortization': ['5114'],
                'Depreciation': ['5114'],
                'Depreciation Expense': ['511401'],
                'Amortization Expense': ['511402'],
                'Government fees': ['510601', '512601', '512602', '512603'],
                'Legal & Government Fee': ['510601'],
                'Trade License': ['512601'],
                'Establishment Card': ['512602'],
                'Visa Fee': ['512603'],
                'Insurance': ['512501'],
                'Office expense': [
                    '510401', '511001', '511002', '511101', '511201', '511202',
                    '511501', '511701', '511801', '512801',
                ],
                'Bank charges': ['512301'],
                'Exchange loss': ['512201'],
                'Other expenses': [
                    '510301', '510501', '511301', '511302', '511601', '511901',
                    '512001', '512101', '512202', '512401', '512701',
                ],
            }

            def _sum_prefixes(totals, prefixes):
                return sum(totals.get(p, 0.0) for p in prefixes)

            r_top_fallback = self._find_row_by_label(sheet, 'Total operating expenses') or 27
            for row_idx in range(r_op_hdr + 1, r_top_fallback):
                label = (sheet.cell(row=row_idx, column=1).value or '').strip()
                if not label:
                    continue
                prefixes = _SOCI_EXPENSE_MAP.get(label)
                if prefixes is None:
                    continue
                current = _sum_prefixes(pt, prefixes)
                prev = _sum_prefixes(ppt, prefixes)
                if not (current or (show_prior and prev)):
                    continue
                operating_expense_lines.append({
                    'label': label,
                    'current': current,
                    'prev': prev,
                })

        op_rows = self._ar_ensure_detail_rows(
            sheet, 'Operating expenses', ['Total operating expenses'], len(operating_expense_lines)
        )
        self._ar_write_dynamic_rows(sheet, op_rows, operating_expense_lines, show_prior, fmt)

        r_top = self._find_row_by_label(sheet, 'Total operating expenses') or 27
        r_inv = (
            self._find_row_by_label(sheet, 'Gain / (loss) on investment', start_row=r_top + 1)
            or self._find_row_by_label(sheet, 'Gain / (loss) on investment')
            or 29
        )
        r_oi = (
            self._find_row_by_label(sheet, 'Other income', start_row=r_inv + 1)
            or self._find_row_by_label(sheet, 'Other income')
            or 30
        )
        r_np = (
            self._find_row_by_keys(sheet, ['Net profit / (loss)', 'Net profit'], start_row=r_oi + 1)
            or self._find_row_by_keys(sheet, ['Net profit / (loss)', 'Net profit'])
            or 32
        )

        # Total operating expenses formula
        sheet[f'B{r_top}'] = f'=SUM(B{r_op_hdr + 1}:B{r_top - 1})'
        sheet[f'C{r_top}'] = f'=SUM(C{r_op_hdr + 1}:C{r_top - 1})'

        # Investment gain / (loss) (5201 negated for P&L display)
        self._ar_set_value(sheet, f'B{r_inv}', -(pt.get('5201', 0.0)), fmt)
        self._ar_set_value(sheet, f'C{r_inv}', -(ppt.get('5201', 0.0)), fmt)

        # Other income (4103) is always shown as a positive display amount.
        self._ar_set_value(sheet, f'B{r_oi}', abs(pt.get('4103', 0.0)), fmt)
        self._ar_set_value(sheet, f'C{r_oi}', abs(ppt.get('4103', 0.0)), fmt)

        # Net profit formula
        sheet[f'B{r_np}'] = f'=B{r_gp}-B{r_top}+B{r_inv}+B{r_oi}'
        sheet[f'C{r_np}'] = f'=C{r_gp}-C{r_top}+C{r_inv}+C{r_oi}'

        if not show_prior:
            sheet['C6'] = None
            for row_idx in range(7, sheet.max_row + 1):
                sheet[f'C{row_idx}'] = None

    # -- SOCE population ---------------------------------------------------

    def _ar_populate_soce(self, sheet, rd):
        """Populate Statement of Changes in Equity from Audit Report data."""
        soce_rows = rd.get('soce_rows') or []
        if not soce_rows:
            return

        fmt = '#,##0.00'
        row_start = 7
        template_last_row = 19
        template_capacity = template_last_row - row_start + 1

        # Expand template when Audit Report produces more SOCE rows than static
        # placeholders; this prevents silently dropping movement lines.
        if len(soce_rows) > template_capacity:
            insert_count = len(soce_rows) - template_capacity
            sheet.insert_rows(template_last_row + 1, insert_count)
            for offset in range(insert_count):
                self._copy_row_style(sheet, template_last_row, template_last_row + 1 + offset, max_col=6)

        # Map SOCE row keys to Excel columns.
        col_map = {
            'share_capital': 'B',
            'owner_current_account': 'C',
            'retained_earnings': 'D',
            'statutory_reserves': 'E',
            'total_equity': 'F',
        }

        max_data_row = row_start + len(soce_rows) - 1
        for index, soce_row in enumerate(soce_rows):
            row_idx = row_start + index
            sheet.row_dimensions[row_idx].hidden = False
            self._ar_write_soce_row(sheet, row_idx, soce_row, col_map, fmt)

        clear_until = max(template_last_row, max_data_row)
        for row_idx in range(max_data_row + 1, clear_until + 1):
            sheet.row_dimensions[row_idx].hidden = True
            sheet.cell(row=row_idx, column=1, value=None)
            for col_letter in ('B', 'C', 'D', 'E', 'F'):
                sheet[f'{col_letter}{row_idx}'] = None

    def _ar_write_soce_row(self, sheet, row_idx, soce_row, col_map, fmt):
        """Write a single SOCE data row into the given *row_idx*."""
        label = soce_row.get('label')
        if label:
            sheet.cell(row=row_idx, column=1, value=label)
        for key, col_letter in col_map.items():
            value = soce_row.get(key)
            if value is None:
                sheet[f'{col_letter}{row_idx}'] = None
            else:
                self._ar_set_value(sheet, f'{col_letter}{row_idx}', value, fmt)

    # -- SOCF population ---------------------------------------------------

    def _ar_populate_socf(self, sheet, rd):
        """Populate Statement of Cash Flows from Audit Report data."""
        show_prior = bool(rd.get('show_prior_year'))
        fmt = '#,##0.00'
        comparative_period_word = rd.get('comparative_period_word') or 'year'

        def has_value(current, prior=0.0):
            return bool(current) or (show_prior and bool(prior))

        def row(label=None, keys=None, **kwargs):
            row_idx = None
            if label:
                row_idx = self._find_row_by_label(sheet, label, **kwargs)
            if row_idx is None and keys:
                row_idx = self._find_row_by_keys(sheet, keys, **kwargs)
            return row_idx

        def write_row(row_idx, *, label=None, current=None, prior=None, visible=True):
            if not row_idx:
                return
            sheet.row_dimensions[row_idx].hidden = not visible
            if label is not None:
                sheet.cell(row=row_idx, column=1, value=label)
            if not visible:
                sheet[f'B{row_idx}'] = None
                sheet[f'C{row_idx}'] = None
                return
            if current is None:
                sheet[f'B{row_idx}'] = None
            else:
                self._ar_set_value(sheet, f'B{row_idx}', current, fmt)
            if show_prior:
                if prior is None:
                    sheet[f'C{row_idx}'] = None
                else:
                    self._ar_set_value(sheet, f'C{row_idx}', prior, fmt)
            else:
                sheet[f'C{row_idx}'] = None

        current_np = rd.get('cashflow_net_profit_amount', rd.get('net_profit_before_tax', 0.0))
        prior_np = rd.get('cashflow_prev_net_profit_amount', rd.get('prev_net_profit_before_tax', 0.0))
        dep_cur = rd.get('current_depreciation_total', 0.0)
        dep_prior = rd.get('prior_depreciation_total', 0.0)
        eosb_adj_cur = rd.get('end_service_benefits_adjustment', 0.0)
        eosb_adj_prior = rd.get('prior_end_service_benefits_adjustment', 0.0)
        operating_before_wc_cur = rd.get('operating_cashflow_before_working_capital', 0.0)
        operating_before_wc_prior = rd.get('prior_operating_cashflow_before_working_capital', 0.0)
        current_assets_change_cur = rd.get('change_in_current_assets', 0.0)
        current_assets_change_prior = rd.get('prior_change_in_current_assets', 0.0)
        current_liabilities_change_cur = rd.get('change_in_current_liabilities', 0.0)
        current_liabilities_change_prior = rd.get('prior_change_in_current_liabilities', 0.0)
        corporate_tax_paid_cur = rd.get('corporate_tax_paid', 0.0)
        corporate_tax_paid_prior = rd.get('prior_corporate_tax_paid', 0.0)
        eosb_paid_cur = rd.get('end_service_benefits_paid', 0.0)
        eosb_paid_prior = rd.get('prior_end_service_benefits_paid', 0.0)
        net_operations_cur = rd.get('net_cash_generated_from_operations', 0.0)
        net_operations_prior = rd.get('prior_net_cash_generated_from_operations', 0.0)
        property_cur = rd.get('current_property', 0.0)
        property_prior = rd.get('prior_property', 0.0)
        net_investing_cur = rd.get('net_cash_generated_from_investing_activities', 0.0)
        net_investing_prior = rd.get('prior_net_cash_generated_from_investing_activities', 0.0)
        paid_up_capital_cur = rd.get('paid_up_capital', 0.0)
        paid_up_capital_prior = rd.get('prior_paid_up_capital', 0.0)
        dividend_paid_cur = rd.get('dividend_paid', 0.0)
        dividend_paid_prior = rd.get('prior_dividend_paid', 0.0)
        owner_ca_cur = rd.get('owner_current_account', 0.0)
        owner_ca_prior = rd.get('prior_owner_current_account', 0.0)
        show_owner_ca_row = bool(rd.get('show_owner_current_account_cashflow_row'))
        related_party_loan_cur = rd.get('related_party_loan', 0.0)
        related_party_loan_prior = rd.get('prior_related_party_loan', 0.0)
        show_related_party_loan_row = bool(rd.get('show_related_party_loan_cashflow_row'))
        security_deposit_cur = rd.get('security_deposit', 0.0)
        security_deposit_prior = rd.get('prior_security_deposit', 0.0)
        show_security_deposit_row = bool(rd.get('show_security_deposit_cashflow_row'))
        net_financing_cur = rd.get('net_cash_generated_from_financing_activities', 0.0)
        net_financing_prior = rd.get('prior_net_cash_generated_from_financing_activities', 0.0)
        net_cash_change_cur = rd.get('net_cash_and_cash_equivalents', 0.0)
        net_cash_change_prior = rd.get('prior_net_cash_and_cash_equivalents', 0.0)
        cash_beginning_cur = rd.get('cash_beginning_year', 0.0)
        cash_beginning_prior = rd.get('prior_cash_beginning_year', 0.0)
        cash_end_cur = rd.get('cash_end_of_year', 0.0)
        cash_end_prior = rd.get('prior_cash_end_of_year', 0.0)

        # Insert dynamic adjustment rows between net profit and operating-before-WC.
        r_np = row(keys=['Net profit for the year', 'Net profit for the period']) or 9
        r_ocf = row(
            keys=[
                'Operating cash flows before changes in working capital',
                'Operating cash flows before working capital changes',
            ]
        ) or 10
        adjustment_lines = []
        if has_value(dep_cur, dep_prior) or has_value(eosb_adj_cur, eosb_adj_prior):
            adjustment_lines.append({'label': 'Adjustments for:', 'current': None, 'prev': None})
        if has_value(dep_cur, dep_prior):
            adjustment_lines.append({'label': 'Depreciation', 'current': dep_cur, 'prev': dep_prior})
        if has_value(eosb_adj_cur, eosb_adj_prior):
            adjustment_lines.append({
                'label': 'End of service benefits',
                'current': eosb_adj_cur,
                'prev': eosb_adj_prior,
            })
        adjustment_rows = self._ar_ensure_rows_between(sheet, r_np, r_ocf, len(adjustment_lines), max_col=3)
        self._ar_write_dynamic_rows(sheet, adjustment_rows, adjustment_lines, show_prior, fmt)

        # Re-resolve shifted anchors after row insertions.
        r_op_hdr = row(label='Cash flows from operating activities') or 8
        r_np = row(keys=['Net profit for the year', 'Net profit for the period']) or 9
        r_ocf = row(
            keys=[
                'Operating cash flows before changes in working capital',
                'Operating cash flows before working capital changes',
            ]
        ) or 10
        r_wc_hdr = row(label='Changes in working capital') or 11
        r_ca_chg = row(keys=['(Increase) / decrease in current assets', 'Increase / decrease in current assets']) or 12
        r_cl_chg = row(
            keys=['Increase / (decrease) in current liabilities', 'Increase / decrease in current liabilities']
        ) or 13
        r_ncfo = row(keys=['Net cash generated from operations', 'Net cash (used in) operations']) or 14

        # Insert optional operating rows before net operations.
        optional_operating_lines = []
        if has_value(corporate_tax_paid_cur, corporate_tax_paid_prior):
            optional_operating_lines.append({
                'label': 'Corporate tax paid',
                'current': corporate_tax_paid_cur,
                'prev': corporate_tax_paid_prior,
            })
        if has_value(eosb_paid_cur, eosb_paid_prior):
            optional_operating_lines.append({
                'label': 'End of service benefits paid',
                'current': eosb_paid_cur,
                'prev': eosb_paid_prior,
            })
        optional_rows = self._ar_ensure_rows_between(sheet, r_cl_chg, r_ncfo, len(optional_operating_lines), max_col=3)
        self._ar_write_dynamic_rows(sheet, optional_rows, optional_operating_lines, show_prior, fmt)

        # Resolve all remaining anchors after insertions.
        r_ncfo = row(keys=['Net cash generated from operations', 'Net cash (used in) operations']) or 14
        r_invest_hdr = row(label='Cash flows from investing activities', keys=['Cash flows from investing activities']) or 15
        r_ncfi = row(
            keys=['Net cash generated from investing activities', 'Net cash (used in) investing activities'],
            start_row=r_invest_hdr,
        ) or 18
        r_fin_hdr = row(label='Cash flows from financing activities', keys=['Cash flows from financing activities']) or 19
        r_ncff = row(
            keys=['Net cash generated from financing activities', 'Net cash (used in) financing activities'],
            start_row=r_fin_hdr,
        ) or 24

        # Operating section values and labels.
        if show_prior:
            current_is_loss = (current_np or 0.0) < 0.0
            prior_is_loss = (prior_np or 0.0) < 0.0
            if current_is_loss and prior_is_loss:
                net_profit_label = f'Net (loss) for the {comparative_period_word}'
            elif (not current_is_loss) and (not prior_is_loss):
                net_profit_label = f'Net profit for the {comparative_period_word}'
            elif (not current_is_loss) and prior_is_loss:
                net_profit_label = f'Net profit / (loss) for the {comparative_period_word}'
            else:
                net_profit_label = f'Net (loss) / profit for the {comparative_period_word}'
        else:
            net_profit_label = 'Net (loss) for the period' if (current_np or 0.0) < 0.0 else 'Net profit for the period'

        write_row(r_op_hdr, label='Cash flows from operating activities', current=None, prior=None, visible=True)
        write_row(r_np, label=net_profit_label, current=current_np, prior=prior_np, visible=True)
        write_row(
            r_ocf,
            label='Operating cash flows before changes in working capital',
            current=operating_before_wc_cur,
            prior=operating_before_wc_prior,
            visible=True,
        )
        write_row(r_wc_hdr, label='Changes in working capital', current=None, prior=None, visible=True)
        write_row(
            r_ca_chg,
            label='(Increase) / decrease in current assets',
            current=current_assets_change_cur,
            prior=current_assets_change_prior,
            visible=True,
        )
        write_row(
            r_cl_chg,
            label='Increase / (decrease) in current liabilities',
            current=current_liabilities_change_cur,
            prior=current_liabilities_change_prior,
            visible=True,
        )
        net_ops_label = 'Net cash (used in) operations' if has_value(
            (net_operations_cur or 0.0) < 0.0,
            (net_operations_prior or 0.0) < 0.0,
        ) else 'Net cash generated from operations'
        write_row(r_ncfo, label=net_ops_label, current=net_operations_cur, prior=net_operations_prior, visible=True)

        # Investing section print/hide.
        investing_lines = []
        if has_value(property_cur, property_prior):
            investing_lines.append({
                'label': 'Property, plant and equipment',
                'current': property_cur,
                'prev': property_prior,
            })
        if show_security_deposit_row:
            investing_lines.append({
                'label': 'Security deposit',
                'current': security_deposit_cur,
                'prev': security_deposit_prior,
            })
        show_investing = bool(investing_lines) or has_value(net_investing_cur, net_investing_prior)
        investing_rows = self._ar_ensure_rows_between(
            sheet,
            r_invest_hdr,
            r_ncfi,
            len(investing_lines),
            max_col=3,
        )
        self._ar_write_dynamic_rows(sheet, investing_rows, investing_lines, show_prior, fmt)
        write_row(r_invest_hdr, label='Cash flows from investing activities', current=None, prior=None, visible=show_investing)
        investing_label = (
            'Net cash (used in) investing activities'
            if has_value((net_investing_cur or 0.0) < 0.0, (net_investing_prior or 0.0) < 0.0)
            else 'Net cash generated from investing activities'
        )
        write_row(
            r_ncfi,
            label=investing_label,
            current=net_investing_cur,
            prior=net_investing_prior,
            visible=show_investing and has_value(net_investing_cur, net_investing_prior),
        )

        financing_lines = []
        if has_value(paid_up_capital_cur, paid_up_capital_prior):
            financing_lines.append({
                'label': 'Paid up capital',
                'current': paid_up_capital_cur,
                'prev': paid_up_capital_prior,
            })
        if has_value(dividend_paid_cur, dividend_paid_prior):
            financing_lines.append({
                'label': 'Dividend paid',
                'current': dividend_paid_cur,
                'prev': dividend_paid_prior,
            })
        if show_owner_ca_row:
            financing_lines.append({
                'label': 'Owner current account',
                'current': owner_ca_cur,
                'prev': owner_ca_prior,
            })
        if show_related_party_loan_row:
            financing_lines.append({
                'label': 'Loan from related party',
                'current': related_party_loan_cur,
                'prev': related_party_loan_prior,
            })
        financing_rows = self._ar_ensure_rows_between(
            sheet,
            r_fin_hdr,
            r_ncff,
            len(financing_lines),
            max_col=3,
        )
        self._ar_write_dynamic_rows(sheet, financing_rows, financing_lines, show_prior, fmt)

        r_ncff = row(
            keys=['Net cash generated from financing activities', 'Net cash (used in) financing activities'],
            start_row=r_fin_hdr,
        ) or 24
        r_net = row(
            keys=[
                'Net increase in cash and cash equivalents',
                'Net increase / (decrease) in cash and cash equivalents',
            ]
        ) or 25
        r_beg = row(
            keys=['Cash and cash equivalents, beginning of the period', 'Cash and cash equivalents at the beginning']
        ) or 26
        r_end = row(
            keys=['Cash and cash equivalents, end of the period', 'Cash and cash equivalents at the end']
        ) or 27

        # Financing section print/hide.
        show_financing = (
            has_value(paid_up_capital_cur, paid_up_capital_prior)
            or has_value(dividend_paid_cur, dividend_paid_prior)
            or show_owner_ca_row
            or show_related_party_loan_row
            or has_value(net_financing_cur, net_financing_prior)
        )
        write_row(r_fin_hdr, label='Cash flows from financing activities', current=None, prior=None, visible=show_financing)
        financing_label = (
            'Net cash (used in) financing activities'
            if has_value((net_financing_cur or 0.0) < 0.0, (net_financing_prior or 0.0) < 0.0)
            else 'Net cash generated from financing activities'
        )
        write_row(
            r_ncff,
            label=financing_label,
            current=net_financing_cur,
            prior=net_financing_prior,
            visible=show_financing,
        )

        # Closing rows.
        write_row(
            r_net,
            label='Net increase in cash and cash equivalents',
            current=net_cash_change_cur,
            prior=net_cash_change_prior,
            visible=True,
        )
        write_row(
            r_beg,
            label='Cash and cash equivalents, beginning of the period',
            current=cash_beginning_cur,
            prior=cash_beginning_prior,
            visible=True,
        )
        write_row(
            r_end,
            label='Cash and cash equivalents, end of the period',
            current=cash_end_cur,
            prior=cash_end_prior,
            visible=has_value(cash_end_cur, cash_end_prior),
        )

        if not show_prior:
            sheet['C6'] = None
            for row_idx in range(7, sheet.max_row + 1):
                sheet[f'C{row_idx}'] = None

    def _get_unique_sheet_name(self, base_name, used_sheet_names):
        max_len = 31
        name = (base_name or 'Sheet').strip()[:max_len]
        if name not in used_sheet_names:
            used_sheet_names.add(name)
            return name

        counter = 1
        while True:
            suffix = f" ({counter})"
            candidate = f"{name[:max_len - len(suffix)]}{suffix}"
            if candidate not in used_sheet_names:
                used_sheet_names.add(candidate)
                return candidate
            counter += 1
